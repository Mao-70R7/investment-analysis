from __future__ import annotations

import json
import math
import re
import sqlite3
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl

from benchmark_comparison_pool import build_comparison_pool


ROOT = next(parent for parent in Path(__file__).resolve().parents if (parent / "AGENTS.md").is_file())
DEFAULT_CATALOG_PATH = ROOT / "业务基线" / "基准.xlsx"
ASSET_TABLE_NAME = "策略基准资产配置"
CATALOG_SHEET_NAME = "基准指数全量表"
OTHER_BUCKET = "其他"
CANONICAL_ASSET_MAJORS = ["权益", "债券", "现金", "商品", "另类"]
CANONICAL_ASSET_CATEGORIES = ["A股", "港股", "海外权益", "债券", "现金", "商品", "另类"]

# These indices represent changing portfolios of funds or multiple asset classes. Their
# names do not provide a stable equity weight, so they must not fall through to the
# generic "中证...指数 = A股" rule.
NON_CLASSIFIABLE_MULTI_ASSET_INDEX_TERMS = [
    "中证开放式基金指数",
    "中证基金指数",
    "中证混合型基金指数",
    "中证普通混合型基金指数",
    "中证灵活配置混合型基金指数",
    "中证工银财富动态配置基金指数",
    "中证目标日期2035指数",
]

DYNAMIC_BENCHMARK_PATTERN = re.compile(
    r"下滑曲线|权益配置中枢比例|目标日期|时间段|年份|"
    r"(?<![A-Za-z])X(?![A-Za-z])|\(\s*1\s*-\s*X\s*\)|"
    r"(?<![A-Za-z])[ABISYZ](?=\s*(?:%|\*|×|\)|$))",
    flags=re.IGNORECASE,
)


def clean_text(value: Any, default: str | None = None) -> str | None:
    if value is None:
        return default
    text = str(value).strip()
    return text if text else default


def as_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        if isinstance(value, float) and math.isnan(value):
            return None
        return float(value)
    text = str(value).strip().replace("%", "").replace("％", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def round_or_none(value: Any, digits: int = 4) -> float | None:
    number = as_float(value)
    if number is None:
        return None
    return round(number, digits)


def compact_text(text: Any) -> str:
    raw = clean_text(text, "") or ""
    return re.sub(r"[\s（）()\[\]【】{}<>《》\-—_:/：,，;；]+", "", raw).upper()


def contains_non_classifiable_multi_asset_index(text: Any) -> bool:
    compact = compact_text(text)
    return any(compact_text(term) in compact for term in NON_CLASSIFIABLE_MULTI_ASSET_INDEX_TERMS)


def is_static_benchmark_formula(text: Any) -> bool:
    raw = clean_text(text, "") or ""
    years = set(re.findall(r"20\d{2}", raw))
    return bool(raw) and len(years) < 2 and not DYNAMIC_BENCHMARK_PATTERN.search(raw)


def split_aliases(value: Any) -> list[str]:
    text = clean_text(value, "") or ""
    parts = re.split(r"[、,，;；\n\r]+", text)
    return [part.strip() for part in parts if part and part.strip()]


def normalize_asset_major(major: str, category: str) -> str:
    """Use category-level truth for broad asset class when the source catalog is too coarse."""
    if category in {"A股", "港股", "海外权益"}:
        return "权益"
    if category == "债券":
        return "债券"
    if category == "现金":
        return "现金"
    if category == "商品":
        return "商品"
    if category == "另类":
        return "另类"
    return major


EXTRA_ALIASES_BY_INDEX_NAME: dict[str, list[str]] = {
    "沪深300指数": ["沪深300", "CSI300", "000300"],
    "沪深300全收益指数": ["沪深300全收益指数", "沪深300全收益", "沪深全收益300指数"],
    "上证综合指数": ["上证指数", "上证综指", "上证综合"],
    "上证国债指数": ["上证国债", "国债指数", "000012"],
    "中证全债指数": ["中证全债", "中证全债指数收益率", "H11001"],
    "中证综合债指数": ["中证综合债券指数", "中证综合债券", "中证综合债", "中证综合债指数收益率"],
    "中证偏股型基金指数": ["中证偏股基金指数", "中证偏股基金指数收益率", "中证偏股基金", "中证偏股混合基金指数", "中证偏股混合基金", "偏股基金指数", "偏股基金"],
    "中证纯债债券型基金指数": ["中证纯债债券型基金指数", "中证纯债债基指数", "中证纯债债基", "中证纯债基金指数", "中证纯债基金", "中证纯债", "纯债债基", "纯债债基指数", "930609.CSI"],
    "中证普通债券型基金指数": ["普通债券基金指数", "普通债券型基金", "930610.CSI"],
    "中证债券型基金指数": ["债券型基金指数", "H11023"],
    "中证货币基金指数": ["中证货币型基金指数", "中证货币市场基金指数", "中证货币指数", "货币基金指数", "货币市场基金指数", "货币基金", "货币市场基金", "H11025.CSI", "H11025"],
    "中债-综合全价(总值)指数": ["中债综合全价", "中债-综合全价", "中债综合全价指数", "中债综合全价总值指数", "中证综合全价", "CBA00203.CS", "彭博巴克莱全球综合指数", "彭博巴克莱全球综合"],
    "中债-综合财富(总值)指数": ["中债综合财富", "中债-综合财富", "中债综合指数", "中债综合", "CBA00201.CS"],
    "中债-新综合全价(总值)指数": ["中债新综合全价", "中债-新综合全价", "中债新综合全价指数", "中债新综合全价总值指数", "CBA00603.CS"],
    "中债-新综合财富(总值)指数": ["中债新综合财富", "中债-新综合财富", "中债-新综合指数", "中债新综合指数", "CBA00601.CS"],
    "中债-总全价(总值)指数": ["中债总全价", "中债-总全价", "中债-总指数全价", "中债总指数全价", "CBA00103.CS"],
    "中债-总财富(总值)指数": ["中债总财富", "中债财富总指数", "中债-总财富", "CBA00101.CS"],
    "恒生指数": ["香港恒生指数", "恒指", "HSI", "HSI.HI"],
    "MSCI World Index": ["MSCI全球指数", "MSCI全球", "MSCI世界指数", "MSCI发达市场指数", "MSCI发达市场", "MSCI新兴市场指数", "MSCI新兴市场", "990100.MI", "891800.MI"],
    "S&P 500 Index": ["标准普尔500", "标普500", "标普 500", "S&P500", "S&P 500", "SP500", "SPX", "SPX.GI"],
    "Nasdaq-100 Index": ["纳斯达克100", "纳斯达克 100", "纳指100", "纳指 100", "NASDAQ100", "NDX", "NDX.GI"],
    "上海黄金交易所Au99.99": ["上海黄金Au99.99", "上海金现货", "上海黄金9999", "黄金9999", "AU99.99", "AU9999", "SHAU.SGE"],
    "中证内地消费主题指数": ["中证内地消费指数", "中证消费", "消费主题指数", "000942", "000942.CSI"],
    "中证医药卫生指数": ["中证医药", "中证医药指数", "000933", "000933.SH", "000933.CSI"],
    "中证新能源指数": ["新能源指数", "000941", "000941.SH", "000941.CSI"],
    "中证科技传媒通信150指数": ["中证TMT指数", "中证TMT", "000998", "000998.CSI"],
    "中证智能制造主题指数": ["中证装备产业指数", "中证装备产业", "H11054", "H11054.CSI"],
    "中证商品期货综合指数": ["中证商品CFCI综合指数", "中证商品CFCI综合", "中证商品CIFI指数", "中证商品CIFI"],
    "现金/存款": ["活期存款利率", "银行人民币活期存款利率", "1年定期存款利率", "一年定期存款利率"],
    "上期所有色金属期货": ["上期所有色金属期货", "IMCI.SHF"],
    "大商所豆粕期货": ["大商所豆粕期货", "DCESMFI.DCE"],
    "郑商所能源化工期货": ["郑商所能源化工期货", "000201.CZC"],
}


@dataclass
class BenchmarkCatalogEntry:
    asset_major: str
    asset_category: str
    index_name: str
    aliases: list[str] = field(default_factory=list)
    penetration: float | None = None
    code: str = ""

    def all_aliases(self) -> list[str]:
        values: list[str] = [self.index_name, *self.aliases]
        values.extend(EXTRA_ALIASES_BY_INDEX_NAME.get(self.index_name, []))
        if self.index_name.endswith("指数") and len(self.index_name) > 2:
            values.append(self.index_name[:-2])
        deduped: list[str] = []
        seen: set[str] = set()
        for value in values:
            text = clean_text(value, "") or ""
            if not text:
                continue
            key = compact_text(text)
            if not key or key in seen:
                continue
            seen.add(key)
            deduped.append(text)
        return deduped


@dataclass
class BenchmarkCatalog:
    entries: list[BenchmarkCatalogEntry]
    asset_majors: list[str]
    asset_categories: list[str]

    @property
    def asset_major_fields(self) -> list[str]:
        return [f"基准资产大类-{name}" for name in [*self.asset_majors, OTHER_BUCKET]]

    @property
    def asset_category_fields(self) -> list[str]:
        return [f"基准资产类别-{name}" for name in [*self.asset_categories, OTHER_BUCKET]]


def load_benchmark_catalog(path: Path = DEFAULT_CATALOG_PATH) -> BenchmarkCatalog:
    if not path.exists():
        raise FileNotFoundError(f"基准分类表不存在：{path}")
    workbook = openpyxl.load_workbook(path, data_only=True)
    if CATALOG_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"基准分类表缺少工作表：{CATALOG_SHEET_NAME}")
    sheet = workbook[CATALOG_SHEET_NAME]
    headers = [clean_text(cell.value, "") for cell in sheet[1]]
    rows: list[BenchmarkCatalogEntry] = []
    asset_majors: list[str] = list(CANONICAL_ASSET_MAJORS)
    asset_categories: list[str] = list(CANONICAL_ASSET_CATEGORIES)
    for values in sheet.iter_rows(min_row=2, values_only=True):
        item = dict(zip(headers, values))
        major = clean_text(item.get("资产大类"), "")
        category = clean_text(item.get("资产类别"), "")
        index_name = clean_text(item.get("指数名称"), "")
        if not major or not category or not index_name:
            continue
        major = normalize_asset_major(str(major), str(category))
        aliases = split_aliases(item.get("基准原始表述方式"))
        entry = BenchmarkCatalogEntry(
            asset_major=str(major),
            asset_category=str(category),
            index_name=str(index_name),
            aliases=aliases,
            penetration=as_float(item.get("类内渗透率")),
            code=f"XLSX:{compact_text(index_name)}",
        )
        rows.append(entry)
        if major not in asset_majors:
            asset_majors.append(str(major))
        if category not in asset_categories:
            asset_categories.append(str(category))
    return BenchmarkCatalog(rows, asset_majors, asset_categories)


def match_catalog_entry(text: str, catalog: BenchmarkCatalog) -> BenchmarkCatalogEntry | None:
    compact = compact_text(text)
    if not compact:
        return None
    if contains_non_classifiable_multi_asset_index(text):
        return None
    if "股票" in compact and any(term in compact for term in ["黄金", "商品", "原油", "能源", "有色金属"]):
        return BenchmarkCatalogEntry("权益", "A股", "资源产业股票指数", [], code=f"SEMANTIC:RESOURCE_EQUITY:{compact[:40]}")
    if any(term in compact for term in ["货币基金", "货币型基金", "货币市场", "活期存款", "定期存款", "存款利率"]):
        return BenchmarkCatalogEntry("现金", "现金", "现金/存款", [], code=f"SEMANTIC:CASH:{compact[:40]}")
    if any(term in compact for term in ["中债", "债券", "国债", "全债", "短债", "综合债", "纯债", "转债", "信用债", "同业存单", "美元债"]):
        return BenchmarkCatalogEntry("债券", "债券", "债券指数", [], code=f"SEMANTIC:BOND:{compact[:40]}")
    if any(term in compact for term in ["CME中国商品", "中证商品", "商品CIFI", "商品CFI", "商品CFCI", "商品期货", "商品综合指数", "AU9999", "原油", "WTI", "BRENT", "GSCI"]):
        return BenchmarkCatalogEntry("商品", "商品", "商品指数", [], code=f"SEMANTIC:COMMODITY:{compact[:40]}")
    best: tuple[int, float, BenchmarkCatalogEntry] | None = None
    for entry in catalog.entries:
        for alias in entry.all_aliases():
            alias_key = compact_text(alias)
            if not alias_key:
                continue
            if alias_key in compact:
                score = (len(alias_key), float(entry.penetration or 0.0), entry)
                if best is None or score[:2] > best[:2]:
                    best = score
    return best[2] if best else None


def infer_generic_index_entry(text: str) -> BenchmarkCatalogEntry | None:
    """Infer broad asset class from benchmark text when the index is absent from the catalog."""
    compact = compact_text(text)
    if not compact or "未披露" in compact or "暂无" in compact:
        return None
    if contains_non_classifiable_multi_asset_index(text):
        return None
    def entry(major: str, category: str, name: str, code: str) -> BenchmarkCatalogEntry:
        return BenchmarkCatalogEntry(major, category, name, [], code=f"GENERIC:{code}:{compact[:40]}")

    if re.fullmatch(r"(?:业绩比较基准)?年化收益率\d+(?:\.\d+)?%?", compact):
        return None
    if "股票" in compact and any(term in compact for term in ["黄金", "商品", "有色金属", "原油", "能源"]):
        category = "港股" if any(term in compact for term in ["港股", "恒生"]) else "A股"
        return entry("权益", category, "资源产业股票指数", "RESOURCE_EQUITY")
    if any(
        term in compact
        for term in [
            "存款",
            "活期",
            "定期",
            "定存",
            "央行定存",
            "通知存款",
            "银行定期",
            "同业存款",
            "同业拆借",
            "银行同业拆借",
            "SHIBOR",
            "HIBOR",
            "LIBOR",
            "SOFR",
            "FR007",
            "DR007",
            "货币市场",
            "货币基金",
        ]
    ):
        return entry("现金", "现金", "现金/存款", "CASH")
    if any(
        term in compact
        for term in [
            "中债",
            "债券",
            "国债",
            "金融债",
            "企业债",
            "公司债",
            "信用债",
            "中诚信",
            "利率债",
            "政策性金融债",
            "同业存单",
            "短融",
            "短期融资券",
            "票据",
            "CFETS",
            "彭博政策性银行债",
            "中国债券",
            "中信全债",
            "全债",
            "短债",
            "综合债",
            "纯债",
            "转债",
            "IBOXX亚债",
            "IBOXX",
            "中资美元债",
            "美元债",
            "上清所",
            "上海清算所",
            "农发债",
            "转债",
            "可转债",
            "JPMORGANASIANCREDIT",
            "ASIANCREDIT",
            "USDCREDIT",
            "BOFA",
            "MERRILL",
            "BROADMARKET",
            "TBILL",
            "T-BILL",
            "GOVERNMENTBOND",
            "GLOBALBOND",
            "美国总体市场",
        ]
    ):
        return entry("债券", "债券", "债券指数", "BOND")
    if "REIT" in compact or "不动产" in compact:
        return entry("另类", "另类", "REITs指数", "ALT")
    if any(term in compact for term in ["黄金", "白银", "商品期货", "商品指数", "商品综合指数", "CME中国商品", "期货指数", "有色金属期货", "能源化工", "豆粕期货", "AU9999", "WTI", "BRENT", "原油", "GSCI", "COMMODITY"]):
        return entry("商品", "商品", "商品指数", "COMMODITY")
    if any(term in compact for term in ["港股", "香港", "恒生", "HK", "H股"]):
        return entry("权益", "港股", "港股权益指数", "HK_EQUITY")
    if "MSCI" in compact and any(term in compact for term in ["中国A股", "A股在岸", "CHINAA"]):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    if any(
        term in compact
        for term in [
            "纳斯达克",
            "标普",
            "标准普尔",
            "MSCI",
            "富时",
            "罗素",
            "日经",
            "海外",
            "美国",
            "全球",
            "德国",
            "法国",
            "越南",
            "VN30",
            "VIETNAM",
            "道琼斯",
            "IBOVESPA",
            "伊博维斯帕",
            "摩根斯坦利新兴市场股票",
            "新交所",
            "泛东南亚",
            "新兴亚洲",
            "印度",
            "INDIA",
            "ETP",
            "东证指数",
            "东京证券交易所股价总指数",
            "TOPIX",
            "费城半导体",
            "彭博全球新能源汽车",
            "PHLXSEMICONDUCTOR",
            "GLOBALX",
            "BLOOMBERG",
            "ELECTRICVEHICLES",
            "SEMICONDUCTOR",
        ]
    ):
        if any(term in compact for term in ["AGGREGATE", "债", "CREDIT", "TREASURY", "BOFA", "MERRILL", "BROADMARKET", "TBILL", "T-BILL", "GOVERNMENTBOND"]):
            return entry("债券", "债券", "债券指数", "BOND")
        return entry("权益", "海外权益", "海外权益指数", "OVERSEAS_EQUITY")
    if "中信" in compact and any(
        term in compact
        for term in [
            "行业",
            "风格",
            "消费",
            "成长",
            "汽车",
            "计算机",
            "农林牧渔",
            "食品饮料",
            "医药",
            "中药",
            "生产",
            "电子",
            "通信",
            "传媒",
        ]
    ):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    if any(term in compact for term in ["360互联网", "180治理", "专精特新小巨人", "中国制造2025", "内地消费", "消费主题"]):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    if "偏股" in compact and "基金指数" in compact:
        return entry("权益", "A股", "偏股基金指数", "PARTIAL_EQUITY_FUND")
    if "指数" in compact and any(term in compact for term in ["主题", "行业", "专精特新", "小巨人"]):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    if any(term in compact for term in ["深证成指", "上证A指", "上证综指", "上证综合", "天相280"]):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    if "中证" in compact and any(
        term in compact
        for term in [
            "科技",
            "消费",
            "医药",
            "新能源",
            "高端制造",
            "智能制造",
            "信息",
            "传媒",
            "通信",
            "电子",
            "半导体",
            "红利",
            "成长",
            "价值",
            "行业",
            "主题",
            "股票",
            "A股",
            "800",
            "500",
            "1000",
            "A500",
        ]
    ):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    if "指数" in compact and any(
        term in compact
        for term in [
            "中证",
            "国证",
            "上证",
            "深证",
            "沪深",
            "创业板",
            "科创",
            "北证",
            "申万",
            "申银万国",
            "天相中盘",
            "天相小盘",
            "天相280",
            "华证",
            "新华",
            "巨潮",
            "大数据100",
            "大数据300",
            "360互联网",
            "中华交易服务",
            "天相小市值",
            "中小企业100",
            "中小企业综合",
            "中创400",
            "央视财经50",
            "粤港澳大湾区创新100",
            "道琼斯中国88",
        ]
    ):
        return entry("权益", "A股", "A股权益指数", "A_SHARE")
    return None


def is_fixed_return_part(part: str) -> bool:
    """Return true for additive absolute-return clauses that do not define an asset weight."""
    compact = compact_text(part)
    if not compact:
        return False
    if any(
        term in compact
        for term in [
            "中证",
            "沪深",
            "上证",
            "深证",
            "国证",
            "创业板",
            "科创",
            "北证",
            "内地消费",
            "恒生",
            "标普",
            "纳斯达克",
            "富时",
            "罗素",
        ]
    ):
        return False
    if "*" in compact:
        return False
    if any(term in compact for term in ["指数", "存款", "活期", "定期", "同业", "债", "股票", "基金", "货币市场", "原油", "黄金", "商品", "MSCI", "WTI", "BRENT", "ETP"]):
        return False
    if re.fullmatch(r"(?:年化)?收益率\d+(?:\.\d+)?%?", compact):
        return True
    if re.fullmatch(r"(?:固定)?业绩(?:比较)?基准\d+(?:\.\d+)?%?", compact):
        return True
    if re.fullmatch(r"\d+(?:\.\d+)?%(?:指年收益率|单利年化|年化)?", compact):
        return True
    if re.match(r"^\d+(?:\.\d+)?%", compact):
        return True
    if any(term in compact for term in ["指年收益率", "单利年化", "固定业绩基准", "目标收益"]):
        return bool(re.search(r"\d+(?:\.\d+)?%", compact))
    return False


def split_formula_parts(text: str) -> list[str]:
    parts: list[str] = []
    start = 0
    depth = 0
    for index, char in enumerate(text):
        if char in "(（":
            depth += 1
        elif char in ")）" and depth > 0:
            depth -= 1
        elif char in "+＋" and depth == 0:
            previous = text[index - 1] if index > 0 else ""
            if previous.upper() == "A":
                continue
            part = text[start:index].strip()
            if part:
                parts.append(part)
            start = index + 1
    tail = text[start:].strip()
    if tail:
        parts.append(tail)
    return parts


def parse_cash_average_formula(text: str, catalog: BenchmarkCatalog) -> dict[str, Any] | None:
    compact = compact_text(text)
    if not re.search(r"1\s*/\s*2", text) or "存款" not in compact:
        return None
    if any(term in compact for term in ["股票", "指数", "债券", "中债", "沪深", "中证", "上证", "深证", "港股", "商品", "黄金"]):
        return None
    cash_entry = match_catalog_entry("现金/存款", catalog) or BenchmarkCatalogEntry("现金", "现金", "现金/存款", [], code="GENERIC:CASH")
    return {
        "components": [component_from_entry(cash_entry, 1.0)],
        "missing": [],
        "missing_parts": [],
        "说明": "现金/存款100.00%（存款利率均值公式按现金基准处理）",
    }


def add_or_update_component(parsed: dict[str, dict[str, Any]], entry: BenchmarkCatalogEntry, weight: float) -> None:
    bucket = parsed.setdefault(
        entry.code,
        {
            "code": entry.code,
            "name": entry.index_name,
            "weight": 0.0,
            "资产大类": entry.asset_major,
            "资产类别": entry.asset_category,
        },
    )
    bucket["weight"] += weight


def normalize_formula_text(text: str) -> str:
    normalized = (
        re.sub(r"_x000D_", "\n", text, flags=re.IGNORECASE)
        .replace("＝", "=")
        .replace("×", "*")
        .replace("＊", "*")
        .replace("＋", "+")
        .replace("（", "(")
        .replace("）", ")")
        .replace("％", "%")
        .replace("|", "\n")
    )
    normalized = re.sub(r"(?<![A-Za-z])([A-Z])\s*值(?=\s*(?:%|\*|\+|\)|$))", r"\1", normalized, flags=re.IGNORECASE)
    normalized = re.sub(r"(?<=[0-9%）\)])\s*[xX]\s*(?=[A-Za-z0-9\u4e00-\u9fff（(])", "*", normalized)
    normalized = re.sub(r"^业绩(?:比较)?基准\s*[=:：]", "", normalized)
    normalized = normalized.replace("360互联网+大数据100", "360互联网与大数据100")
    normalized = re.sub(r"\s*[+]\s*利差\s*$", "", normalized)
    normalized = re.sub(r"\s*加\s*利差\s*$", "", normalized)
    return normalized


def parse_weight_from_part(part: str) -> float | None:
    # Commas often occur inside an index qualifier before the actual weight,
    # e.g. "Bloomberg ... Index, 按估值汇率折算)*95%".  Cutting at
    # punctuation would discard the disclosed 95% and turn a clear equity
    # benchmark into an unresolved component.
    weight_area = re.split(r"其中|其\s*中|下表|年份|注[:：]|说明[:：]", part, maxsplit=1)[0]
    compact = (
        re.sub(r"\s+", "", weight_area)
        .replace("×", "*")
        .replace("＊", "*")
        .replace("％", "%")
        .replace("#", "%")
        .replace("【", "")
        .replace("】", "")
    )
    percent_matches = list(re.finditer(r"(\d+(?:\.\d+)?)%", compact))
    if percent_matches:
        depth = 0
        depth_at: list[int] = [0] * (len(compact) + 1)
        for index, char in enumerate(compact):
            depth_at[index] = depth
            if char == "(":
                depth += 1
            elif char == ")" and depth > 0:
                depth -= 1
        depth_at[len(compact)] = depth
        top_level = [match for match in percent_matches if depth_at[match.start()] == 0]
        candidates = top_level or percent_matches
        multiplied = []
        for match in candidates:
            before = compact[max(0, match.start() - 1) : match.start()]
            after = compact[match.end() : match.end() + 1]
            if before == "*" or after == "*":
                multiplied.append(match)
        selected = (multiplied or candidates)[-1]
        return float(selected.group(1)) / 100.0
    match = re.search(r"\*(\d+(?:\.\d+)?)(?:$|[^\d.])", compact)
    if not match:
        match = re.match(r"^(\d+(?:\.\d+)?)(?:\*|[^\d.])", compact)
    if match:
        value = float(match.group(1))
        return value / 100.0 if value > 1.0 else value
    return None


def parse_year_weight_table(text: str) -> dict[int, float]:
    weights: dict[int, float] = {}
    for year_text, pct_text in re.findall(r"(20\d{2})\s*(?:及以后)?\s+(\d+(?:\.\d+)?)\s*%", text):
        pct = float(pct_text)
        if 0.0 <= pct <= 100.0:
            weights[int(year_text)] = pct / 100.0
    return dict(sorted(weights.items()))


def year_weight_for_date(weight_by_year: dict[int, float], date_value: str) -> float:
    years = sorted(weight_by_year)
    if not years:
        return 0.0
    current_year = int(str(date_value)[:4])
    selected = years[0]
    for year in years:
        if year <= current_year:
            selected = year
        else:
            break
    return float(weight_by_year.get(selected, 0.0))


def component_from_entry(entry: BenchmarkCatalogEntry, weight: float) -> dict[str, Any]:
    return {
        "code": entry.code,
        "name": entry.index_name,
        "weight": round(weight, 8),
        "资产大类": entry.asset_major,
        "资产类别": entry.asset_category,
    }


def dynamic_table_segments(text: str) -> list[str]:
    plain = re.sub(r"_x000D_", "\n", text, flags=re.IGNORECASE)
    base_segments: list[str] = []
    for line in plain.splitlines():
        line = line.strip()
        if not line:
            continue
        pieces = re.split(r"(?<![-/.至—])(?=20\d{2})", line)
        base_segments.extend(piece.strip() for piece in pieces if piece.strip())
    segments = list(base_segments)
    # PDF/F10 table extraction often wraps a year range across two lines, for
    # example "2026年至" followed by "2028年 46-71 61 39". Keep an adjacent
    # pair as an additional candidate so the disclosed interval remains intact.
    segments.extend(
        f"{base_segments[index]} {base_segments[index + 1]}"
        for index in range(len(base_segments) - 1)
        if (
            (
                any(term in base_segments[index] for term in ["至", "—", "成立", "生效", "设立"])
                or re.search(r"20\d{2}\s*[-/]?\s*$", base_segments[index])
            )
            and not segment_weight_values(base_segments[index])
            and re.search(r"20\d{2}", base_segments[index + 1])
            and segment_weight_values(base_segments[index + 1])
        )
    )
    return segments


def segment_year_interval(segment: str) -> tuple[int, int] | None:
    years = [int(value) for value in re.findall(r"20\d{2}", segment)]
    if not years:
        return None
    if len(years) >= 2:
        return min(years), max(years)
    year = years[0]
    before = segment[: segment.find(str(year))]
    after = segment[segment.find(str(year)) + 4 :]
    if any(term in before for term in ["成立", "生效", "设立"]):
        return 1900, year
    if re.search(r"及以后|以后|起", after):
        return year, 9999
    return year, year


def segment_weight_values(segment: str) -> list[float]:
    value_area = re.sub(r"20\d{2}\s*年(?:\s*\d{1,2}\s*月(?:\s*\d{1,2}\s*日)?)?", " ", segment)
    value_area = re.sub(r"20\d{2}(?:[-./]\d{1,2}){0,2}", " ", value_area)
    values: list[float] = []
    for raw in re.findall(r"(?<![\d.])(\d+(?:\.\d+)?)(?:\s*%)?", value_area):
        value = float(raw)
        if 0.0 <= value <= 100.0:
            values.append(value)
    return values


def disclosed_dynamic_weight_for_year(text: str, year: int) -> tuple[float, str, bool] | None:
    candidates: list[tuple[int, int, float, str, bool]] = []
    compact = compact_text(text)
    x_precedes_bounds = bool(
        re.search(r"(?:下滑曲线中枢值X|中枢值X|X值?).{0,20}(?:下限).{0,20}(?:上限)", compact, flags=re.IGNORECASE)
    )
    for order, segment in enumerate(dynamic_table_segments(text)):
        interval = segment_year_interval(segment)
        if not interval or not (interval[0] <= year <= interval[1]):
            continue
        values = segment_weight_values(segment)
        if not values:
            continue
        used_complement_pair = len(values) >= 2 and abs(values[-2] + values[-1] - 100.0) <= 0.2
        if x_precedes_bounds and len(values) >= 3:
            selected = values[0]
            used_complement_pair = False
        else:
            selected = values[-2] if used_complement_pair else values[-1]
        span = interval[1] - interval[0]
        candidates.append((span, -order, selected, segment, used_complement_pair))
    if not candidates:
        return None
    _, _, value, segment, used_pair = min(candidates, key=lambda item: (item[0], item[1]))
    return value / 100.0, segment, used_pair


def disclosed_dynamic_xy_for_year(text: str, year: int) -> tuple[float, float, str] | None:
    candidates: list[tuple[int, int, float, float, str]] = []
    for order, segment in enumerate(dynamic_table_segments(text)):
        interval = segment_year_interval(segment)
        if not interval or not (interval[0] <= year <= interval[1]):
            continue
        values = segment_weight_values(segment)
        if len(values) < 2:
            continue
        x_value, y_value = values[-2:]
        if x_value + y_value > 100.0001:
            continue
        candidates.append((interval[1] - interval[0], -order, x_value, y_value, segment))
    if not candidates:
        return None
    _, _, x_value, y_value, segment = min(candidates, key=lambda item: (item[0], item[1]))
    return x_value / 100.0, y_value / 100.0, segment


def dynamic_formula_line(text: str) -> str:
    """Return the disclosed formula line rather than a surrounding period heading."""
    lines = [line.strip() for line in re.split(r"_x000D_|[\r\n]", text, flags=re.IGNORECASE) if line.strip()]
    for line in lines:
        if line.count("指数") < 2:
            continue
        if not any(term in compact_text(line) for term in ["债", "国债", "全债", "固定收益"]):
            continue
        if DYNAMIC_BENCHMARK_PATTERN.search(line):
            formula = re.sub(r"^.*?业绩(?:比较)?基准\s*[=:：]", "", line).strip()
            return re.split(r"[,，]\s*其中", formula, maxsplit=1)[0].strip()
    return lines[0] if lines else text


def strip_balanced_outer_parentheses(text: str) -> str:
    value = text.strip()
    while value.startswith("(") and value.endswith(")"):
        depth = 0
        encloses_all = True
        for index, char in enumerate(value):
            if char == "(":
                depth += 1
            elif char == ")":
                depth -= 1
                if depth == 0 and index != len(value) - 1:
                    encloses_all = False
                    break
        if not encloses_all or depth != 0:
            break
        value = value[1:-1].strip()
    return value


def strip_dynamic_variable(part: str, variable: str) -> str:
    value = part.strip()
    token = re.escape(variable)
    value = re.sub(rf"^\s*{token}\s*%?\s*\*\s*", "", value, flags=re.IGNORECASE)
    value = re.sub(rf"\s*\*\s*{token}\s*%?\s*$", "", value, flags=re.IGNORECASE)
    return strip_balanced_outer_parentheses(value)


def dynamic_variable_component(
    part: str,
    variable: str,
    variable_weight: float,
    catalog: BenchmarkCatalog,
) -> list[dict[str, Any]] | None:
    expression = strip_dynamic_variable(part, variable)
    if not expression or re.search(rf"(?<![A-Za-z]){re.escape(variable)}(?![A-Za-z])", expression, flags=re.IGNORECASE):
        return None
    parsed = parse_benchmark_formula_with_catalog(expression, catalog, select_dated_formula=False)
    inner_components = parsed.get("components") or []
    if parsed.get("missing") or not inner_components:
        return None
    if any(item.get("资产大类") != "权益" for item in inner_components):
        return None
    return [
        {
            **item,
            "weight": round(float(item.get("weight") or 0.0) * variable_weight, 8),
        }
        for item in inner_components
        if float(item.get("weight") or 0.0) > 0
    ]


def parse_dynamic_two_variable_formula(text: str, catalog: BenchmarkCatalog) -> dict[str, Any] | None:
    formula_line = dynamic_formula_line(text)
    parts = split_formula_parts(formula_line)
    if len(parts) != 2:
        return None
    variables: list[str] = []
    for part in parts:
        matches = re.findall(r"(?<![A-Za-z])([AB])(?=\s*(?:%|\*))", part, flags=re.IGNORECASE)
        if len(set(match.upper() for match in matches)) != 1:
            return None
        variables.append(matches[0].upper())
    if set(variables) != {"A", "B"}:
        return None
    disclosed = disclosed_dynamic_xy_for_year(text, datetime.now().year)
    if disclosed is None:
        return None
    first_weight, second_weight, source_segment = disclosed
    weights_by_variable = {"A": first_weight, "B": second_weight}
    components: list[dict[str, Any]] = []
    for part, variable in zip(parts, variables):
        expression = strip_dynamic_variable(part, variable)
        parsed = parse_benchmark_formula_with_catalog(expression, catalog, select_dated_formula=False)
        inner = parsed.get("components") or []
        if parsed.get("missing") or len(inner) != 1:
            return None
        components.append({**inner[0], "weight": round(weights_by_variable[variable], 8)})
    total = sum(float(item.get("weight") or 0.0) for item in components)
    if abs(total - 1.0) > 0.002:
        return None
    description = (
        f"按{datetime.now().year}年披露表格解析："
        + "、".join(f"{item['资产大类']}{item['weight'] * 100:.2f}%" for item in components)
        + f"；依据={source_segment[:160]}"
    )
    return {"components": components, "missing": [], "missing_parts": [], "说明": description}


def parse_dynamic_xy_formula(text: str, catalog: BenchmarkCatalog) -> dict[str, Any] | None:
    formula_line = re.split(r"_x000D_|[\r\n]", text, maxsplit=1, flags=re.IGNORECASE)[0]
    if not (
        re.search(r"X\s*%", formula_line, flags=re.IGNORECASE)
        and re.search(r"Y\s*%", formula_line, flags=re.IGNORECASE)
        and re.search(r"100\s*-\s*X\s*-\s*Y", formula_line, flags=re.IGNORECASE)
    ):
        return None
    if not any(term in formula_line for term in ["债", "国债", "全债", "固定收益"]):
        return None
    disclosed = disclosed_dynamic_xy_for_year(text, datetime.now().year)
    if disclosed is None:
        return None
    x_weight, y_weight, source_segment = disclosed
    equity_weight = x_weight + y_weight
    bond_weight = 1.0 - equity_weight
    if bond_weight < -0.0001:
        return None
    components = [
        component_from_entry(BenchmarkCatalogEntry("权益", "A股", "动态基准权益部分", [], code="DYNAMIC:EQUITY"), equity_weight),
        component_from_entry(BenchmarkCatalogEntry("债券", "债券", "动态基准债券部分", [], code="DYNAMIC:BOND"), max(0.0, bond_weight)),
    ]
    description = (
        f"按{datetime.now().year}年披露表格解析：权益{equity_weight * 100:.2f}%"
        f"（X={x_weight * 100:.2f}%、Y={y_weight * 100:.2f}%）、债券{bond_weight * 100:.2f}%"
        f"；依据={source_segment[:160]}"
    )
    return {"components": components, "missing": [], "missing_parts": [], "说明": description}


def parse_dynamic_single_variable_formula(text: str, catalog: BenchmarkCatalog) -> dict[str, Any] | None:
    formula_line = dynamic_formula_line(text)
    complement_match = re.search(
        r"(?:1\s*-\s*([A-Z])\s*%?|\d+(?:\.\d+)?\s*%?\s*-\s*([A-Z]))",
        formula_line,
        flags=re.IGNORECASE,
    )
    variable = next((value.upper() for value in (complement_match.groups() if complement_match else ()) if value), "")
    if not variable and re.search(r"下滑曲线值|权益类资产的中枢值", formula_line):
        variable = "X"
    if not variable:
        return None
    if variable not in {"A", "I", "S", "X"}:
        return None
    disclosed = disclosed_dynamic_weight_for_year(text, datetime.now().year)
    if disclosed is None:
        return None
    equity_weight, source_segment, used_complement_pair = disclosed
    compact_formula = compact_text(formula_line)
    multiplier_match = re.search(rf"{re.escape(variable)}\s*\*\s*(0\.\d+)", formula_line, flags=re.IGNORECASE)
    if multiplier_match and not used_complement_pair:
        equity_weight *= float(multiplier_match.group(1))
    equity_weight = min(1.0, max(0.0, equity_weight))

    fixed_weights = {"现金": 0.0, "商品": 0.0, "另类": 0.0}
    formula_parts = split_formula_parts(formula_line)
    variable_pattern = rf"(?<![A-Za-z]){re.escape(variable)}(?![A-Za-z])|下滑曲线值|中枢值"
    variable_parts = [
        part
        for part in formula_parts
        if re.search(variable_pattern, part, flags=re.IGNORECASE)
        and not re.search(rf"(?:1\s*-\s*{re.escape(variable)}|\d+(?:\.\d+)?\s*%?\s*-\s*{re.escape(variable)})", part, flags=re.IGNORECASE)
    ]
    for part in formula_parts:
        if re.search(variable_pattern, part, flags=re.IGNORECASE):
            continue
        entry = match_catalog_entry(part, catalog) or infer_generic_index_entry(part)
        weight = parse_weight_from_part(part)
        if entry and weight is not None and entry.asset_major in fixed_weights:
            fixed_weights[entry.asset_major] += weight
    fixed_total = sum(fixed_weights.values())
    if equity_weight + fixed_total > 1.0001:
        return None
    residual = max(0.0, 1.0 - equity_weight - fixed_total)
    if not any(term in compact_formula for term in ["债", "国债", "全债", "固定收益"]):
        return None

    components: list[dict[str, Any]] = []
    if len(variable_parts) == 1:
        components = dynamic_variable_component(variable_parts[0], variable, equity_weight, catalog) or []
    if not components:
        components = [component_from_entry(BenchmarkCatalogEntry("权益", "A股", "动态基准权益部分", [], code="DYNAMIC:EQUITY"), equity_weight)]
    if residual > 0:
        components.append(component_from_entry(BenchmarkCatalogEntry("债券", "债券", "动态基准债券部分", [], code="DYNAMIC:BOND"), residual))
    for major, weight in fixed_weights.items():
        if weight <= 0:
            continue
        category = "现金" if major == "现金" else major
        components.append(component_from_entry(BenchmarkCatalogEntry(major, category, f"动态基准{major}部分", [], code=f"DYNAMIC:{major}"), weight))
    description = (
        f"按{datetime.now().year}年披露表格解析：权益{equity_weight * 100:.2f}%、"
        f"债券{residual * 100:.2f}%"
    )
    if fixed_total > 0:
        description += "、" + "、".join(f"{major}{weight * 100:.2f}%" for major, weight in fixed_weights.items() if weight > 0)
    description += f"；依据={source_segment[:160]}"
    return {"components": components, "missing": [], "missing_parts": [], "说明": description}


def parse_equity_center_formula(text: str, catalog: BenchmarkCatalog) -> dict[str, Any] | None:
    if "权益配置中枢比例" not in text:
        return None
    match = re.search(r"初始基准\s*(\d+(?:\.\d+)?)\s*%", text)
    if not match:
        return None
    equity_weight = float(match.group(1)) / 100.0
    cash_weight = 0.05 if "货币" in text else 0.0
    bond_weight = max(0.0, 1.0 - equity_weight - cash_weight)
    equity_entry = match_catalog_entry("沪深300", catalog)
    bond_entry = match_catalog_entry("上证国债", catalog)
    cash_entry = match_catalog_entry("货币基金指数", catalog)
    if not equity_entry or not bond_entry:
        return None
    components = [component_from_entry(equity_entry, equity_weight), component_from_entry(bond_entry, bond_weight)]
    if cash_weight and cash_entry:
        components.append(component_from_entry(cash_entry, cash_weight))
    description = " + ".join(f'{item["name"]}{item["weight"] * 100:.2f}%' for item in components)
    return {"components": components, "missing": [], "missing_parts": [], "说明": f"{description}（按披露初始基准还原）"}


def strip_static_definition_suffix(text: str) -> str:
    """Remove a trailing index-definition equation after a complete 100% formula."""
    if not is_static_benchmark_formula(text):
        return text
    match = re.search(r"[,，]\s*(?:复合指数|成分指数)\s*=", text)
    if not match:
        return text
    candidate = text[: match.start()].strip()
    weights = [float(value) for value in re.findall(r"(\d+(?:\.\d+)?)\s*%", candidate)]
    if candidate.count("指数") < 2 or not weights or abs(sum(weights) - 100.0) > 0.02:
        return text
    return candidate


def effective_dated_benchmark_formula(text: str, year: int) -> tuple[str, str] | None:
    normalized = re.sub(r"_x000D_", "\n", text, flags=re.IGNORECASE)
    candidates: list[tuple[int, int, str, str]] = []
    for order, segment in enumerate(re.split(r"[\r\n;；]+", normalized)):
        segment = segment.strip()
        if segment.count("指数") < 2 or len(re.findall(r"\d+(?:\.\d+)?\s*%", segment)) < 2:
            continue
        years = [int(value) for value in re.findall(r"20\d{2}", segment)]
        if not years:
            continue
        start_year = min(years)
        end_year = max(years)
        prefix = segment[: segment.find(str(years[0]))]
        suffix_after_last_year = segment[segment.rfind(str(years[-1])) + 4 :]
        if any(term in prefix for term in ["成立", "生效", "设立"]):
            start_year = 1900
        if any(term in suffix_after_last_year[:40] for term in ["以后", "及以后", "起"]):
            end_year = 9999
        if not (start_year <= year <= end_year):
            continue
        formula_match = re.search(
            r"(?=(?:沪深|中证|中债|上证|深证|新华|中国债券|恒生|MSCI|标普|纳斯达克|银行|[一1]年期))",
            segment,
            flags=re.IGNORECASE,
        )
        if not formula_match:
            continue
        formula = segment[formula_match.start() :].strip(" ,，:\t")
        if formula.count("指数") < 2:
            continue
        candidates.append((end_year - start_year, -order, formula, segment))
    if not candidates:
        return None
    _, _, formula, source_segment = min(candidates, key=lambda item: (item[0], item[1]))
    return formula, source_segment


def parse_disclosed_date(text: str) -> date | None:
    match = re.search(r"(20\d{2})\s*(?:[./-]|\s*年\s*)(\d{1,2})\s*(?:[./-]|\s*月\s*)(\d{1,2})\s*(?:日)?", text)
    if not match:
        return None
    try:
        return date(int(match.group(1)), int(match.group(2)), int(match.group(3)))
    except ValueError:
        return None


def effective_numbered_period_formula(text: str, as_of: date) -> tuple[str, str] | None:
    """Select a numbered legal-disclosure block when its effective date is explicit."""
    normalized = re.sub(r"_x000D_", "\n", text, flags=re.IGNORECASE)
    blocks = [block.strip() for block in re.split(r"(?=[(（]\s*\d+\s*[)）])", normalized) if block.strip()]
    candidates: list[tuple[int, str, str]] = []
    for block in blocks:
        lines = [line.strip() for line in block.splitlines() if line.strip()]
        if len(lines) < 2 or not re.match(r"^[(（]\s*\d+\s*[)）]", lines[0]):
            continue
        heading = lines[0]
        disclosed = parse_disclosed_date(heading)
        if disclosed is None:
            continue
        if any(term in heading for term in ["之后", "以后", "后", "起"]):
            start, end = disclosed, date.max
        elif any(term in heading for term in ["至", "～", "~", "-", "截至"]):
            start, end = date.min, disclosed
        else:
            continue
        if not (start <= as_of <= end):
            continue
        formula_index = next((index for index, line in enumerate(lines[1:], start=1) if "业绩" in line and "基准" in line), None)
        if formula_index is None:
            continue
        selected_lines = lines[formula_index:]
        formula = "\n".join(selected_lines)
        candidates.append(((end - start).days if end != date.max and start != date.min else 10**9, formula, heading))
    if not candidates:
        return None
    _, formula, heading = min(candidates, key=lambda item: item[0])
    return formula, heading


def parse_benchmark_formula_with_catalog(
    benchmark_text: Any,
    catalog: BenchmarkCatalog,
    *,
    select_dated_formula: bool = True,
) -> dict[str, Any]:
    raw = clean_text(benchmark_text, "") or ""
    if not raw:
        return {"components": [], "missing": ["未披露"], "missing_parts": [], "说明": "未披露业绩基准"}
    normalized = normalize_formula_text(raw)
    numbered_formula = effective_numbered_period_formula(normalized, datetime.now().date()) if select_dated_formula else None
    if numbered_formula:
        effective_formula, source_segment = numbered_formula
        effective_result = parse_benchmark_formula_with_catalog(effective_formula, catalog, select_dated_formula=False)
        effective_result["说明"] = (
            f"按{datetime.now().date().isoformat()}生效区间解析：{effective_result.get('说明') or ''}；"
            f"依据={source_segment[:180]}"
        )
        return effective_result
    dated_formula = effective_dated_benchmark_formula(normalized, datetime.now().year) if select_dated_formula else None
    if dated_formula:
        effective_formula, source_segment = dated_formula
        effective_result = parse_benchmark_formula_with_catalog(
            effective_formula,
            catalog,
            select_dated_formula=False,
        )
        effective_result["说明"] = (
            f"按{datetime.now().year}年生效区间解析：{effective_result.get('说明') or ''}；"
            f"依据={source_segment[:180]}"
        )
        return effective_result
    dynamic_xy = parse_dynamic_xy_formula(normalized, catalog)
    if dynamic_xy:
        return dynamic_xy
    dynamic_two_variable = parse_dynamic_two_variable_formula(normalized, catalog)
    if dynamic_two_variable:
        return dynamic_two_variable
    dynamic = parse_dynamic_single_variable_formula(normalized, catalog)
    if dynamic:
        return dynamic
    equity_center = parse_equity_center_formula(normalized, catalog)
    if equity_center:
        return equity_center
    if DYNAMIC_BENCHMARK_PATTERN.search(normalized):
        return {"components": [], "missing": ["动态基准当期变量值"], "missing_parts": [], "说明": f"暂未解析动态权重基准：{raw}"}
    cash_average = parse_cash_average_formula(normalized, catalog)
    if cash_average:
        return cash_average

    formula_body = strip_static_definition_suffix(normalized)
    parts = [part for part in split_formula_parts(formula_body) if part.strip()]
    fixed_return_parts = [part.strip() for part in parts if is_fixed_return_part(part)]
    asset_parts = [part for part in parts if not is_fixed_return_part(part)]
    parsed: dict[str, dict[str, Any]] = {}
    missing: list[str] = []
    missing_parts: list[dict[str, Any]] = []
    for part in parts or [normalized]:
        if is_fixed_return_part(part):
            continue
        entry = match_catalog_entry(part, catalog)
        if not entry:
            entry = infer_generic_index_entry(part)
        weight = parse_weight_from_part(part)
        if not entry:
            if "%" in part or "指数" in part or weight is not None or len(parts) <= 1:
                missing.append(f"未映射组件：{part.strip()}")
                missing_parts.append({"text": part.strip(), "weight": round(weight, 8) if weight is not None else None})
            continue
        if weight is None and (len(parts) <= 1 or (fixed_return_parts and len(asset_parts) == 1)):
            weight = 1.0
        if weight is None:
            missing.append(f"未解析权重：{part.strip()}")
            missing_parts.append({"text": part.strip(), "weight": None})
            continue
        add_or_update_component(parsed, entry, weight)

    components = list(parsed.values())
    total_weight = sum(float(item["weight"]) for item in components)
    missing_weight = sum(float(item["weight"]) for item in missing_parts if item.get("weight") is not None)
    if fixed_return_parts and components and not missing and 0.0 < total_weight < 0.995:
        cash_entry = match_catalog_entry("现金/存款", catalog) or BenchmarkCatalogEntry("现金", "现金", "现金/固定收益残余", [], code="GENERIC:CASH_RESIDUAL")
        add_or_update_component(parsed, cash_entry, 1.0 - total_weight)
        components = list(parsed.values())
        total_weight = sum(float(item["weight"]) for item in components)
    if components and not missing and 0.0 < total_weight <= 1.5:
        for item in components:
            item["weight"] = round(float(item["weight"]) / total_weight, 8)
    elif components and missing_weight <= 0.0 and 0.0 < total_weight < 0.995:
        missing.append(f"权重合计不足：{round(total_weight * 100, 4)}%")
    elif components and total_weight + missing_weight > 1.5:
        missing.append(f"权重合计异常：{round((total_weight + missing_weight) * 100, 4)}%")
    description = " + ".join(f'{item["name"]}{item["weight"] * 100:.2f}%' for item in components) if components else "未解析出可计算组件"
    if fixed_return_parts and components:
        description += f"（固定收益/加点组件按现金残余或已有现金基准处理：{'、'.join(fixed_return_parts[:3])}）"
    return {"components": components, "missing": missing, "missing_parts": missing_parts, "说明": description}


def equity_level(equity_pct: float | None, parsed: bool) -> str | None:
    if not parsed or equity_pct is None:
        return None
    if equity_pct <= 0:
        return "L0"
    return f"L{min(10, max(1, math.ceil(equity_pct / 10.0)))}"


def compute_benchmark_asset_mix(benchmark_text: Any, catalog: BenchmarkCatalog) -> dict[str, Any]:
    parsed = parse_benchmark_formula_with_catalog(benchmark_text, catalog)
    major_weights = {name: 0.0 for name in [*catalog.asset_majors, OTHER_BUCKET]}
    category_weights = {name: 0.0 for name in [*catalog.asset_categories, OTHER_BUCKET]}
    mapped_weight = 0.0
    for item in parsed.get("components") or []:
        weight_pct = float(item.get("weight") or 0.0) * 100.0
        major = clean_text(item.get("资产大类"), OTHER_BUCKET) or OTHER_BUCKET
        category = clean_text(item.get("资产类别"), OTHER_BUCKET) or OTHER_BUCKET
        if major not in major_weights:
            major_weights[OTHER_BUCKET] += weight_pct
        else:
            major_weights[major] += weight_pct
        if category not in category_weights:
            category_weights[OTHER_BUCKET] += weight_pct
        else:
            category_weights[category] += weight_pct
        mapped_weight += weight_pct

    missing_weight = 0.0
    missing_without_weight = 0
    for item in parsed.get("missing_parts") or []:
        weight = item.get("weight")
        if weight is None:
            missing_without_weight += 1
            continue
        missing_weight += float(weight) * 100.0
    if missing_weight > 0:
        major_weights[OTHER_BUCKET] += missing_weight
        category_weights[OTHER_BUCKET] += missing_weight

    parsed_any = mapped_weight > 0 or missing_weight > 0
    if not parsed_any:
        mix: dict[str, Any] = {
            "基准公式解析": parsed.get("说明"),
            "基准缺失组件": parsed.get("missing", []),
            "基准资产大类JSON": None,
            "基准资产类别JSON": None,
            "基准映射置信度": "未解析",
            "基准资产已映射权重": None,
            "基准资产未映射权重": None,
            "基准风险资产权重": None,
            "基准风险资产权重_百分比": None,
            "基准结构类型": "未知",
            "非权益比较轨道": "未纳入",
            "正式可比池": "",
            "可比池样本资格": "否",
            "可比池说明": "基准未解析",
            "是否多元策略": 0,
            "多元策略标签": "",
        }
        for field_name in [*catalog.asset_major_fields, *catalog.asset_category_fields]:
            mix[field_name] = None
        return mix

    for weights in (major_weights, category_weights):
        total = sum(weights.values())
        if 0.0 < total <= 150.0 and abs(total - 100.0) > 0.01:
            scale = 100.0 / total
            for key in list(weights):
                weights[key] *= scale

    rounded_major = {key: round_or_none(value) for key, value in major_weights.items()}
    rounded_category = {key: round_or_none(value) for key, value in category_weights.items()}
    equity_pct = rounded_major.get("权益") or 0.0
    overseas_pct = (rounded_category.get("港股") or 0.0) + (rounded_category.get("海外权益") or 0.0)
    commodity_pct = rounded_major.get("商品") or 0.0
    alternative_pct = rounded_major.get("另类") or 0.0
    a_share_pct = rounded_category.get("A股") or 0.0
    is_diversified = int(a_share_pct > 0 and overseas_pct > 0 and (commodity_pct + alternative_pct) > 0)
    if is_diversified:
        diversified_label = "A股+海外+另类"
    else:
        active_parts = [name for name, value in rounded_category.items() if name != OTHER_BUCKET and (value or 0) > 0]
        diversified_label = "、".join(active_parts) if len(active_parts) >= 3 else ""

    confidence = "高"
    if parsed.get("missing") or missing_without_weight:
        confidence = "中" if mapped_weight >= 70.0 and missing_without_weight == 0 else "低"
    risk_asset_pct = equity_pct + commodity_pct + alternative_pct
    risk_asset_complete = mapped_weight > 0 and missing_weight <= 0 and missing_without_weight == 0
    bucket = equity_level(risk_asset_pct, risk_asset_complete)
    comparison = build_comparison_pool(
        bucket=bucket or "",
        equity=equity_pct,
        bond=rounded_major.get("债券"),
        cash=rounded_major.get("现金"),
        commodity=commodity_pct,
        alternative=alternative_pct,
        unknown=rounded_major.get(OTHER_BUCKET),
    )
    if confidence not in {"高", "中"}:
        comparison.update(
            {
                "基准结构类型": "未知",
                "非权益比较轨道": "未纳入",
                "正式可比池": "",
                "可比池样本资格": "否",
                "可比池说明": f"基准映射置信度={confidence}",
            }
        )
    mix = {
        "基准公式解析": parsed.get("说明"),
        "基准缺失组件": parsed.get("missing", []),
        "基准资产大类JSON": json.dumps(rounded_major, ensure_ascii=False),
        "基准资产类别JSON": json.dumps(rounded_category, ensure_ascii=False),
        "基准映射置信度": confidence,
        "基准资产已映射权重": round_or_none(mapped_weight),
        "基准资产未映射权重": round_or_none(missing_weight),
        "基准风险资产权重": bucket,
        "基准风险资产权重_百分比": round_or_none(risk_asset_pct) if risk_asset_complete else None,
        "基准港股权益权重": rounded_category.get("港股") or 0.0,
        "基准海外权益权重": rounded_category.get("海外权益") or 0.0,
        **comparison,
        "是否多元策略": is_diversified,
        "多元策略标签": diversified_label,
    }
    for key, value in rounded_major.items():
        mix[f"基准资产大类-{key}"] = value
    for key, value in rounded_category.items():
        mix[f"基准资产类别-{key}"] = value
    return mix


def build_strategy_benchmark_asset_rows(conn: sqlite3.Connection, catalog: BenchmarkCatalog | None = None) -> list[dict[str, Any]]:
    catalog = catalog or load_benchmark_catalog()
    now = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    rows: list[dict[str, Any]] = []
    query = """
        SELECT s."统一策略ID", s."渠道ID", c."渠道名称", s."渠道策略ID", s."策略名称", s."投顾机构",
               s."业绩基准" AS "策略业绩基准", b."业绩基准文本" AS "状态业绩基准"
        FROM "策略信息" s
        LEFT JOIN "渠道信息" c ON c."渠道ID" = s."渠道ID"
        LEFT JOIN "策略基准费率状态" b ON b."统一策略ID" = s."统一策略ID"
        ORDER BY s."渠道ID", s."投顾机构", s."策略名称"
    """
    for row in conn.execute(query):
        item = dict(row)
        benchmark_text = clean_text(item.get("策略业绩基准") or item.get("状态业绩基准"), "")
        mix = compute_benchmark_asset_mix(benchmark_text, catalog)
        output = {
            "统一策略ID": item.get("统一策略ID"),
            "渠道ID": item.get("渠道ID"),
            "渠道名称": item.get("渠道名称"),
            "渠道策略ID": item.get("渠道策略ID"),
            "策略名称": item.get("策略名称"),
            "投顾机构": item.get("投顾机构"),
            "业绩基准文本": benchmark_text,
            **mix,
            "最近更新时间": now,
        }
        rows.append(output)
    return rows


def column_type(column: str) -> str:
    if column == "是否多元策略":
        return "INTEGER"
    if column.startswith("基准资产大类-") or column.startswith("基准资产类别-"):
        return "REAL"
    if column in {"基准资产已映射权重", "基准资产未映射权重", "基准港股权益权重", "基准海外权益权重", "基准互斥权重合计_百分比", "基准风险资产权重_百分比"}:
        return "REAL"
    return "TEXT"


def write_strategy_benchmark_asset_table(conn: sqlite3.Connection, rows: list[dict[str, Any]]) -> dict[str, Any]:
    conn.execute(f'DROP TABLE IF EXISTS "{ASSET_TABLE_NAME}"')
    if not rows:
        conn.execute(
            f'''
            CREATE TABLE "{ASSET_TABLE_NAME}" (
                "统一策略ID" TEXT PRIMARY KEY
            )
            '''
        )
        return {"asset_rows": 0, "asset_parsed": 0, "asset_high_confidence": 0}
    columns: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                columns.append(key)
    definitions = []
    for column in columns:
        suffix = " PRIMARY KEY" if column == "统一策略ID" else ""
        definitions.append(f'"{column}" {column_type(column)}{suffix}')
    conn.execute(f'CREATE TABLE "{ASSET_TABLE_NAME}" ({", ".join(definitions)})')
    placeholders = ",".join(["?"] * len(columns))
    quoted_columns = ",".join(f'"{column}"' for column in columns)
    conn.executemany(
        f'INSERT INTO "{ASSET_TABLE_NAME}" ({quoted_columns}) VALUES ({placeholders})',
        [[json.dumps(value, ensure_ascii=False) if isinstance(value, (list, dict)) else value for value in (row.get(column) for column in columns)] for row in rows],
    )
    return {
        "asset_rows": len(rows),
        "asset_parsed": sum(1 for row in rows if row.get("基准风险资产权重")),
        "asset_high_confidence": sum(1 for row in rows if row.get("基准映射置信度") == "高"),
        "asset_diversified": sum(1 for row in rows if int(row.get("是否多元策略") or 0) == 1),
    }


def export_rows_csv(rows: list[dict[str, Any]], path: Path) -> None:
    import csv

    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    fieldnames: list[str] = []
    seen: set[str] = set()
    for row in rows:
        for key in row:
            if key not in seen:
                seen.add(key)
                fieldnames.append(key)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)
