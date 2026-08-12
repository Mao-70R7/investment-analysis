# -*- coding: utf-8 -*-
from __future__ import annotations

import argparse
import json
import math
import random
import re
import sqlite3
import time
from collections import defaultdict
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime, timedelta
from html import escape
from pathlib import Path
from statistics import median, stdev
from typing import Any

import requests

from basic_data_navigation import SIDEBAR_CSS, render_system_topbar


PROJECT_ROOT = next(parent for parent in Path(__file__).resolve().parents if (parent / "AGENTS.md").is_file())
DEFAULT_DB_PATH = PROJECT_ROOT / "data" / "analysis_zh_current.sqlite"
DEFAULT_REPORT_ROOT = PROJECT_ROOT / "site"
DEFAULT_SITE_DIR = DEFAULT_REPORT_ROOT / "basic_data"
DEFAULT_PURCHASE_CACHE = PROJECT_ROOT / "outputs" / "qd_fund_purchase_status" / "latest.json"
DEFAULT_LIMIT_ANNOUNCEMENT_CACHE = PROJECT_ROOT / "outputs" / "qd_fund_limit_announcements" / "latest.json"
MAX_LIMIT_NOTICE_CHECKS = 6
AMOUNT_RE = re.compile(
    r"(?P<num>(?:\d{1,3}(?:,\d{3})+|\d+)(?:\.\d+)?)\s*(?P<wan>万)?\s*(?P<currency>元人民币|人民币元|万元人民币|人民币|元|万美元|美元|港元|港币)"
)


def q(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def as_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        num = float(value)
    except (TypeError, ValueError):
        return None
    return num if math.isfinite(num) else None


def as_bool(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if value is None:
        return False
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "是"}


def pct(value: float | None, digits: int = 2, signed: bool = False) -> str:
    if value is None:
        return "--"
    sign = "+" if signed and value > 0 else ""
    return f"{sign}{value:.{digits}f}%"


def compact_text(value: Any, limit: int = 42) -> str:
    text = re.sub(r"\s+", " ", str(value or "")).strip()
    if not text:
        return ""
    return text[:limit] + ("..." if len(text) > limit else "")


def normalize_target_series_name(name: str) -> str:
    text = re.sub(r"\s+", "", name or "")
    text = re.sub(r"[（(][^）)]*(?:第[零一二三四五六七八九十百千万\d]+期|止盈|到期|运行中|目标盈)[^）)]*[）)]", "", text)
    text = re.sub(r"第[零一二三四五六七八九十百千万\d]+期", "", text)
    text = re.sub(r"\d{4}[-年]?\d{1,2}[-月]?\d{0,2}日?", "", text)
    text = re.sub(r"\d{6,8}$", "", text)
    text = re.sub(r"(目标盈|止盈|到期|已停止|已到期|运行中)", "", text)
    text = re.sub(r"[-_]+$", "", text)
    return text or (name or "未命名目标盈系列")


def strategy_unit(row: dict[str, Any]) -> tuple[str, str, str]:
    sid = str(row.get("strategy_id") or "")
    name = str(row.get("strategy_name") or sid or "未命名策略")
    advisor = str(row.get("advisor") or "未披露机构")
    if as_bool(row.get("is_target")):
        series = normalize_target_series_name(name)
        return f"target::{advisor}::{series}", series, "目标盈系列合并"
    return f"strategy::{sid}", name, str(row.get("governance_status") or "常规策略")


def load_json(value: Any, fallback: Any) -> Any:
    if value in (None, ""):
        return fallback
    try:
        return json.loads(str(value))
    except Exception:
        return fallback


def theme_labels(value: Any, limit: int = 3) -> list[str]:
    raw = load_json(value, [])
    items = raw.values() if isinstance(raw, dict) else raw if isinstance(raw, list) else []
    labels: list[str] = []
    for item in items:
        if isinstance(item, dict):
            label = item.get("主题名称") or item.get("名称") or item.get("name") or item.get("label")
        else:
            label = str(item or "")
        label = re.sub(r"\s+", "", str(label or "")).strip()
        if label and label not in labels and label not in {"-", "--", "未识别", "未分类"}:
            labels.append(label)
        if len(labels) >= limit:
            break
    return labels


def is_qd_fund(row: dict[str, Any]) -> bool:
    text = " ".join(
        str(row.get(key) or "")
        for key in ("fund_name", "fund_company", "fund_type2", "region", "standard_sub_asset", "advisor_asset_bucket", "expo_sub_asset")
    )
    return as_bool(row.get("is_qdii")) or "QDII" in text or "海外/QDII" in text


def normalize_qd_fund_base_name(name: str) -> str:
    text = re.sub(r"\s+", "", str(name or ""))
    text = text.replace("（", "(").replace("）", ")")
    text = re.sub(r"[\-_]?(?:A|B|C|D|E|F|H|I|Y)(?=\((?:人民币|人民币份额|美元|美元现汇|美元现钞|港元|港币)[^)]*\)$)", "", text)
    text = re.sub(r"(人民币|人民币份额|美元现汇|美元现钞|美元|港元|港币)(?:A|B|C|D|E|F|H|I|Y)(?:类)?$", r"\1", text)
    text = re.sub(r"[\-_]?(?:A|B|C|D|E|F|H|I|Y)(?:类)?$", "", text)
    text = re.sub(r"(?:收费|份额)$", "", text)
    return text or str(name or "")


def qd_share_class(name: str) -> str:
    text = re.sub(r"\s+", "", str(name or "")).replace("（", "(").replace("）", ")")
    patterns = [
        r"(?:人民币|人民币份额|美元现汇|美元现钞|美元|港元|港币)(A|B|C|D|E|F|H|I|Y)(?:类)?$",
        r"[\-_]?(A|B|C|D|E|F|H|I|Y)(?:类)?$",
        r"[\-_]?(A|B|C|D|E|F|H|I|Y)(?=\((?:人民币|人民币份额|美元|美元现汇|美元现钞|港元|港币)[^)]*\)$)",
    ]
    for pattern in patterns:
        match = re.search(pattern, text)
        if match:
            return match.group(1)
    return ""


def qd_fund_group_key(row: dict[str, Any]) -> str:
    company = re.sub(r"\s+", "", str(row.get("fund_company") or ""))
    base = normalize_qd_fund_base_name(str(row.get("fund_name") or row.get("fund_code") or ""))
    return f"{company}::{base}"


def qd_main_share_score(code: str, row: dict[str, Any], purchase: dict[str, Any] | None, announcement: dict[str, Any] | None, returns: dict[str, Any] | None) -> tuple[int, int, int, int, str]:
    share = qd_share_class(str(row.get("fund_name") or ""))
    share_rank = {"A": 0, "": 1, "I": 2, "B": 3, "D": 4, "E": 5, "F": 6, "H": 7, "Y": 8, "C": 9}.get(share, 6)
    ann = merged_announcement_limit(announcement)
    has_limit = 0 if ann.get("personal") or ann.get("institution") or purchase else 1
    has_return = 0 if returns else 1
    is_c_like = 1 if share in {"C", "F", "Y"} else 0
    return (has_limit, share_rank, has_return, is_c_like, code)


def qd_direction(row: dict[str, Any]) -> str:
    text = " ".join(
        str(row.get(key) or "")
        for key in ("fund_name", "fund_type2", "region", "standard_sub_asset", "advisor_asset_bucket", "expo_sub_asset", "themes")
    )
    if any(token in text for token in ("美元债", "海外债", "亚洲债", "全球债", "高收益债", "票息")):
        return "海外债"
    if any(token in text for token in ("纳斯达克", "纳指", "标普", "美国", "美股", "道琼斯")):
        return "美股"
    if any(token in text for token in ("恒生", "港股", "香港", "中概", "大中华")):
        return "港股/中概"
    if "日本" in text:
        return "日本"
    if any(token in text for token in ("德国", "欧洲", "DAX")):
        return "欧洲"
    if "印度" in text:
        return "印度"
    if any(token in text for token in ("越南", "东南亚", "亚太")):
        return "亚太"
    if any(token in text for token in ("黄金", "抗通胀", "商品")):
        return "商品/黄金"
    if any(token in text for token in ("原油", "油气")):
        return "原油/能源"
    if any(token in text for token in ("全球", "海外", "发达市场")):
        return "全球配置"
    return "其他QDII"


def money_text(value: Any) -> str:
    amount = as_float(value)
    if amount is None:
        return "未披露"
    if amount <= 0 or amount >= 100_000_000:
        return "无明显限制"
    if amount >= 10_000:
        return f"{amount / 10000:.0f}万元"
    return f"{amount:.0f}元"


def parsed_amount_text(parsed: dict[str, Any] | None) -> str:
    if not parsed:
        return "未单独披露"
    amount = parsed.get("amount")
    currency = parsed.get("currency") or "CNY"
    if amount is None:
        return "未单独披露"
    unit = "美元" if currency == "USD" else "港元" if currency == "HKD" else "元"
    if amount >= 10000:
        return f"{amount / 10000:g}万{unit}"
    return f"{amount:g}{unit}"


def sales_daily_limit_amount(purchase: dict[str, Any] | None) -> float | None:
    if not purchase:
        return None
    amount = as_float(purchase.get("日累计限定金额"))
    if amount is None:
        return None
    if amount <= 0:
        return 100_000_000
    return amount


def sales_limit_text(purchase: dict[str, Any] | None) -> str:
    amount = sales_daily_limit_amount(purchase)
    if amount is None:
        return "销售端未采集"
    if amount >= 100_000_000:
        return "销售端未设日限额"
    return f"销售端参考 {money_text(amount)}"


def investor_limit_text(parsed: dict[str, Any] | None, purchase: dict[str, Any] | None, source_type: str = "") -> str:
    if parsed:
        prefix = "公告" if source_type == "explicit" else "通用公告参考"
        return f"{prefix} {parsed_amount_text(parsed)}"
    return sales_limit_text(purchase)


def investor_limit_amount(parsed: dict[str, Any] | None, purchase: dict[str, Any] | None) -> float | None:
    if parsed and parsed.get("currency") == "CNY":
        return as_float(parsed.get("amount"))
    return sales_daily_limit_amount(purchase)


def public_limit_detail(limit_detail: dict[str, Any]) -> dict[str, Any]:
    parse_note = str(limit_detail.get("parseNote") or "")
    parse_note = parse_note.replace("非个人/机构", "非个人")
    parse_note = re.sub(r"；?产品户限额未单独披露；?", "", parse_note)
    return {
        "personal": limit_detail.get("personal"),
        "nonPerson": limit_detail.get("institution"),
        "general": limit_detail.get("general"),
        "personalSourceType": limit_detail.get("personalSourceType") or "",
        "nonPersonSourceType": limit_detail.get("institutionSourceType") or "",
        "businessScope": limit_detail.get("businessScope") or "",
        "effectiveDate": limit_detail.get("effectiveDate") or "",
        "caliber": limit_detail.get("caliber") or "",
        "sourceTitle": limit_detail.get("sourceTitle") or "",
        "sourceDate": limit_detail.get("sourceDate") or "",
        "sourceUrl": limit_detail.get("sourceUrl") or "",
        "parseStatus": limit_detail.get("parseStatus") or "",
        "parseNote": parse_note.strip("；"),
    }


def parse_amount(match: re.Match[str]) -> dict[str, Any]:
    num = float(match.group("num").replace(",", ""))
    currency_text = match.group("currency") or ""
    multiplier = 10000 if match.group("wan") or currency_text.startswith("万") else 1
    if "美元" in currency_text:
        currency = "USD"
    elif "港" in currency_text:
        currency = "HKD"
    else:
        currency = "CNY"
    return {"amount": num * multiplier, "currency": currency, "raw": match.group(0)}


def first_amount(text: str) -> dict[str, Any] | None:
    for match in AMOUNT_RE.finditer(text):
        parsed = parse_amount(match)
        if as_float(parsed.get("amount")) and float(parsed.get("amount") or 0) > 0:
            return parsed
    return None


def clean_notice_text(text: str) -> str:
    text = re.sub(r"\s+", " ", text or "")
    return text.strip()


def notice_is_relevant(title: str) -> bool:
    title = title or ""
    if any(token in title for token in ("终止", "销售业务", "转托管", "节假日", "红利", "清算", "基金合同", "招募说明书")):
        return False
    return any(token in title for token in ("限额", "金额限制", "大额申购", "暂停申购", "恢复申购", "限制大额"))


def find_context_amount(text: str, keywords: list[str], *, before_chars: int = 80, after_chars: int = 420) -> tuple[dict[str, Any] | None, str]:
    for keyword in keywords:
        for match in re.finditer(re.escape(keyword), text):
            start = max(0, match.start() - before_chars)
            end = min(len(text), match.end() + after_chars)
            segment = text[start:end]
            amount = first_amount(segment)
            if amount:
                return amount, segment
    return None, ""


def find_context_amount_patterns(text: str, patterns: list[str]) -> tuple[dict[str, Any] | None, str]:
    for pattern in patterns:
        for match in re.finditer(pattern, text):
            start = match.start()
            end = min(len(text), match.end() + 420)
            segment = text[start:end]
            amount = first_amount(segment)
            if amount:
                return amount, segment
    return None, ""


def parse_limit_notice(title: str, publish_date: str, notice_id: str, content: str, attach_url: str = "") -> dict[str, Any]:
    text = clean_notice_text(content)
    is_pause = "暂停申购" in title and "恢复" not in title
    is_resume = "恢复申购" in title
    general_amount, general_context = find_context_amount(text, ["限制金额", "限额", "不超过", "超过", "累计金额应不超过"])
    personal_amount, personal_context = find_context_amount_patterns(text, [r"(?<!非)个人投资者", r"(?<!非)个人客户", r"(?<!非)个人"])
    institution_amount, institution_context = find_context_amount_patterns(text, [r"非个人投资者", r"非个人客户", r"机构投资者", r"机构客户", r"非个人", r"机构"])
    product_amount, product_context = find_context_amount(text, ["产品户", "产品账户", "资管产品", "资产管理产品", "产品投资者", "基金产品"])
    business_scope = "申购、定投、转换转入" if any(k in text for k in ("定投", "转换转入")) else "申购"
    effective_match = re.search(r"(?:自|起始日)\s*(\d{4})\s*年\s*(\d{1,2})\s*月\s*(\d{1,2})\s*日", text)
    effective_date = ""
    if effective_match:
        effective_date = f"{int(effective_match.group(1)):04d}-{int(effective_match.group(2)):02d}-{int(effective_match.group(3)):02d}"
    status = "暂停申购" if is_pause else "恢复申购" if is_resume else "限大额" if general_amount else "公告待人工确认"
    parse_status = "parsed" if any([general_amount, personal_amount, institution_amount, product_amount, is_pause, is_resume]) else "unparsed"
    parse_note = ""
    if personal_amount is None and general_amount is not None:
        parse_note += "个人限额未单独披露，使用通用限额参考；"
    if institution_amount is None and general_amount is not None:
        parse_note += "非个人限额未单独披露，使用通用限额参考；"
    return {
        "公告ID": notice_id,
        "公告标题": title,
        "公告日期": publish_date,
        "公告URL": f"https://np-cnotice-stock.eastmoney.com/api/content/ann?art_code={notice_id}&client_source=web_fund",
        "PDF链接": attach_url,
        "公告状态": status,
        "通用日限额": general_amount,
        "个人日限额": personal_amount,
        "非个人机构日限额": institution_amount,
        "产品户日限额": product_amount,
        "业务范围": business_scope,
        "生效日期": effective_date or publish_date,
        "限额口径": "单日单个基金账户/销售渠道口径，以公告原文为准",
        "解析状态": parse_status,
        "解析说明": parse_note.rstrip("；"),
        "证据摘录": compact_text(general_context or personal_context or institution_context or product_context or text, 220),
    }


def fetch_limit_announcement_for_code(code: str) -> tuple[str, dict[str, Any]]:
    def get_json(url: str, *, params: dict[str, Any], headers: dict[str, str], retries: int = 5) -> dict[str, Any]:
        last_error: Exception | None = None
        for attempt in range(retries):
            try:
                resp = requests.get(url, params=params, headers=headers, timeout=20)
                resp.raise_for_status()
                try:
                    return resp.json()
                except ValueError:
                    text = resp.text.strip()
                    jsonp = re.match(r"^[\w$]+\((.*)\);?$", text, flags=re.S)
                    if jsonp:
                        return json.loads(jsonp.group(1))
                    raise
            except Exception as exc:
                last_error = exc
                time.sleep(0.8 * (attempt + 1) + random.uniform(0.05, 0.35))
        raise RuntimeError(str(last_error))

    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124 Safari/537.36",
        "Accept": "application/json,text/plain,*/*",
        "Referer": f"https://fundf10.eastmoney.com/jjgg_{code}_5.html",
        "Origin": "https://fundf10.eastmoney.com",
        "Connection": "close",
    }
    list_json = get_json(
        "https://api.fund.eastmoney.com/f10/JJGG",
        params={"callback": "qdNoticeCallback", "fundcode": code, "pageIndex": "1", "pageSize": "40", "type": "5", "_": round(datetime.now().timestamp() * 1000)},
        headers=headers,
    )
    announcements = list_json.get("Data") or []
    parsed_by_type: dict[str, dict[str, Any]] = {}
    checked: list[dict[str, Any]] = []
    for ann in announcements:
        title = ann.get("TITLE") or ""
        if not notice_is_relevant(title):
            continue
        notice_id = ann.get("ID") or ""
        publish_date = ann.get("PUBLISHDATEDesc") or str(ann.get("PUBLISHDATE") or "")[:10]
        content_headers = {
            "User-Agent": headers["User-Agent"],
            "Accept": "application/json,text/plain,*/*",
            "Referer": "https://fundf10.eastmoney.com/",
        }
        try:
            content_json = get_json(
                "https://np-cnotice-stock.eastmoney.com/api/content/ann",
                params={"art_code": notice_id, "client_source": "web_fund"},
                headers=content_headers,
            )
        except Exception:
            content_json = get_json(
                "https://np-cnotice-fund.eastmoney.com/api/content/ann",
                params={"art_code": notice_id, "client_source": "web_fund"},
                headers=content_headers,
            )
        data = content_json.get("data") or {}
        parsed = parse_limit_notice(title, publish_date, notice_id, data.get("notice_content") or "", data.get("attach_url_web") or data.get("attach_url") or "")
        checked.append(parsed)
        if len(checked) >= MAX_LIMIT_NOTICE_CHECKS and parsed["解析状态"] != "parsed":
            break
        if parsed["解析状态"] != "parsed":
            continue
        if parsed.get("个人日限额") and "personal" not in parsed_by_type:
            parsed_by_type["personal"] = parsed
        if parsed.get("非个人机构日限额") and "institution" not in parsed_by_type:
            parsed_by_type["institution"] = parsed
        if parsed.get("产品户日限额") and "product" not in parsed_by_type:
            parsed_by_type["product"] = parsed
        if parsed.get("通用日限额") and "general" not in parsed_by_type:
            parsed_by_type["general"] = parsed
        if parsed.get("公告状态") in {"暂停申购", "恢复申购"} and "status" not in parsed_by_type:
            parsed_by_type["status"] = parsed
        if parsed_by_type.get("general") or parsed_by_type.get("status") or len(checked) >= MAX_LIMIT_NOTICE_CHECKS:
            break
    chosen = {
        "checked": checked[:8],
        "general": parsed_by_type.get("general") or {},
        "personal": parsed_by_type.get("personal") or {},
        "institution": parsed_by_type.get("institution") or {},
        "product": parsed_by_type.get("product") or {},
        "status": parsed_by_type.get("status") or {},
    }
    return code, chosen


def load_limit_announcements(
    codes: set[str],
    cache_path: Path,
    *,
    skip: bool = False,
    refresh: bool = False,
    workers: int = 8,
) -> tuple[dict[str, dict[str, Any]], str]:
    if skip:
        return {}, "已跳过限额公告解析"
    cached_rows: dict[str, dict[str, Any]] = {}
    cached_errors: dict[str, str] = {}
    cache_is_fresh = False
    if cache_path.exists() and not refresh:
        try:
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            cached_codes = set((cached.get("rows") or {}).keys())
            cached_rows = cached.get("rows") or {}
            cached_errors = cached.get("errors") or {}
            generated_at = datetime.fromisoformat(str(cached.get("generatedAt")).replace("Z", "+00:00"))
            cache_is_fresh = (datetime.now().astimezone() - generated_at).total_seconds() < 24 * 3600
            if codes.issubset(cached_codes) and not cached_errors:
                source = "限额公告解析缓存"
                if not cache_is_fresh:
                    source += "（超过24小时，未强制刷新）"
                return cached_rows, source
        except Exception:
            cached_rows = {}
            cached_errors = {}
    rows: dict[str, dict[str, Any]] = {
        code: row for code, row in cached_rows.items() if code in codes and isinstance(row, dict) and row
    }
    errors: dict[str, str] = {code: msg for code, msg in cached_errors.items() if code in codes}
    fetch_codes = set(codes)
    if cache_is_fresh and not refresh:
        fetch_codes = codes - set(rows) - set(errors)
    with ThreadPoolExecutor(max_workers=max(1, min(workers, 4))) as pool:
        futures = {pool.submit(fetch_limit_announcement_for_code, code): code for code in sorted(fetch_codes)}
        for future in as_completed(futures):
            code = futures[future]
            try:
                got_code, row = future.result()
                rows[got_code] = row
            except Exception as exc:
                errors[code] = str(exc)
    payload = {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "source": "东方财富-天天基金基金公告 type=5 + 公告正文",
        "rows": rows,
        "errors": errors,
    }
    cache_path.parent.mkdir(parents=True, exist_ok=True)
    if len(errors) < len(codes):
        cache_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    source = "东方财富/天天基金限额公告解析"
    if errors:
        source += f"（{len(errors)} 只失败）"
    return rows, source


def merged_announcement_limit(announcement: dict[str, Any] | None) -> dict[str, Any]:
    announcement = announcement or {}
    general = announcement.get("general") or {}
    status = announcement.get("status") or {}
    personal = announcement.get("personal") or {}
    institution = announcement.get("institution") or {}
    product = announcement.get("product") or {}
    def date_key(item: dict[str, Any]) -> str:
        return str((item or {}).get("公告日期") or (item or {}).get("生效日期") or "")

    def choose_limit(specific: dict[str, Any], specific_key: str) -> tuple[dict[str, Any] | None, str]:
        if specific and specific.get(specific_key):
            if general and general.get("通用日限额") and date_key(general) > date_key(specific):
                return general.get("通用日限额"), "general"
            return specific.get(specific_key), "explicit"
        if general and general.get("通用日限额"):
            return general.get("通用日限额"), "general"
        return None, ""

    def choose_source(*items: dict[str, Any]) -> dict[str, Any]:
        valid = [item for item in items if item]
        if not valid:
            return {}
        return sorted(valid, key=date_key, reverse=True)[0]

    source = choose_source(general, personal, institution, product, status)
    personal_limit, personal_source_type = choose_limit(personal, "个人日限额")
    institution_limit, institution_source_type = choose_limit(institution, "非个人机构日限额")
    return {
        "status": (status or general).get("公告状态") or "",
        "personal": personal_limit,
        "institution": institution_limit,
        "personalSourceType": personal_source_type,
        "institutionSourceType": institution_source_type,
        "product": product.get("产品户日限额"),
        "general": general.get("通用日限额"),
        "businessScope": source.get("业务范围") or "",
        "effectiveDate": source.get("生效日期") or "",
        "caliber": source.get("限额口径") or "",
        "sourceTitle": source.get("公告标题") or "",
        "sourceDate": source.get("公告日期") or "",
        "sourceUrl": source.get("公告URL") or "",
        "parseStatus": "parsed" if any([general, personal, institution, product, status]) else "",
        "parseNote": "；".join(
            part
            for part in [
                (personal or {}).get("解析说明"),
                (institution or {}).get("解析说明"),
                (product or {}).get("解析说明"),
                (general or {}).get("解析说明"),
            ]
            if part
        ),
    }


def limit_status(
    row: dict[str, Any],
    purchase: dict[str, Any] | None = None,
    announcement: dict[str, Any] | None = None,
) -> tuple[str, str, dict[str, Any]]:
    ann = merged_announcement_limit(announcement)
    if purchase:
        buy_state = str(purchase.get("申购状态") or "").strip()
        daily_limit = purchase.get("日累计限定金额")
        min_buy = purchase.get("购买起点")
        fee = purchase.get("手续费")
        parts = []
        if buy_state:
            parts.append(buy_state)
        parts.append(f"日限额 {money_text(daily_limit)}")
        if as_float(min_buy) is not None:
            parts.append(f"起点 {money_text(min_buy)}")
        if as_float(fee) is not None:
            parts.append(f"费率 {as_float(fee):.2f}%")
        level = "known"
        if any(token in buy_state for token in ("封闭", "不可", "停止")):
            level = "risk"
        elif "暂停" in buy_state:
            level = "known"
        elif any(
            (item or {}).get("amount") is not None and (item or {}).get("currency") == "CNY" and float((item or {}).get("amount") or 0) < 10_000
            for item in [ann.get("personal"), ann.get("institution"), ann.get("product")]
        ):
            level = "risk"
        elif (as_float(daily_limit) or 0) > 0 and (as_float(daily_limit) or 0) < 10_000:
            level = "risk"
        return "；".join(parts), level, ann
    status = str(row.get("fund_status") or "").strip()
    if any(token in status for token in ("暂停", "限购", "限制", "封闭")):
        return status, "risk", ann
    if status and status.lower() not in {"active", "normal"}:
        return status, "known", ann
    if status:
        return "基础状态正常；销售端状态待核验", "unknown", ann
    return "限额待核验", "unknown", ann


def load_purchase_status(codes: set[str], cache_path: Path, skip: bool = False) -> tuple[dict[str, dict[str, Any]], str]:
    if skip:
        return {}, "已跳过天天申购状态补充"
    try:
        import akshare as ak  # type: ignore

        df = ak.fund_purchase_em()
        if "基金代码" not in df.columns:
            raise RuntimeError("fund_purchase_em 缺少基金代码列")
        rows: dict[str, dict[str, Any]] = {}
        for item in df.to_dict(orient="records"):
            code = str(item.get("基金代码") or "").zfill(6)
            if code in codes:
                rows[code] = {
                    key: (None if str(value) == "NaT" else value)
                    for key, value in item.items()
                    if key in {"基金代码", "基金简称", "申购状态", "赎回状态", "下一开放日", "购买起点", "日累计限定金额", "手续费"}
                }
        cache_path.parent.mkdir(parents=True, exist_ok=True)
        cache_path.write_text(
            json.dumps({"generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"), "rows": rows}, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )
        return rows, "天天基金申购状态"
    except Exception as exc:
        if cache_path.exists():
            cached = json.loads(cache_path.read_text(encoding="utf-8"))
            rows = cached.get("rows") or {}
            return rows, f"天天基金申购状态缓存（接口失败：{exc}）"
        return {}, f"未采集到限额（接口失败：{exc}）"


def load_current_holding_rows(conn: sqlite3.Connection) -> list[dict[str, Any]]:
    sql = f"""
    WITH latest_expo AS (
      SELECT e.*
      FROM {q('基金经济暴露快照')} e
      JOIN (
        SELECT {q('基金代码')} AS code, MAX({q('报告期')}) AS max_report
        FROM {q('基金经济暴露快照')}
        GROUP BY {q('基金代码')}
      ) x
      ON x.code = e.{q('基金代码')} AND x.max_report = e.{q('报告期')}
    )
    SELECT
      h.{q('统一策略ID')} AS strategy_id,
      h.{q('渠道ID')} AS channel_id,
      h.{q('渠道策略ID')} AS channel_strategy_id,
      h.{q('持仓日期')} AS holding_date,
      h.{q('披露日期')} AS disclosure_date,
      h.{q('基金代码')} AS fund_code,
      COALESCE(d.{q('标准基金名称')}, h.{q('基金名称')}, f.{q('基金名称')}, e.{q('基金名称')}) AS fund_name,
      h.{q('基金权重_百分比')} AS weight,
      s.{q('策略名称')} AS strategy_name,
      s.{q('投顾机构')} AS advisor,
      s.{q('策略类型')} AS strategy_type,
      g.{q('治理状态')} AS governance_status,
      g.{q('分析分组')} AS analysis_group,
      g.{q('是否测试组合')} AS is_test,
      g.{q('是否信号类组合')} AS is_signal,
      g.{q('是否目标盈期次')} AS is_target,
      g.{q('是否已停止')} AS is_stopped,
      g.{q('是否纳入常规排名')} AS include_regular_rank,
      d.{q('基金公司')} AS fund_company,
      d.{q('天天基金二级分类')} AS fund_type2,
      d.{q('市场地域标签')} AS region,
      d.{q('是否QDII')} AS is_qdii,
      d.{q('是否ETF')} AS is_etf,
      d.{q('是否ETF联接')} AS is_etf_link,
      d.{q('是否FOF')} AS is_fof,
      d.{q('标准资产细类')} AS standard_sub_asset,
      d.{q('投顾资产分类桶')} AS advisor_asset_bucket,
      d.{q('主题标签JSON')} AS dict_themes,
      f.{q('基金状态')} AS fund_status,
      e.{q('标准资产大类')} AS expo_asset,
      e.{q('标准资产细类')} AS expo_sub_asset,
      e.{q('主题标签JSON')} AS expo_themes,
      e.{q('质量状态')} AS exposure_quality
    FROM {q('策略当前持仓')} h
    LEFT JOIN {q('策略信息')} s ON s.{q('统一策略ID')} = h.{q('统一策略ID')}
    LEFT JOIN {q('策略治理标签')} g ON g.{q('统一策略ID')} = h.{q('统一策略ID')}
    LEFT JOIN {q('基金标准分类字典')} d ON d.{q('基金代码')} = h.{q('基金代码')}
    LEFT JOIN {q('基金信息')} f ON f.{q('基金代码')} = h.{q('基金代码')}
    LEFT JOIN latest_expo e ON e.{q('基金代码')} = h.{q('基金代码')}
    WHERE COALESCE(h.{q('基金权重_百分比')}, 0) > 0
    """
    return [dict(row) for row in conn.execute(sql)]


def valid_current_row(row: dict[str, Any]) -> bool:
    if as_bool(row.get("is_test")):
        return False
    if as_bool(row.get("is_signal")):
        return False
    if as_bool(row.get("is_stopped")):
        return False
    return as_bool(row.get("include_regular_rank")) or as_bool(row.get("is_target"))


def build_units(rows: list[dict[str, Any]]) -> tuple[dict[str, dict[str, Any]], float]:
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
    meta: dict[str, dict[str, Any]] = {}
    for row in rows:
        if not valid_current_row(row):
            continue
        key, name, unit_type = strategy_unit(row)
        grouped[key].append(row)
        meta.setdefault(
            key,
            {
                "unitId": key,
                "unitName": name,
                "unitType": unit_type,
                "strategyType": row.get("strategy_type") or row.get("analysis_group") or unit_type,
                "advisor": row.get("advisor") or "未披露机构",
                "strategyIds": set(),
            },
        )
        if row.get("strategy_id"):
            meta[key]["strategyIds"].add(str(row["strategy_id"]))

    units: dict[str, dict[str, Any]] = {}
    total_weight = 0.0
    for key, unit_rows in grouped.items():
        latest_date = max(str(row.get("holding_date") or "") for row in unit_rows)
        latest_rows = [row for row in unit_rows if str(row.get("holding_date") or "") == latest_date]
        by_sid: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
        sid_meta: dict[str, dict[str, Any]] = {}
        fund_meta: dict[str, dict[str, Any]] = {}
        for row in latest_rows:
            sid = str(row.get("strategy_id") or key)
            code = str(row.get("fund_code") or "").strip()
            weight = as_float(row.get("weight")) or 0.0
            if not code or weight <= 0:
                continue
            by_sid[sid][code] += weight
            sid_meta.setdefault(sid, row)
            fund_meta.setdefault(code, row)
        fund_weights: dict[str, float] = {}
        sids = list(by_sid)
        for code in sorted(fund_meta):
            values = [funds.get(code, 0.0) for funds in by_sid.values()]
            if as_bool(latest_rows[0].get("is_target")) and values:
                value = median(values)
            else:
                value = sum(values)
            if value > 0:
                fund_weights[code] = float(value)
        unit_total = sum(fund_weights.values())
        total_weight += unit_total
        units[key] = {
            **meta[key],
            "strategyIds": sorted(meta[key]["strategyIds"]),
            "holdingDate": latest_date,
            "fundWeights": fund_weights,
            "fundMeta": fund_meta,
            "unitTotalWeight": unit_total,
            "sourceStrategyCount": len(sids),
        }
    return units, total_weight


def load_recent_rebalance(conn: sqlite3.Connection, qd_codes: set[str], latest_date: str | None) -> dict[str, dict[str, Any]]:
    if not qd_codes:
        return {}
    if not latest_date:
        row = conn.execute(f"SELECT MAX({q('调仓日期')}) AS d FROM {q('策略调仓明细')}").fetchone()
        latest_date = row["d"] if row else None
    if not latest_date:
        return {}
    start = (datetime.strptime(str(latest_date)[:10], "%Y-%m-%d") - timedelta(days=30)).date().isoformat()
    placeholders = ",".join("?" for _ in qd_codes)
    sql = f"""
    SELECT
      d.{q('统一策略ID')} AS strategy_id,
      d.{q('调仓日期')} AS rebalance_date,
      d.{q('基金代码')} AS fund_code,
      d.{q('基金名称')} AS fund_name,
      d.{q('调前权重_百分比')} AS before_weight,
      d.{q('调后权重_百分比')} AS after_weight,
      d.{q('权重变化_百分比')} AS weight_change,
      e.{q('调仓标题')} AS event_title,
      e.{q('调仓原因')} AS event_reason,
      s.{q('策略名称')} AS strategy_name,
      s.{q('投顾机构')} AS advisor,
      g.{q('治理状态')} AS governance_status,
      g.{q('是否测试组合')} AS is_test,
      g.{q('是否信号类组合')} AS is_signal,
      g.{q('是否目标盈期次')} AS is_target,
      g.{q('是否已停止')} AS is_stopped,
      g.{q('是否纳入常规排名')} AS include_regular_rank
    FROM {q('策略调仓明细')} d
    LEFT JOIN {q('策略调仓事件')} e ON e.{q('调仓事件ID')} = d.{q('调仓事件ID')}
    LEFT JOIN {q('策略信息')} s ON s.{q('统一策略ID')} = d.{q('统一策略ID')}
    LEFT JOIN {q('策略治理标签')} g ON g.{q('统一策略ID')} = d.{q('统一策略ID')}
    WHERE d.{q('基金代码')} IN ({placeholders})
      AND d.{q('调仓日期')} >= ?
    """
    rows = [dict(row) for row in conn.execute(sql, [*sorted(qd_codes), start])]
    per_unit_date: dict[tuple[str, str, str], list[float]] = defaultdict(list)
    examples: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in rows:
        if not valid_current_row(row):
            continue
        key, unit_name, unit_type = strategy_unit(row)
        code = str(row.get("fund_code") or "")
        before = as_float(row.get("before_weight")) or 0.0
        after = as_float(row.get("after_weight")) or 0.0
        change = as_float(row.get("weight_change"))
        if change is None:
            change = after - before
        per_unit_date[(key, code, str(row.get("rebalance_date") or ""))].append(change)
        if len(examples[code]) < 5 and abs(change) >= 0.5:
            examples[code].append(
                {
                    "strategyName": unit_name,
                    "advisor": row.get("advisor") or "未披露机构",
                    "date": row.get("rebalance_date") or "",
                    "change": round(change, 2),
                    "reason": compact_text(row.get("event_reason") or row.get("event_title"), 64),
                }
            )

    out: dict[str, dict[str, Any]] = {}
    touched_units: dict[str, set[str]] = defaultdict(set)
    for (unit, code, _date), values in per_unit_date.items():
        if not values:
            continue
        change = float(median(values))
        item = out.setdefault(code, {"netChange": 0.0, "addUnits": 0, "cutUnits": 0, "changedUnits": 0, "examples": []})
        item["netChange"] += change
        touched_units[code].add(unit)
        if change > 0:
            item["addUnits"] += 1
        elif change < 0:
            item["cutUnits"] += 1
    for code, units in touched_units.items():
        out[code]["changedUnits"] = len(units)
        out[code]["netChange"] = round(out[code]["netChange"], 2)
        out[code]["examples"] = examples.get(code, [])
    return out


def load_fund_returns(conn: sqlite3.Connection, codes: set[str]) -> dict[str, dict[str, Any]]:
    if not codes:
        return {}
    placeholders = ",".join("?" for _ in codes)
    sql = f"""
    SELECT
      {q('基金代码')} AS code,
      {q('交易日期')} AS trade_date,
      COALESCE({q('累计净值')}, {q('单位净值')}) AS nav
    FROM {q('基金日度净值')}
    WHERE {q('基金代码')} IN ({placeholders})
      AND COALESCE({q('累计净值')}, {q('单位净值')}) IS NOT NULL
    ORDER BY {q('基金代码')}, {q('交易日期')}
    """
    rows_by_code: dict[str, list[tuple[str, float]]] = defaultdict(list)
    for row in conn.execute(sql, sorted(codes)):
        nav = as_float(row["nav"])
        if nav and nav > 0:
            rows_by_code[str(row["code"])].append((str(row["trade_date"])[:10], nav))
    out: dict[str, dict[str, Any]] = {}
    for code, series in rows_by_code.items():
        if not series:
            continue
        latest_date, latest_nav = series[-1]
        latest_day = datetime.strptime(latest_date, "%Y-%m-%d").date()

        def value_on_or_before(target) -> float | None:
            base = None
            for trade_date, nav in series:
                if datetime.strptime(trade_date, "%Y-%m-%d").date() <= target:
                    base = nav
                else:
                    break
            return base

        def value_on_or_after(target) -> float | None:
            for trade_date, nav in series:
                if datetime.strptime(trade_date, "%Y-%m-%d").date() >= target:
                    return nav
            return None

        def calc_return(base: float | None) -> float | None:
            if not base:
                return None
            return (latest_nav / base - 1) * 100

        def ret(days: int) -> float | None:
            return calc_return(value_on_or_before(latest_day - timedelta(days=days)))

        ytd_start = latest_day.replace(month=1, day=1)
        ytd_base = value_on_or_before(ytd_start) or value_on_or_after(ytd_start)
        one_year_start = latest_day - timedelta(days=365)
        one_year_series = [
            (datetime.strptime(trade_date, "%Y-%m-%d").date(), nav)
            for trade_date, nav in series
            if datetime.strptime(trade_date, "%Y-%m-%d").date() >= one_year_start
        ]
        max_drawdown = None
        if one_year_series:
            peak = one_year_series[0][1]
            drawdowns = []
            for _trade_date, nav in one_year_series:
                peak = max(peak, nav)
                drawdowns.append((nav / peak - 1) * 100 if peak else 0)
            max_drawdown = min(drawdowns) if drawdowns else None
        volatility = None
        sharpe_ratio = None
        daily_returns: list[float] = []
        if len(one_year_series) >= 12:
            daily_returns = [
                one_year_series[i][1] / one_year_series[i - 1][1] - 1
                for i in range(1, len(one_year_series))
                if one_year_series[i - 1][1] > 0
            ]
            if len(daily_returns) >= 10:
                daily_std = stdev(daily_returns)
                volatility = daily_std * math.sqrt(252) * 100
                if daily_std > 0:
                    sharpe_ratio = (sum(daily_returns) / len(daily_returns)) / daily_std * math.sqrt(252)

        out[code] = {
            "latestNavDate": latest_date,
            "return1w": round(ret(7), 2) if ret(7) is not None else None,
            "return1m": round(ret(30), 2) if ret(30) is not None else None,
            "return3m": round(ret(90), 2) if ret(90) is not None else None,
            "return6m": round(ret(183), 2) if ret(183) is not None else None,
            "return1y": round(ret(365), 2) if ret(365) is not None else None,
            "returnYtd": round(calc_return(ytd_base), 2) if calc_return(ytd_base) is not None else None,
            "maxDrawdown1y": round(max_drawdown, 2) if max_drawdown is not None else None,
            "volatility1y": round(volatility, 2) if volatility is not None else None,
            "sharpeRatio1y": round(sharpe_ratio, 2) if sharpe_ratio is not None else None,
            "navSampleCount1y": len(one_year_series),
        }
    return out


def aggregate(
    units: dict[str, dict[str, Any]],
    total_weight: float,
    rebalance: dict[str, dict[str, Any]],
    purchase_status: dict[str, dict[str, Any]],
    limit_announcements: dict[str, dict[str, Any]],
    fund_returns: dict[str, dict[str, Any]],
) -> dict[str, Any]:
    funds: dict[str, dict[str, Any]] = {}
    direction_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"funds": set(), "units": set(), "totalWeight": 0.0, "netChange": 0.0})
    for unit in units.values():
        for code, weight in unit["fundWeights"].items():
            row = unit["fundMeta"].get(code) or {}
            if not is_qd_fund(row):
                continue
            group_key = qd_fund_group_key(row)
            direction = qd_direction(row)
            item = funds.setdefault(
                group_key,
                {
                    "groupKey": group_key,
                    "baseName": normalize_qd_fund_base_name(str(row.get("fund_name") or code)),
                    "company": row.get("fund_company") or "未披露",
                    "direction": direction,
                    "fundType": row.get("fund_type2") or row.get("standard_sub_asset") or row.get("expo_sub_asset") or "",
                    "isQdii": bool(as_bool(row.get("is_qdii"))),
                    "themes": theme_labels(row.get("expo_themes") or row.get("dict_themes")),
                    "exposureQuality": row.get("exposure_quality") or "",
                    "shareRows": {},
                    "shareNames": {},
                    "shareWeights": defaultdict(float),
                    "unitRows": {},
                    "unitIds": set(),
                    "advisors": set(),
                    "totalWeight": 0.0,
                    "maxWeight": 0.0,
                },
            )
            item["shareRows"][code] = row
            item["shareNames"][code] = row.get("fund_name") or code
            item["shareWeights"][code] += weight
            item["totalWeight"] += weight
            item["maxWeight"] = max(item["maxWeight"], weight)
            item["unitIds"].add(unit["unitId"])
            item["advisors"].add(unit["advisor"])
            unit_entry = item["unitRows"].setdefault(
                unit["unitId"],
                {
                    "unitId": unit["unitId"],
                    "strategyName": unit["unitName"],
                    "strategyIds": unit["strategyIds"],
                    "advisor": unit["advisor"],
                    "unitType": unit["unitType"],
                    "holdingDate": unit["holdingDate"],
                    "weight": 0.0,
                    "shareCodes": set(),
                    "shareNames": set(),
                    "sourceStrategyCount": unit["sourceStrategyCount"],
                },
            )
            unit_entry["weight"] += weight
            unit_entry["shareCodes"].add(code)
            unit_entry["shareNames"].add(row.get("fund_name") or code)
            direction_totals[direction]["funds"].add(group_key)
            direction_totals[direction]["units"].add(unit["unitId"])
            direction_totals[direction]["totalWeight"] += weight

    fund_rows: list[dict[str, Any]] = []
    for group_key, item in funds.items():
        share_codes = sorted(item["shareRows"])
        main_code = min(
            share_codes,
            key=lambda c: qd_main_share_score(c, item["shareRows"][c], purchase_status.get(c), limit_announcements.get(c), fund_returns.get(c)),
        )
        main_row = item["shareRows"][main_code]
        main_purchase = purchase_status.get(main_code) or {}
        main_announcement = limit_announcements.get(main_code) or {}
        main_return = fund_returns.get(main_code) or {}
        limit_text, limit_level, limit_detail = limit_status(main_row, main_purchase, main_announcement)
        unit_count = len(item["unitIds"])
        advisor_count = len(item["advisors"])
        recent = {"netChange": 0.0, "addUnits": 0, "cutUnits": 0, "changedUnits": 0, "examples": []}
        seen_add: set[str] = set()
        seen_cut: set[str] = set()
        seen_changed: set[str] = set()
        for share_code in share_codes:
            share_recent = rebalance.get(share_code) or {}
            recent["netChange"] += as_float(share_recent.get("netChange")) or 0.0
            for example in share_recent.get("examples", []) or []:
                recent["examples"].append({**example, "shareCode": share_code})
            for key, target in [("addUnits", seen_add), ("cutUnits", seen_cut), ("changedUnits", seen_changed)]:
                if share_recent.get(key):
                    for example in share_recent.get("examples", []) or []:
                        unit_id = example.get("unitId") or f"{example.get('advisor')}::{example.get('strategyName')}"
                        target.add(str(unit_id))
        recent["addUnits"] = len(seen_add)
        recent["cutUnits"] = len(seen_cut)
        recent["changedUnits"] = len(seen_changed)
        recent["examples"] = sorted(recent["examples"], key=lambda x: (str(x.get("date") or ""), abs(as_float(x.get("change")) or 0.0)), reverse=True)[:8]
        strategy_units = []
        for row in item["unitRows"].values():
            out_row = dict(row)
            out_row["weight"] = round(as_float(out_row.get("weight")) or 0.0, 2)
            out_row["shareCodes"] = sorted(out_row["shareCodes"])
            out_row["shareNames"] = sorted(out_row["shareNames"])
            out_row["recentChange"] = recent.get("netChange", 0.0)
            strategy_units.append(out_row)
        strategy_units.sort(key=lambda x: (-x["weight"], x["advisor"], x["strategyName"]))
        total_ratio = item["totalWeight"] / total_weight * 100 if total_weight else None
        personal_amount = investor_limit_amount(limit_detail.get("personal"), main_purchase)
        non_person_amount = investor_limit_amount(limit_detail.get("institution"), main_purchase)
        row = {
            "code": main_code,
            "name": main_row.get("fund_name") or item["baseName"] or main_code,
            "baseName": item["baseName"],
            "company": item["company"],
            "direction": item["direction"],
            "fundType": item["fundType"],
            "isQdii": item["isQdii"],
            "shareCount": len(share_codes),
            "shareCodes": share_codes,
            "shareNames": [item["shareNames"].get(c) or c for c in share_codes],
            "shareWeights": {c: round(item["shareWeights"].get(c, 0.0), 2) for c in share_codes},
            "limit": limit_text,
            "limitLevel": limit_level,
            "purchaseStatus": main_purchase,
            "limitDetail": public_limit_detail(limit_detail),
            "personalLimit": investor_limit_text(limit_detail.get("personal"), main_purchase, str(limit_detail.get("personalSourceType") or "")),
            "nonPersonLimit": investor_limit_text(limit_detail.get("institution"), main_purchase, str(limit_detail.get("institutionSourceType") or "")),
            "personalLimitSource": limit_detail.get("personalSourceType") or ("sales" if main_purchase else ""),
            "nonPersonLimitSource": limit_detail.get("institutionSourceType") or ("sales" if main_purchase else ""),
            "personalLimitAmount": personal_amount,
            "nonPersonLimitAmount": non_person_amount,
            "salesDailyLimitAmount": sales_daily_limit_amount(main_purchase),
            "limitSourceDate": limit_detail.get("sourceDate") or "",
            "limitSourceTitle": limit_detail.get("sourceTitle") or "",
            "limitSourceUrl": limit_detail.get("sourceUrl") or "",
            "limitParseStatus": limit_detail.get("parseStatus") or "",
            "return1w": main_return.get("return1w"),
            "return1m": main_return.get("return1m"),
            "return3m": main_return.get("return3m"),
            "return6m": main_return.get("return6m"),
            "return1y": main_return.get("return1y"),
            "returnYtd": main_return.get("returnYtd"),
            "maxDrawdown1y": main_return.get("maxDrawdown1y"),
            "volatility1y": main_return.get("volatility1y"),
            "sharpeRatio1y": main_return.get("sharpeRatio1y"),
            "navSampleCount1y": main_return.get("navSampleCount1y"),
            "latestNavDate": main_return.get("latestNavDate"),
            "themes": item["themes"],
            "exposureQuality": item["exposureQuality"],
            "strategyCount": unit_count,
            "advisorCount": advisor_count,
            "totalWeight": round(item["totalWeight"], 2),
            "totalWeightRatio": round(total_ratio, 4) if total_ratio is not None else None,
            "averageWeight": round(item["totalWeight"] / unit_count, 2) if unit_count else None,
            "maxWeight": round(item["maxWeight"], 2),
            "recentNetInflow": recent.get("netChange", 0.0),
            "recentAddStrategies": recent.get("addUnits", 0),
            "recentCutStrategies": recent.get("cutUnits", 0),
            "recentChangedStrategies": recent.get("changedUnits", 0),
            "recentExamples": recent.get("examples", []),
            "strategyUnits": strategy_units,
        }
        fund_rows.append(row)
        direction_totals[row["direction"]]["netChange"] += row["recentNetInflow"]
    fund_rows.sort(key=lambda x: (-x["strategyCount"], -x["advisorCount"], -x["totalWeight"], x["code"]))

    directions = []
    for direction, item in direction_totals.items():
        directions.append(
            {
                "direction": direction,
                "fundCount": len(item["funds"]),
                "strategyCount": len(item["units"]),
                "totalWeight": round(item["totalWeight"], 2),
                "totalWeightRatio": round(item["totalWeight"] / total_weight * 100, 4) if total_weight else None,
                "recentNetInflow": round(item["netChange"], 2),
            }
        )
    directions.sort(key=lambda x: (-x["strategyCount"], -x["totalWeight"], x["direction"]))
    qd_unit_weights: dict[str, float] = defaultdict(float)
    qd_unit_totals: dict[str, float] = {}
    type_totals: dict[str, dict[str, Any]] = defaultdict(lambda: {"total": 0, "qd": 0, "weights": [], "topDirections": defaultdict(float)})
    for unit in units.values():
        unit_type = unit.get("strategyType") or unit.get("unitType") or "未分类"
        type_totals[unit_type]["total"] += 1
        qd_weight = 0.0
        for code, weight in unit["fundWeights"].items():
            row = unit["fundMeta"].get(code) or {}
            if is_qd_fund(row):
                qd_weight += weight
                type_totals[unit_type]["topDirections"][qd_direction(row)] += weight
        if qd_weight > 0:
            qd_unit_weights[unit["unitId"]] = qd_weight
            qd_unit_totals[unit["unitId"]] = unit.get("unitTotalWeight") or 100.0
            type_totals[unit_type]["qd"] += 1
            type_totals[unit_type]["weights"].append(qd_weight / max(unit.get("unitTotalWeight") or 100.0, 1e-9) * 100)
    type_rows = []
    for unit_type, item in type_totals.items():
        if not item["qd"]:
            continue
        top_direction = ""
        if item["topDirections"]:
            top_direction = max(item["topDirections"].items(), key=lambda x: x[1])[0]
        type_rows.append(
            {
                "strategyType": unit_type,
                "totalStrategyCount": item["total"],
                "qdStrategyCount": item["qd"],
                "penetration": round(item["qd"] / item["total"] * 100, 2) if item["total"] else None,
                "medianQdWeight": round(median(item["weights"]), 2) if item["weights"] else None,
                "averageQdWeight": round(sum(item["weights"]) / len(item["weights"]), 2) if item["weights"] else None,
                "topDirection": top_direction,
            }
        )
    type_rows.sort(key=lambda x: (-x["qdStrategyCount"], -(x["medianQdWeight"] or 0), x["strategyType"]))

    limit_risk = [row for row in fund_rows if row["limitLevel"] == "risk"]
    unknown_limit = [row for row in fund_rows if row["limitLevel"] == "unknown"]
    announcement_count = sum(1 for row in fund_rows if row.get("limitSourceTitle"))
    personal_limit_count = sum(1 for row in fund_rows if not str(row.get("personalLimit") or "").startswith("销售端未采集"))
    non_person_limit_count = sum(1 for row in fund_rows if not str(row.get("nonPersonLimit") or "").startswith("销售端未采集"))
    qd_weight_shares = [
        qd_unit_weights[unit_id] / max(qd_unit_totals.get(unit_id) or 100.0, 1e-9) * 100 for unit_id in qd_unit_weights
    ]
    return {
        "generatedAt": datetime.now().astimezone().isoformat(timespec="seconds"),
        "funds": fund_rows,
        "directions": directions,
        "strategyTypeRows": type_rows,
        "summary": {
            "fundCount": len(fund_rows),
            "strategyCount": len({unit_id for row in fund_rows for unit_id in [u["unitId"] for u in row["strategyUnits"]]}),
            "advisorCount": len({u["advisor"] for row in fund_rows for u in row["strategyUnits"]}),
            "totalWeightRatio": round(sum(row["totalWeight"] for row in fund_rows) / total_weight * 100, 4) if total_weight else None,
            "penetration": round(len(qd_unit_weights) / len(units) * 100, 2) if units else None,
            "medianQdWeightInQdStrategies": round(median(qd_weight_shares), 2) if qd_weight_shares else None,
            "averageQdWeightInQdStrategies": round(sum(qd_weight_shares) / len(qd_weight_shares), 2) if qd_weight_shares else None,
            "limitRiskCount": len(limit_risk),
            "limitUnknownCount": len(unknown_limit),
            "limitAnnouncementCount": announcement_count,
            "personalLimitKnownCount": personal_limit_count,
            "nonPersonLimitKnownCount": non_person_limit_count,
            "totalComparableUnits": len(units),
            "totalComparableWeight": round(total_weight, 2),
        },
    }


def persist_limit_snapshot(conn: sqlite3.Connection, data: dict[str, Any]) -> None:
    conn.execute(f"DROP TABLE IF EXISTS {q('QD基金限额快照')}")
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {q('QD基金限额快照')} (
          {q('基金代码')} TEXT PRIMARY KEY,
          {q('基金名称')} TEXT,
          {q('主基金代码')} TEXT,
          {q('合并份额代码')} TEXT,
          {q('基金公司')} TEXT,
          {q('QD方向')} TEXT,
          {q('申购状态')} TEXT,
          {q('销售端日限额')} TEXT,
          {q('个人日限额')} TEXT,
          {q('非个人日限额')} TEXT,
          {q('限额口径')} TEXT,
          {q('业务范围')} TEXT,
          {q('生效日期')} TEXT,
          {q('公告日期')} TEXT,
          {q('公告标题')} TEXT,
          {q('公告URL')} TEXT,
          {q('解析状态')} TEXT,
          {q('解析说明')} TEXT,
          {q('近1周收益率_百分比')} REAL,
          {q('近1月收益率_百分比')} REAL,
          {q('近3月收益率_百分比')} REAL,
          {q('近6月收益率_百分比')} REAL,
          {q('近1年收益率_百分比')} REAL,
          {q('今年以来收益率_百分比')} REAL,
          {q('近1年最大回撤_百分比')} REAL,
          {q('近1年年化波动率_百分比')} REAL,
          {q('近1年夏普')} REAL,
          {q('最新净值日期')} TEXT,
          {q('生成时间')} TEXT
        )
        """
    )
    now = datetime.now().astimezone().isoformat(timespec="seconds")
    rows = []
    for fund in data.get("funds", []):
        detail = fund.get("limitDetail") or {}
        rows.append(
            (
                fund.get("code"),
                fund.get("name"),
                fund.get("code"),
                ",".join(fund.get("shareCodes") or []),
                fund.get("company"),
                fund.get("direction"),
                fund.get("limit"),
                money_text((fund.get("purchaseStatus") or {}).get("日累计限定金额")),
                fund.get("personalLimit"),
                fund.get("nonPersonLimit"),
                detail.get("caliber"),
                detail.get("businessScope"),
                detail.get("effectiveDate"),
                fund.get("limitSourceDate"),
                fund.get("limitSourceTitle"),
                fund.get("limitSourceUrl"),
                fund.get("limitParseStatus") or "sales_status_only",
                detail.get("parseNote"),
                fund.get("return1w"),
                fund.get("return1m"),
                fund.get("return3m"),
                fund.get("return6m"),
                fund.get("return1y"),
                fund.get("returnYtd"),
                fund.get("maxDrawdown1y"),
                fund.get("volatility1y"),
                fund.get("sharpeRatio1y"),
                fund.get("latestNavDate"),
                now,
            )
        )
    conn.executemany(
        f"""
        INSERT OR REPLACE INTO {q('QD基金限额快照')} (
          {q('基金代码')}, {q('基金名称')}, {q('主基金代码')}, {q('合并份额代码')},
          {q('基金公司')}, {q('QD方向')}, {q('申购状态')},
          {q('销售端日限额')}, {q('个人日限额')}, {q('非个人日限额')},
          {q('限额口径')}, {q('业务范围')}, {q('生效日期')}, {q('公告日期')}, {q('公告标题')},
          {q('公告URL')}, {q('解析状态')}, {q('解析说明')}, {q('近1周收益率_百分比')},
          {q('近1月收益率_百分比')}, {q('近3月收益率_百分比')}, {q('近6月收益率_百分比')},
          {q('近1年收益率_百分比')}, {q('今年以来收益率_百分比')},
          {q('近1年最大回撤_百分比')}, {q('近1年年化波动率_百分比')},
          {q('近1年夏普')},
          {q('最新净值日期')}, {q('生成时间')}
        ) VALUES ({",".join("?" for _ in range(29))})
        """,
        rows,
    )
    conn.commit()


FIELD_DESCRIPTIONS = {
    "QD基金数量": "当前持仓中被识别为 QDII、海外 QDII 或海外资产工具的底层基金数量。同一基金的 A/C/F 等份额合并计为一只。",
    "配置策略数": "配置了至少一只 QD 基金的有效策略单元数量。目标盈系列按同机构同系列合并，同一底层基金的多份额合并后计数。",
    "策略渗透率": "配置了 QD 基金的有效策略单元数 / 全部有效可比策略单元数。测试、信号、已停止等不可比策略不进入分母。",
    "QD策略中位仓位": "只在已经配置 QD 基金的策略内计算：每个策略的 QD 基金仓位合计 / 该策略总披露仓位，再取中位数。",
    "限额风险基金": "根据当前销售端日累计限额和公告明确限额判断。销售端日限额过低、限大额或公告明确披露个人/非个人低限额会进入风险池；普通暂停申购仅在申购状态中展示，不覆盖为限额风险。",
    "公告证据覆盖": "能解析到东方财富/天天基金公告标题和日期的 QD 基金数 / QD 基金总数。没有公告证据不等于没有限额，页面会继续使用销售端申购状态和日累计限额作交易可得性参考。",
    "渗透率": "当前行所属策略类型内，配置了 QD 基金的策略单元数 / 该类型全部有效策略单元数。",
    "含QD策略": "当前策略类型内，至少配置一只 QD 基金的有效策略单元数量。",
    "策略数": "配置该底层基金或命中当前条件的有效策略单元数量。目标盈系列和基金多份额均已合并后计数。",
    "QD中位仓位": "当前行所属策略类型内，只统计含 QD 基金策略的 QD 仓位中位数。",
    "QD策略均仓": "只在配置该基金的策略单元内计算，该基金仓位的简单平均值。",
    "个人限额": "优先展示主基金公告正文中明确写给个人投资者的单日限额；若公告只有统一账户限额，则标为通用公告参考；若无公告证据，则回落到销售端日累计限额。",
    "非个人限额": "优先展示主基金公告正文中明确写给非个人/机构投资者的单日限额；若公告只有统一账户限额，则标为通用公告参考，不视作机构专属限额；若无公告证据，则回落到销售端日累计限额。",
    "QD方向": "基于基金名称、QDII标识、基金类型、经济暴露和主题标签归类到美股、港股/中概、海外债、印度、日本、欧洲、商品/黄金等方向。",
    "合并份额": "该底层基金被合并统计的份额代码。仓位、策略数、渗透率和调仓按这些份额汇总；业绩和限额取主基金份额。",
    "配置总仓位比例": "该底层 QD 基金所有份额在有效策略单元中的仓位合计 / 有效策略总披露仓位。用于看市场共识权重，不等同于单只策略仓位。",
    "配置总仓位比例(%)": "该底层 QD 基金所有份额在有效策略单元中的仓位合计 / 有效策略总披露仓位。用于看市场共识权重，不等同于单只策略仓位。",
    "近1月收益": "使用本地基金日度净值，以最新净值相对约 30 天前可得净值计算。",
    "近1周收益率": "使用本地基金日度净值，以最新净值相对约 7 天前可得净值计算。",
    "近1月收益率": "使用本地基金日度净值，以最新净值相对约 30 天前可得净值计算。",
    "近3月收益率": "使用本地基金日度净值，以最新净值相对约 90 天前可得净值计算。",
    "近6月收益率": "使用本地基金日度净值，以最新净值相对约 183 天前可得净值计算。",
    "近1年收益率": "使用本地基金日度净值，以最新净值相对约 365 天前可得净值计算。",
    "今年以来收益率": "使用本地基金日度净值，以最新净值相对当年年初附近可得净值计算。",
    "近1年最大回撤": "使用最近 365 天本地基金日度净值，逐日计算相对历史峰值的回撤，取最低值，负值越大表示回撤越深。",
    "近1年年化波动率": "使用最近 365 天基金日收益率标准差乘以 sqrt(252) 年化，样本过少时为空。",
    "近1年夏普": "使用最近 365 天基金日收益率均值 / 日收益率标准差 * sqrt(252) 计算，暂不扣无风险收益率，样本过少或波动率为 0 时为空。",
    "最大限额": "个人日限额和非个人日限额中较大的有效金额，用于观察产品最大申购空间；若要看真实瓶颈，应同时看保守限额容量。",
    "保守限额容量": "个人日限额和非个人日限额中较小的有效金额，用于衡量对不同客户类型都较稳妥的容量瓶颈。",
    "净调入": "最近调仓窗口中，该基金调入仓位比例合计减去调出仓位比例合计。",
    "近1月净调入": "最近调仓窗口中，涉及该基金的调入比例合计减调出比例合计。目标盈系列按合并策略单元处理。",
}
FIELD_DESCRIPTIONS.update(
    {
        "基金申购状态": "天天基金销售端当前申购状态，单独展示为开放申购、限大额、暂停申购等状态。",
        "个人单日限额": "XLSX 导出专用数值字段，优先取个人日限额金额；暂停申购记为 0，销售端未设置明显限额记为 100000000，便于筛选和排序。",
        "非个人单日限额": "XLSX 导出专用数值字段，优先取非个人/机构日限额金额；暂停申购记为 0，销售端未设置明显限额记为 100000000，便于筛选和排序。",
        "配置机构汇总": "列出配置该 QD 基金的全部投顾机构，并按配置策略数量降序展示为“机构（策略数）”。",
    }
)


def info_label(label: str) -> str:
    note = FIELD_DESCRIPTIONS.get(label)
    if not note:
        return escape(label)
    return (
        f'{escape(label)} <button type="button" class="info-eye" '
        f'data-info-title="{escape(label)}" data-info="{escape(note)}" aria-label="{escape(label)}口径说明"></button>'
    )


def info_script() -> str:
    notes = json.dumps(FIELD_DESCRIPTIONS, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    return f"""
(function(){{
  const notes = {notes};
  window.qdInfoButton = function(label){{
    const note = notes[label] || "";
    if (!note) return label;
    const safeLabel = String(label).replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[s]));
    const safeNote = String(note).replace(/[&<>"']/g, s => ({{'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}}[s]));
    return `${{safeLabel}} <button type="button" class="info-eye" data-info-title="${{safeLabel}}" data-info="${{safeNote}}" aria-label="${{safeLabel}}口径说明"></button>`;
  }};
  function closeInfo(){{
    const old = document.querySelector(".info-modal");
    if (old) old.remove();
  }}
  document.addEventListener("click", function(ev){{
    const btn = ev.target.closest(".info-eye");
    if (!btn) {{
      if (ev.target.classList && ev.target.classList.contains("info-backdrop")) closeInfo();
      return;
    }}
    ev.preventDefault();
    closeInfo();
    const modal = document.createElement("div");
    modal.className = "info-modal";
    modal.innerHTML = `<div class="info-backdrop"></div><div class="info-card" role="dialog" aria-modal="true"><button type="button" class="info-close" aria-label="关闭">×</button><h3>${{btn.dataset.infoTitle || "口径说明"}}</h3><p>${{btn.dataset.info || ""}}</p></div>`;
    document.body.appendChild(modal);
    modal.querySelector(".info-close").addEventListener("click", closeInfo);
  }});
  document.addEventListener("keydown", function(ev){{ if (ev.key === "Escape") closeInfo(); }});
}})();
"""


RETURN_METRIC_KEYS = {"return1w", "return1m", "return3m", "return6m", "return1y", "returnYtd", "maxDrawdown1y"}


def value_class(value: Any) -> str:
    num = as_float(value)
    if num is None or num == 0:
        return "neutral"
    return "pos" if num > 0 else "neg"


def pct_html(value: Any, signed: bool = False) -> str:
    return f'<span class="{value_class(value)}">{escape(pct(as_float(value), signed=signed))}</span>'


def report_cards_html(data: dict[str, Any]) -> str:
    summary = data["summary"]
    cards = [
        ("QD基金数量", f"{summary['fundCount']}", "严格 QDII/海外 QDII 主口径"),
        ("配置策略数", f"{summary['strategyCount']}", "目标盈系列合并后"),
        ("策略渗透率", pct(summary["penetration"]), "含QD策略/有效策略"),
        ("QD策略中位仓位", pct(summary["medianQdWeightInQdStrategies"]), "只在含QD策略内计算"),
        ("限额风险基金", f"{summary['limitRiskCount']}", "低额度/限大额"),
        ("公告证据覆盖", f"{summary['limitAnnouncementCount']}/{summary['fundCount']}", "缺公告时用销售端参考"),
    ]
    return "".join(
        f'<div class="metric"><b>{escape(value)}</b><span>{info_label(label)}</span><small>{escape(note)}</small></div>'
        for label, value, note in cards
    )


def top_table(rows: list[dict[str, Any]], columns: list[tuple[str, str]], limit: int = 8) -> str:
    head = "".join(f"<th>{info_label(label)}</th>" for label, _ in columns)
    body = []
    for row in rows[:limit]:
        cells = []
        for _label, key in columns:
            value = row.get(key)
            if key == "name":
                code = escape(str(row.get("code") or ""))
                name = escape(str(row.get("name") or ""))
                value = f'<a class="fund-link" href="./fund.html?code={code}">{code} {name}</a>'
            elif key in RETURN_METRIC_KEYS or key == "recentNetInflow":
                value = pct_html(value, signed=(key in RETURN_METRIC_KEYS or key == "recentNetInflow"))
            elif key in {"totalWeightRatio", "averageWeight", "medianQdWeight", "penetration"}:
                value = escape(pct(as_float(value)))
            elif value is None:
                value = "--"
            else:
                value = escape(str(value))
            cells.append(f"<td>{value}</td>")
        body.append("<tr>" + "".join(cells) + "</tr>")
    return f"<table><thead><tr>{head}</tr></thead><tbody>{''.join(body)}</tbody></table>"


def bar_chart_html(rows: list[dict[str, Any]], label_key: str, value_key: str, note_keys: list[str] | None = None, limit: int = 8) -> str:
    rows = rows[:limit]
    max_value = max([as_float(row.get(value_key)) or 0 for row in rows] or [0])
    if max_value <= 0:
        return '<div class="empty">暂无可展示数据</div>'
    parts = []
    for row in rows:
        value = as_float(row.get(value_key)) or 0.0
        width = max(3.0, value / max_value * 100)
        notes = []
        for key in note_keys or []:
            got = row.get(key)
            if got is None:
                continue
            if key == "fundCount":
                notes.append(f"{got}只基金")
            elif key == "totalWeightRatio":
                notes.append(f"总仓位{pct(as_float(got))}")
            elif key.endswith("Ratio") or key in {"recentNetInflow", "return1m", "return3m", "medianQdWeight"}:
                notes.append(pct(as_float(got), signed=(key == "recentNetInflow")))
            else:
                notes.append(str(got))
        parts.append(
            f"""<div class="bar-row"><div><b>{escape(str(row.get(label_key) or ""))}</b><span>{escape(" · ".join(notes))}</span></div><div class="bar"><i style="width:{width:.1f}%"></i></div><em>{value:g}</em></div>"""
        )
    return '<div class="bars">' + "".join(parts) + "</div>"


def grouped_bar_chart_html(
    rows: list[dict[str, Any]],
    label_key: str,
    left_key: str,
    right_key: str,
    left_label: str = "个人",
    right_label: str = "非个人",
    limit: int = 8,
) -> str:
    rows = rows[:limit]
    max_value = max(
        [as_float(row.get(left_key)) or 0 for row in rows] + [as_float(row.get(right_key)) or 0 for row in rows] or [0]
    )
    if max_value <= 0:
        return '<div class="empty">暂无可展示数据</div>'
    parts = [
        f'<div class="chart-legend"><span><i class="legend-personal"></i>{escape(left_label)}</span><span><i class="legend-non-person"></i>{escape(right_label)}</span></div>'
    ]
    for row in rows:
        left_value = as_float(row.get(left_key)) or 0.0
        right_value = as_float(row.get(right_key)) or 0.0
        left_width = max(3.0, left_value / max_value * 100) if left_value > 0 else 0.0
        right_width = max(3.0, right_value / max_value * 100) if right_value > 0 else 0.0
        parts.append(
            f"""<div class="grouped-bar-row"><div class="grouped-label"><b>{escape(str(row.get(label_key) or ""))}</b></div><div class="grouped-bars"><div class="bar personal"><i style="width:{left_width:.1f}%"></i></div><div class="bar non-person"><i style="width:{right_width:.1f}%"></i></div></div><div class="grouped-values"><span>{left_value:g}</span><span>{right_value:g}</span></div></div>"""
        )
    return '<div class="grouped-bars-wrap">' + "".join(parts) + "</div>"


def limit_bucket(amount: float | None, level: str | None = None) -> str:
    if level == "risk" or (amount is not None and amount < 10_000):
        return "1万以下/暂停"
    if amount is None or amount >= 100_000_000:
        return "未设日限额"
    if amount >= 1_000_000:
        return "100万以上"
    if amount >= 100_000:
        return "10万-100万"
    return "1万-10万"


def limit_bucket_rows(funds: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = ["未设日限额", "100万以上", "10万-100万", "1万-10万", "1万以下/暂停"]
    buckets: dict[str, dict[str, Any]] = {
        key: {"bucket": key, "personalFundCount": 0, "nonPersonFundCount": 0} for key in order
    }
    for row in funds:
        personal_bucket = limit_bucket(as_float(row.get("personalLimitAmount")), row.get("limitLevel"))
        non_person_bucket = limit_bucket(as_float(row.get("nonPersonLimitAmount")), row.get("limitLevel"))
        buckets[personal_bucket]["personalFundCount"] += 1
        buckets[non_person_bucket]["nonPersonFundCount"] += 1
    return [buckets[key] for key in order if buckets[key]["personalFundCount"] or buckets[key]["nonPersonFundCount"]]


def limit_source_bucket(text: str, source_type: str) -> str:
    if "未采集" in text or "待核验" in text:
        return "待核验"
    if "未设日限额" in text:
        return "未设日限额"
    if source_type == "explicit":
        return "明确公告"
    if source_type == "general":
        return "通用公告参考"
    if source_type == "sales" or text.startswith("销售端"):
        return "销售端参考"
    return "待核验"


def limit_source_rows(funds: list[dict[str, Any]]) -> list[dict[str, Any]]:
    order = ["明确公告", "通用公告参考", "销售端参考", "未设日限额", "待核验"]
    buckets: dict[str, dict[str, Any]] = {
        key: {"bucket": key, "personalFundCount": 0, "nonPersonFundCount": 0} for key in order
    }
    for row in funds:
        personal_bucket = limit_source_bucket(str(row.get("personalLimit") or ""), str(row.get("personalLimitSource") or ""))
        non_person_bucket = limit_source_bucket(str(row.get("nonPersonLimit") or ""), str(row.get("nonPersonLimitSource") or ""))
        buckets[personal_bucket]["personalFundCount"] += 1
        buckets[non_person_bucket]["nonPersonFundCount"] += 1
    return [buckets[key] for key in order if buckets[key]["personalFundCount"] or buckets[key]["nonPersonFundCount"]]


def qd_nav_html(active: str = "qd") -> str:
    return render_system_topbar(active)


def _scatter_chart_html_legacy(data: dict[str, Any]) -> str:
    rows = []
    for row in data.get("funds", []):
        rows.append(
            {
                "code": row.get("code"),
                "name": row.get("name"),
                "company": row.get("company"),
                "direction": row.get("direction"),
                "strategyCount": row.get("strategyCount"),
                "totalWeightRatio": row.get("totalWeightRatio"),
                "return1w": row.get("return1w"),
                "return1m": row.get("return1m"),
                "return3m": row.get("return3m"),
                "return6m": row.get("return6m"),
                "return1y": row.get("return1y"),
                "returnYtd": row.get("returnYtd"),
                "maxDrawdown1y": row.get("maxDrawdown1y"),
                "volatility1y": row.get("volatility1y"),
                "personalLimit": row.get("personalLimit"),
                "nonPersonLimit": row.get("nonPersonLimit"),
            }
        )
    payload = json.dumps(rows, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    return f"""
<div class="scatter-controls">
  <label>纵轴指标
    <select id="qdScatterMetric">
      <option value="return1m">近1月收益率</option>
      <option value="return1w">近1周收益率</option>
      <option value="return3m">近3月收益率</option>
      <option value="return6m">近6月收益率</option>
      <option value="return1y">近1年收益率</option>
      <option value="returnYtd">今年以来收益率</option>
      <option value="maxDrawdown1y">近1年最大回撤</option>
      <option value="volatility1y">近1年年化波动率</option>
    </select>
  </label>
</div>
<div id="qdScatter" class="scatter-chart"></div>
<script>
(function(){{
  const rows = {payload};
  const metricNames = {{
    return1w:"近1周收益率", return1m:"近1月收益率", return3m:"近3月收益率",
    return6m:"近6月收益率", return1y:"近1年收益率", returnYtd:"今年以来收益率",
    maxDrawdown1y:"近1年最大回撤", volatility1y:"近1年年化波动率"
  }};
  const el = document.getElementById("qdScatter");
  const select = document.getElementById("qdScatterMetric");
  const fmt = (v, sign=false) => v == null || Number.isNaN(Number(v)) ? "--" : `${{sign && Number(v)>0 ? "+" : ""}}${{Number(v).toFixed(2)}}%`;
  const cls = (metric, value) => metric.startsWith("return") && Number(value) > 0 ? "pos" : metric.startsWith("return") && Number(value) < 0 ? "neg" : "neutral";
  function render(){{
    const metric = select.value;
    const usable = rows.filter(d => Number.isFinite(Number(d.strategyCount)) && Number.isFinite(Number(d[metric])));
    if (!usable.length) {{
      el.innerHTML = '<div class="empty">当前指标暂无可展示数据</div>';
      return;
    }}
    const w = Math.max(760, el.clientWidth || 760);
    const h = 390;
    const m = {{l:58,r:26,t:26,b:52}};
    const xMax = Math.max(...usable.map(d => Number(d.strategyCount)), 1);
    const yVals = usable.map(d => Number(d[metric]));
    let yMin = Math.min(...yVals), yMax = Math.max(...yVals);
    if (yMin === yMax) {{ yMin -= 1; yMax += 1; }}
    const pad = (yMax - yMin) * 0.12;
    yMin -= pad; yMax += pad;
    const sizeMax = Math.max(...usable.map(d => Number(d.totalWeightRatio) || 0), 0.01);
    const x = v => m.l + (Number(v) / xMax) * (w - m.l - m.r);
    const y = v => m.t + (yMax - Number(v)) / (yMax - yMin) * (h - m.t - m.b);
    const r = v => 4 + Math.sqrt(Math.max(Number(v) || 0, 0) / sizeMax) * 12;
    const xTicks = [0, Math.round(xMax/4), Math.round(xMax/2), Math.round(xMax*3/4), xMax].filter((v,i,a)=>i===0||v!==a[i-1]);
    const yTicks = [yMin, yMin+(yMax-yMin)/4, yMin+(yMax-yMin)/2, yMin+(yMax-yMin)*3/4, yMax];
    const points = usable.map(d => {{
      const cx=x(d.strategyCount), cy=y(d[metric]), rr=r(d.totalWeightRatio);
      const title = `${{d.code}} ${{d.name}}\\n${{metricNames[metric]}}：${{fmt(d[metric], metric.startsWith("return"))}}\\n策略数：${{d.strategyCount}}\\n总仓位占比：${{fmt(d.totalWeightRatio)}}\\n个人：${{d.personalLimit || "--"}}\\n非个人：${{d.nonPersonLimit || "--"}}`;
      return `<circle class="${{cls(metric,d[metric])}}" cx="${{cx.toFixed(1)}}" cy="${{cy.toFixed(1)}}" r="${{rr.toFixed(1)}}" tabindex="0"><title>${{title.replace(/[&<>]/g, s=>({{'&':'&amp;','<':'&lt;','>':'&gt;'}}[s]))}}</title></circle>`;
    }}).join("");
    el.innerHTML = `<svg viewBox="0 0 ${{w}} ${{h}}" role="img" aria-label="QD基金配置点阵图">
      <rect class="plot-bg" x="${{m.l}}" y="${{m.t}}" width="${{w-m.l-m.r}}" height="${{h-m.t-m.b}}"></rect>
      ${{yTicks.map(t=>`<g><line class="grid-line" x1="${{m.l}}" x2="${{w-m.r}}" y1="${{y(t)}}" y2="${{y(t)}}"></line><text class="axis-text" x="${{m.l-10}}" y="${{y(t)+4}}" text-anchor="end">${{fmt(t)}}</text></g>`).join("")}}
      ${{xTicks.map(t=>`<g><line class="grid-line" x1="${{x(t)}}" x2="${{x(t)}}" y1="${{m.t}}" y2="${{h-m.b}}"></line><text class="axis-text" x="${{x(t)}}" y="${{h-20}}" text-anchor="middle">${{t}}</text></g>`).join("")}}
      <line class="axis-line" x1="${{m.l}}" x2="${{w-m.r}}" y1="${{h-m.b}}" y2="${{h-m.b}}"></line>
      <line class="axis-line" x1="${{m.l}}" x2="${{m.l}}" y1="${{m.t}}" y2="${{h-m.b}}"></line>
      <text class="axis-title" x="${{w/2}}" y="${{h-2}}" text-anchor="middle">配置策略数</text>
      <text class="axis-title" transform="translate(16 ${{h/2}}) rotate(-90)" text-anchor="middle">${{metricNames[metric]}}</text>
      ${{points}}
    </svg>`;
  }}
  select.addEventListener("change", render);
  window.addEventListener("resize", render);
  render();
}})();
</script>
"""


def scatter_chart_html(data: dict[str, Any]) -> str:
    rows: list[dict[str, Any]] = []
    for row in data.get("funds", []):
        strategy_units = []
        for item in row.get("strategyUnits", []) or []:
            strategy_units.append(
                {
                    "unitId": item.get("unitId"),
                    "strategyName": item.get("strategyName"),
                    "strategyIds": item.get("strategyIds") or [],
                    "advisor": item.get("advisor"),
                    "unitType": item.get("unitType"),
                    "holdingDate": item.get("holdingDate"),
                    "weight": item.get("weight"),
                    "shareCodes": item.get("shareCodes") or [],
                    "sourceStrategyCount": item.get("sourceStrategyCount"),
                }
            )
        rows.append(
            {
                "code": row.get("code"),
                "name": row.get("name"),
                "company": row.get("company"),
                "direction": row.get("direction"),
                "fundType": row.get("fundType") or "未披露",
                "strategyCount": row.get("strategyCount"),
                "totalWeightRatio": row.get("totalWeightRatio"),
                "return1w": row.get("return1w"),
                "return1m": row.get("return1m"),
                "return3m": row.get("return3m"),
                "return6m": row.get("return6m"),
                "return1y": row.get("return1y"),
                "returnYtd": row.get("returnYtd"),
                "maxDrawdown1y": row.get("maxDrawdown1y"),
                "volatility1y": row.get("volatility1y"),
                "sharpeRatio1y": row.get("sharpeRatio1y"),
                "latestNavDate": row.get("latestNavDate"),
                "personalLimit": row.get("personalLimit"),
                "nonPersonLimit": row.get("nonPersonLimit"),
                "personalLimitAmount": row.get("personalLimitAmount"),
                "nonPersonLimitAmount": row.get("nonPersonLimitAmount"),
                "personalLimitSource": row.get("personalLimitSource"),
                "nonPersonLimitSource": row.get("nonPersonLimitSource"),
                "limitSourceTitle": row.get("limitSourceTitle"),
                "limitSourceDate": row.get("limitSourceDate"),
                "strategyUnits": strategy_units,
            }
        )
    payload = json.dumps(rows, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")
    template = """
<div class="scatter-controls">
  <label>Y轴收益区间
    <select id="qdScatterReturnMetric">
      <option value="return1m">近1月收益率</option>
      <option value="return1w">近1周收益率</option>
      <option value="return3m">近3月收益率</option>
      <option value="return6m">近6月收益率</option>
      <option value="return1y">近1年收益率</option>
      <option value="returnYtd">今年以来收益率</option>
    </select>
  </label>
  <label>X轴风控指标
    <select id="qdScatterRiskMetric">
      <option value="maxDrawdown1y">近1年最大回撤</option>
      <option value="volatility1y">近1年年化波动率</option>
      <option value="sharpeRatio1y">近1年夏普</option>
    </select>
  </label>
  <label>点大小
    <select id="qdScatterSizeMetric">
      <option value="strategyCount">配置策略数</option>
      <option value="totalWeightRatio">策略总仓位占比</option>
      <option value="maxLimitAmount">最大限额</option>
      <option value="limitCapacityAmount">保守限额容量</option>
    </select>
  </label>
</div>
<div id="qdScatterLegend" class="scatter-legend"></div>
<div id="qdScatter" class="scatter-chart"></div>
<div id="qdScatterDetail" class="scatter-detail empty">点击点位后，在这里查看该 QD 基金对应的配置策略、仓位和关键业绩指标。</div>
<script>
(function(){
  const rows = __PAYLOAD__;
  const metricNames = {
    return1w:"近1周收益率", return1m:"近1月收益率", return3m:"近3月收益率",
    return6m:"近6月收益率", return1y:"近1年收益率", returnYtd:"今年以来收益率",
    maxDrawdown1y:"近1年最大回撤", volatility1y:"近1年年化波动率", sharpeRatio1y:"近1年夏普"
  };
  const palette = ["#1d4ed8","#c026d3","#ea580c","#0891b2","#be123c","#65a30d","#7c3aed","#b45309","#0f766e","#e11d48","#0369a1","#4d7c0f"];
  const el = document.getElementById("qdScatter");
  const legend = document.getElementById("qdScatterLegend");
  const detail = document.getElementById("qdScatterDetail");
  const returnSelect = document.getElementById("qdScatterReturnMetric");
  const riskSelect = document.getElementById("qdScatterRiskMetric");
  const sizeSelect = document.getElementById("qdScatterSizeMetric");
  const fundTypes = [...new Set(rows.map(d => d.fundType || "未披露"))].sort((a,b) => a.localeCompare(b, "zh-CN"));
  const colorByType = new Map(fundTypes.map((type, idx) => [type, palette[idx % palette.length]]));
  const sizeMetricNames = {
    strategyCount: "配置策略数",
    totalWeightRatio: "策略总仓位占比",
    maxLimitAmount: "最大限额",
    limitCapacityAmount: "保守限额容量"
  };
  let activeTypes = new Set(fundTypes);
  let selectedCode = "";
  let detailSort = { key: "weight", dir: "desc" };

  const esc = (v) => String(v ?? "").replace(/[&<>"']/g, s => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[s]));
  const fmtNum = (v) => v == null || Number.isNaN(Number(v)) ? "--" : Number(v).toFixed(2);
  const pctText = (v, sign=false) => v == null || Number.isNaN(Number(v)) ? "--" : `${sign && Number(v)>0 ? "+" : ""}${Number(v).toFixed(2)}%`;
  const moneyText = (v) => {
    const n = Number(v);
    if (!Number.isFinite(n) || n <= 0) return "--";
    if (n >= 100000000) return `${(n / 100000000).toFixed(n >= 1000000000 ? 1 : 2)}亿`;
    if (n >= 10000) return `${(n / 10000).toFixed(n >= 1000000 ? 0 : 1)}万`;
    return `${Math.round(n)}元`;
  };
  const limitCapacityAmount = (row) => {
    const values = [Number(row.personalLimitAmount), Number(row.nonPersonLimitAmount)].filter(Number.isFinite).filter(v => v > 0);
    return values.length ? Math.min(...values) : null;
  };
  const maxLimitAmount = (row) => {
    const values = [Number(row.personalLimitAmount), Number(row.nonPersonLimitAmount)].filter(Number.isFinite).filter(v => v > 0);
    return values.length ? Math.max(...values) : null;
  };
  const limitCapacityText = (row) => {
    const value = limitCapacityAmount(row);
    return value == null ? "--" : moneyText(value);
  };
  const maxLimitText = (row) => {
    const value = maxLimitAmount(row);
    return value == null ? "--" : moneyText(value);
  };
  const sizeValue = (row, metric) => {
    if (metric === "limitCapacityAmount") return limitCapacityAmount(row);
    if (metric === "maxLimitAmount") return maxLimitAmount(row);
    return Number(row[metric]);
  };
  const metricText = (metric, value) => {
    if (value == null || Number.isNaN(Number(value))) return "--";
    const n = Number(value);
    if (metric === "sharpeRatio1y") return n.toFixed(2);
    return pctText(n, metric.startsWith("return"));
  };
  const sizeText = (row, metric) => {
    const value = sizeValue(row, metric);
    if (!Number.isFinite(Number(value))) return "--";
    return metric === "limitCapacityAmount" || metric === "maxLimitAmount" ? moneyText(value) : metric === "strategyCount" ? `${Number(value).toFixed(0)}只` : pctText(value);
  };
  const pctHtml = (v, sign=false, color=true) => {
    if (v == null || Number.isNaN(Number(v))) return "--";
    const n = Number(v);
    const cls = !color || n === 0 ? "neutral" : n > 0 ? "pos" : "neg";
    return `<span class="${cls}">${sign && n > 0 ? "+" : ""}${n.toFixed(2)}%</span>`;
  };
  const median = (values) => {
    const nums = values.map(Number).filter(Number.isFinite).sort((a,b) => a-b);
    if (!nums.length) return 0;
    const mid = Math.floor(nums.length / 2);
    return nums.length % 2 ? nums[mid] : (nums[mid - 1] + nums[mid]) / 2;
  };
  const fundLink = (row) => `<a class="fund-link" href="./fund.html?code=${encodeURIComponent(row.code || "")}">${esc(row.code || "")} ${esc(row.name || "")}</a>`;
  const strategyLink = (item) => {
    const id = (item.strategyIds || [])[0] || item.unitId || "";
    return `<a class="strategy-link" href="./strategy.html?id=${encodeURIComponent(id)}">${esc(item.strategyName || "--")}</a>`;
  };

  function renderLegend(){
    const allActive = activeTypes.size === fundTypes.length;
    legend.innerHTML = `<button type="button" class="legend-filter ${allActive ? "is-active" : ""}" data-type="__all__">全部</button>` + fundTypes.map(type => {
      const color = colorByType.get(type);
      const isActive = activeTypes.has(type);
      return `<button type="button" class="legend-filter ${isActive ? "is-active" : "is-muted"}" data-type="${esc(type)}"><i style="background:${color}"></i>${esc(type)}</button>`;
    }).join("");
    legend.querySelectorAll(".legend-filter").forEach(btn => btn.addEventListener("click", () => {
      const type = btn.dataset.type || "";
      if (type === "__all__") {
        activeTypes = new Set(fundTypes);
      } else if (activeTypes.has(type)) {
        activeTypes.delete(type);
        if (!activeTypes.size) activeTypes = new Set(fundTypes);
      } else {
        activeTypes.add(type);
      }
      selectedCode = "";
      renderLegend();
      render();
    }));
  }

  function sortedStrategies(row){
    const strategies = (row.strategyUnits || []).slice();
    const dir = detailSort.dir === "asc" ? 1 : -1;
    const valueOf = (item) => {
      if (detailSort.key === "weight") return Number(item.weight) || 0;
      if (detailSort.key === "holdingDate") return String(item.holdingDate || "");
      if (detailSort.key === "shareCodes") return (item.shareCodes || []).join(",");
      return String(item[detailSort.key] || "").toLowerCase();
    };
    strategies.sort((a,b) => {
      const av = valueOf(a), bv = valueOf(b);
      if (av < bv) return -1 * dir;
      if (av > bv) return 1 * dir;
      return String(a.strategyName || "").localeCompare(String(b.strategyName || ""), "zh-CN");
    });
    return strategies;
  }

  function sortButton(label, key){
    const mark = detailSort.key === key ? (detailSort.dir === "asc" ? " ▲" : " ▼") : "";
    return `<button type="button" class="sort-head" data-sort="${key}">${esc(label)}${mark}</button>`;
  }

  function renderDetail(row){
    if (!row) {
      detail.className = "scatter-detail empty";
      detail.innerHTML = "点击点位后，在这里查看该 QD 基金对应的配置策略、仓位和关键业绩指标。";
      return;
    }
    const typeColor = colorByType.get(row.fundType || "未披露") || "#64748b";
    const strategies = sortedStrategies(row);
    detail.className = "scatter-detail";
    detail.innerHTML = `
      <div class="scatter-detail-head">
        <div>
          <h3>${fundLink(row)}</h3>
          <p class="muted">${esc(row.company || "--")} · ${esc(row.direction || "--")} · 最新净值 ${esc(row.latestNavDate || "--")}</p>
        </div>
        <div class="scatter-type-pill" style="--type-color:${typeColor}">${esc(row.fundType || "未披露")}</div>
      </div>
      <div class="metrics compact">
        <div class="metric"><b>${pctHtml(row.return1m, true)}</b><span>近1月收益</span></div>
        <div class="metric"><b>${pctHtml(row.return3m, true)}</b><span>近3月收益</span></div>
        <div class="metric"><b>${pctHtml(row.return1y, true)}</b><span>近1年收益</span></div>
        <div class="metric"><b>${pctHtml(row.maxDrawdown1y, true)}</b><span>近1年最大回撤</span></div>
        <div class="metric"><b>${pctHtml(row.volatility1y, false, false)}</b><span>近1年波动率</span></div>
        <div class="metric"><b>${row.sharpeRatio1y == null || Number.isNaN(Number(row.sharpeRatio1y)) ? "--" : Number(row.sharpeRatio1y).toFixed(2)}</b><span>近1年夏普</span></div>
        <div class="metric"><b>${pctHtml(row.totalWeightRatio, false, false)}</b><span>配置总仓位占比</span></div>
        <div class="metric limit-metric"><b>${esc(row.personalLimit || "--")}</b><span>个人限额</span><small>${esc(row.personalLimitSource || "")}</small></div>
        <div class="metric limit-metric"><b>${esc(row.nonPersonLimit || "--")}</b><span>非个人限额</span><small>${esc(row.nonPersonLimitSource || "")}</small></div>
        <div class="metric limit-metric"><b>${maxLimitText(row)}</b><span>最大限额</span><small>个人/非个人较大值</small></div>
        <div class="metric limit-metric"><b>${limitCapacityText(row)}</b><span>限额容量口径</span><small>取个人/非个人较小值</small></div>
      </div>
      ${row.limitSourceTitle ? `<p class="small">限额证据：${esc(row.limitSourceDate || "--")} · ${esc(row.limitSourceTitle)}</p>` : ""}
      <details class="scatter-strategy-details">
        <summary>配置策略列表（${strategies.length}）</summary>
        <table class="compact-table sortable-table">
          <thead><tr><th>${sortButton("配置策略", "strategyName")}</th><th>${sortButton("投顾机构", "advisor")}</th><th>${sortButton("策略类型", "unitType")}</th><th>${sortButton("持仓日期", "holdingDate")}</th><th class="num">${sortButton("策略内仓位", "weight")}</th><th>${sortButton("份额代码", "shareCodes")}</th></tr></thead>
          <tbody>${strategies.map(item => `
            <tr>
              <td>${strategyLink(item)}${Number(item.sourceStrategyCount || 0) > 1 ? `<div class="small">合并 ${esc(item.sourceStrategyCount)} 个期次</div>` : ""}</td>
              <td>${esc(item.advisor || "--")}</td>
              <td>${esc(item.unitType || "--")}</td>
              <td>${esc(item.holdingDate || "--")}</td>
              <td class="num">${pctText(item.weight)}</td>
              <td>${esc((item.shareCodes || []).join(","))}</td>
            </tr>`).join("")}</tbody>
        </table>
      </details>
    `;
    detail.querySelectorAll(".sort-head").forEach(btn => btn.addEventListener("click", () => {
      const key = btn.dataset.sort || "weight";
      if (detailSort.key === key) {
        detailSort.dir = detailSort.dir === "asc" ? "desc" : "asc";
      } else {
        detailSort = { key, dir: key === "weight" ? "desc" : "asc" };
      }
      renderDetail(row);
      const panel = detail.querySelector(".scatter-strategy-details");
      if (panel) panel.open = true;
    }));
  }

  function render(){
    const returnMetric = returnSelect.value;
    const riskMetric = riskSelect.value;
    const sizeMetric = sizeSelect.value;
    const usable = rows.filter(d => activeTypes.has(d.fundType || "未披露") && Number.isFinite(Number(d[riskMetric])) && Number.isFinite(Number(d[returnMetric])));
    if (!usable.length) {
      el.innerHTML = '<div class="empty">当前指标暂无可展示数据</div>';
      renderDetail(null);
      return;
    }
    const w = Math.max(820, el.clientWidth || 820);
    const h = 430;
    const m = {l:70,r:30,t:34,b:62};
    const xVals = usable.map(d => Number(d[riskMetric]));
    const yVals = usable.map(d => Number(d[returnMetric]));
    let xMin = Math.min(...xVals), xMax = Math.max(...xVals);
    if (xMin === xMax) { xMin -= 1; xMax += 1; }
    const xPad = Math.max((xMax - xMin) * 0.12, riskMetric === "sharpeRatio1y" ? 0.1 : 0.5);
    xMin -= xPad; xMax += xPad;
    let yMin = Math.min(...yVals), yMax = Math.max(...yVals);
    if (yMin === yMax) { yMin -= 1; yMax += 1; }
    const pad = Math.max((yMax - yMin) * 0.12, 0.5);
    yMin -= pad; yMax += pad;
    const sizeMax = Math.max(...usable.map(d => Number(sizeValue(d, sizeMetric)) || 0), 0.01);
    const x = v => m.l + (Number(v) - xMin) / (xMax - xMin) * (w - m.l - m.r);
    const y = v => m.t + (yMax - Number(v)) / (yMax - yMin) * (h - m.t - m.b);
    const r = v => 4.5 + Math.sqrt(Math.max(Number(v) || 0, 0) / sizeMax) * 12;
    const xTicks = Array.from({length:7}, (_,i) => xMin + (xMax - xMin) * i / 6);
    const yTicks = Array.from({length:7}, (_,i) => yMin + (yMax - yMin) * i / 6);
    const xMedian = median(xVals);
    const yMedian = median(yVals);
    const yZeroLine = yMin < 0 && yMax > 0 ? `<g><line class="zero-line" x1="${m.l}" x2="${w-m.r}" y1="${y(0)}" y2="${y(0)}"></line><text class="zero-label" x="${w-m.r-4}" y="${y(0)-5}" text-anchor="end">收益0线</text></g>` : "";
    const xZeroLine = xMin < 0 && xMax > 0 ? `<g><line class="zero-line" x1="${x(0)}" x2="${x(0)}" y1="${m.t}" y2="${h-m.b}"></line><text class="zero-label" x="${x(0)+6}" y="${h-m.b-8}">风控0线</text></g>` : "";
    const medianLines = `<g><line class="median-line" x1="${x(xMedian)}" x2="${x(xMedian)}" y1="${m.t}" y2="${h-m.b}"></line><text class="axis-text" x="${x(xMedian)+6}" y="${m.t+14}">风控中位</text><line class="median-line" x1="${m.l}" x2="${w-m.r}" y1="${y(yMedian)}" y2="${y(yMedian)}"></line><text class="axis-text" x="${w-m.r-4}" y="${y(yMedian)-7}" text-anchor="end">收益中位</text></g>`;
    const points = usable.map(d => {
      const cx = x(d[riskMetric]), cy = y(d[returnMetric]), rr = r(sizeValue(d, sizeMetric));
      const type = d.fundType || "未披露";
      const title = `${d.code || ""} ${d.name || ""}\\n基金类型：${type}\\nY轴 ${metricNames[returnMetric]}：${metricText(returnMetric, d[returnMetric])}\\nX轴 ${metricNames[riskMetric]}：${metricText(riskMetric, d[riskMetric])}\\n点大小：${sizeMetricNames[sizeMetric]} ${sizeText(d, sizeMetric)}\\n策略数：${d.strategyCount}\\n配置总仓位：${pctText(d.totalWeightRatio)}\\n最大限额：${maxLimitText(d)}\\n个人限额：${d.personalLimit || "--"}\\n非个人限额：${d.nonPersonLimit || "--"}`;
      return `<circle class="fund-point${selectedCode === d.code ? " selected" : ""}" data-code="${esc(d.code || "")}" cx="${cx.toFixed(1)}" cy="${cy.toFixed(1)}" r="${rr.toFixed(1)}" style="--fund-color:${colorByType.get(type) || "#64748b"}" tabindex="0"><title>${esc(title)}</title></circle>`;
    }).join("");
    el.innerHTML = `<svg viewBox="0 0 ${w} ${h}" role="img" aria-label="QD基金配置点阵图">
      <rect class="plot-bg" x="${m.l}" y="${m.t}" width="${w-m.l-m.r}" height="${h-m.t-m.b}"></rect>
      ${yTicks.map((t,i)=>`<g><line class="grid-line${i===0||i===yTicks.length-1 ? " strong" : ""}" x1="${m.l}" x2="${w-m.r}" y1="${y(t)}" y2="${y(t)}"></line><text class="axis-text" x="${m.l-10}" y="${y(t)+4}" text-anchor="end">${metricText(returnMetric, t)}</text></g>`).join("")}
      ${xTicks.map(t=>`<g><line class="grid-line" x1="${x(t)}" x2="${x(t)}" y1="${m.t}" y2="${h-m.b}"></line><text class="axis-text" x="${x(t)}" y="${h-22}" text-anchor="middle">${metricText(riskMetric, t)}</text></g>`).join("")}
      ${yZeroLine}
      ${xZeroLine}
      ${medianLines}
      <line class="axis-line" x1="${m.l}" x2="${w-m.r}" y1="${h-m.b}" y2="${h-m.b}"></line>
      <line class="axis-line" x1="${m.l}" x2="${m.l}" y1="${m.t}" y2="${h-m.b}"></line>
      <text class="axis-title" x="${w/2}" y="${h-4}" text-anchor="middle">${metricNames[riskMetric]}</text>
      <text class="axis-title" transform="translate(18 ${h/2}) rotate(-90)" text-anchor="middle">${metricNames[returnMetric]}</text>
      ${points}
    </svg>`;
    el.querySelectorAll(".fund-point").forEach(node => {
      const code = node.dataset.code || "";
      const row = rows.find(item => String(item.code || "") === code);
      node.addEventListener("click", () => {
        selectedCode = code;
        renderDetail(row);
        render();
      });
      node.addEventListener("keydown", (ev) => {
        if (ev.key === "Enter" || ev.key === " ") {
          ev.preventDefault();
          selectedCode = code;
          renderDetail(row);
          render();
        }
      });
    });
    if (selectedCode) {
      const row = rows.find(item => String(item.code || "") === String(selectedCode));
      if (row) renderDetail(row);
    }
  }
  renderLegend();
  returnSelect.addEventListener("change", render);
  riskSelect.addEventListener("change", render);
  sizeSelect.addEventListener("change", render);
  window.addEventListener("resize", render);
  render();
})();
</script>
"""
    return template.replace("__PAYLOAD__", payload)


def build_report_html(data: dict[str, Any]) -> str:
    funds = data["funds"]
    directions = data["directions"]
    strategy_types = data.get("strategyTypeRows") or []
    top_consensus = sorted(funds, key=lambda x: (-x["strategyCount"], -x["advisorCount"], -x["totalWeight"]))
    top_inflow = sorted(funds, key=lambda x: (-x["recentNetInflow"], -x["strategyCount"]))
    top_outflow = sorted(funds, key=lambda x: (x["recentNetInflow"], -x["strategyCount"]))
    risk_funds = sorted([row for row in funds if row["limitLevel"] == "risk"], key=lambda x: (-x["strategyCount"], -x["advisorCount"]))
    top_gf = [row for row in funds if "广发" in (row.get("company") or "") or str(row.get("name") or "").startswith("广发")]
    value_candidates = []
    for row in funds:
        amounts = [as_float(row.get("personalLimitAmount")), as_float(row.get("nonPersonLimitAmount"))]
        amounts = [v for v in amounts if v is not None]
        min_limit = min(amounts) if amounts else None
        if row.get("limitLevel") == "risk":
            continue
        if (row.get("return1m") is None or row.get("return1m") <= 0) and (row.get("return3m") is None or row.get("return3m") <= 0):
            continue
        if row.get("strategyCount", 0) < 2:
            continue
        if min_limit is not None and min_limit < 100_000:
            continue
        candidate = dict(row)
        perf_part = f"近1月{pct(row.get('return1m'), signed=True)}" if row.get("return1m") is not None else f"近3月{pct(row.get('return3m'), signed=True)}"
        candidate["valueReason"] = (
            f"{perf_part}，{row.get('strategyCount')}个策略、{row.get('advisorCount')}家机构已配置；"
            f"个人{row.get('personalLimit')}，非个人{row.get('nonPersonLimit')}。"
        )
        value_candidates.append(candidate)
    value_candidates.sort(
        key=lambda x: (
            -(as_float(x.get("return1m")) or -999),
            -x.get("strategyCount", 0),
            -x.get("advisorCount", 0),
            -(as_float(x.get("personalLimitAmount")) or 0),
        )
    )
    direction_line = "、".join(item["direction"] for item in directions[:4]) or "暂无"
    best = top_consensus[0] if top_consensus else {}
    inflow = top_inflow[0] if top_inflow else {}
    limit_note = (
        f"{data['summary']['limitRiskCount']} 只基金存在低限额或限大额风险"
        if data["summary"]["limitRiskCount"]
        else "未识别到低限额或限大额风险，但需以交易端实时状态为准"
    )
    conclusion = (
        f"当前 QD 配置主线集中在 {direction_line}；"
        f"{best.get('name', '暂无基金')} 是配置策略数最高的工具，"
        f"{inflow.get('name', '暂无基金')} 近 1 月净调入最明显。{limit_note}。"
    )
    type_html = top_table(
        strategy_types,
        [("策略类型", "strategyType"), ("含QD策略", "qdStrategyCount"), ("渗透率", "penetration"), ("QD中位仓位", "medianQdWeight"), ("主方向", "topDirection")],
        8,
    )
    direction_blocks = []
    for direction in directions[:6]:
        top3 = sorted([row for row in funds if row["direction"] == direction["direction"]], key=lambda x: (-x["strategyCount"], -x["advisorCount"], -x["totalWeight"]))[:3]
        items = "".join(
            f"""<li><b>{escape(row['name'])}</b><span>{row['strategyCount']} 策略 · 均仓 {pct(row['averageWeight'])} · 1月 {pct(row.get('return1m'), signed=True)} · 个人 {escape(row['personalLimit'])} / 非个人 {escape(row['nonPersonLimit'])}</span></li>"""
            for row in top3
        )
        direction_blocks.append(f"<div class='top3-card'><h3>{escape(direction['direction'])}</h3><ul>{items}</ul></div>")
    direction_html = "".join(direction_blocks)
    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>QD基金配置一页报告</title>
  <style>{CSS}</style>
</head>
<body>
  {qd_nav_html('qd')}
  <main class="page">
    <header class="hero">
      <div>
        <p class="eyebrow">全市场投顾组合 · QD基金配置</p>
        <h1>QD基金配置一页报告</h1>
        <p class="lead">{escape(conclusion)}</p>
      </div>
      <aside>
        <b>统计时间</b>
        <strong>{escape(data['generatedAt'][:10])}</strong>
        <span>当前持仓主口径；目标盈系列合并计数。</span>
        <a href="./qd-fund-detail.html">查看明细</a>
        <a href="./qd-fund-detail.xlsx">下载XLSX</a>
      </aside>
    </header>
    <section class="metrics">{report_cards_html(data)}</section>
    <section class="grid two">
      <article class="panel">
        <h2>QD方向配置分布</h2>
        <p>按合并后的底层基金统计，条形长度和右侧数字代表配置该方向的策略数；方向名下方的“基金数 / 总仓位”分别表示该方向底层基金数量、该方向仓位合计占全部有效策略披露仓位的比例。</p>
        {bar_chart_html(directions, 'direction', 'strategyCount', ['fundCount', 'totalWeightRatio'], 8)}
      </article>
      <article class="panel">
        <h2>限额可得性分布</h2>
        <p>个人和非个人分别分桶统计，蓝色代表个人，绿色代表非个人；“通用公告参考”表示公告没有拆分个人/非个人，“未设日限额”表示销售端未返回明确日累计上限，不是采集失败。</p>
        {grouped_bar_chart_html(limit_bucket_rows(funds), 'bucket', 'personalFundCount', 'nonPersonFundCount', '个人', '非个人', 8)}
        <h3>来源明确度</h3>
        {grouped_bar_chart_html(limit_source_rows(funds), 'bucket', 'personalFundCount', 'nonPersonFundCount', '个人', '非个人', 8)}
      </article>
    </section>
    <section class="panel">
      <h2>QD基金策略数-业绩点阵</h2>
      <p>横轴是配置该基金的策略数，纵轴可切换收益、回撤、波动率；点越大，表示该基金在全部有效策略披露仓位中的合计占比越高。收益率指标按正收益红色、负收益绿色显示。</p>
      {scatter_chart_html(data)}
    </section>
    <section class="grid two">
      <article class="panel">
        <h2>策略类型内配置强度</h2>
        <p>只在各策略类型内部计算，避免纯现金或纯国内策略稀释 QD 仓位判断。</p>
        {type_html}
      </article>
      <article class="panel">
        <h2>限额风险</h2>
        <p>高共识但限大额或额度过低的基金，不适合作为自建策略必要成分；暂停申购状态在明细中单独查看。</p>
        {top_table(risk_funds, [('基金', 'name'), ('方向', 'direction'), ('策略数', 'strategyCount'), ('个人限额', 'personalLimit'), ('非个人限额', 'nonPersonLimit')], 8)}
      </article>
    </section>
    <section class="panel">
      <h2>配置价值观察：额度较高且近期表现较好</h2>
      <p>筛选逻辑：剔除低限额/限大额风险，个人和非个人参考限额不低于 10 万元或销售端未设日限额，且近 1 月或近 3 月收益为正、至少 2 个策略配置。</p>
      {top_table(value_candidates, [('基金', 'name'), ('方向', 'direction'), ('策略数', 'strategyCount'), ('近1月收益', 'return1m'), ('个人限额', 'personalLimit'), ('非个人限额', 'nonPersonLimit'), ('观察理由', 'valueReason')], 8)}
    </section>
    <section class="grid two">
      <article class="panel">
        <h2>近1月增减持信号</h2>
        <h3>净调入 Top3</h3>
        {top_table([r for r in top_inflow if r['recentNetInflow'] > 0], [('基金', 'name'), ('方向', 'direction'), ('净调入', 'recentNetInflow'), ('近1月收益', 'return1m'), ('个人限额', 'personalLimit'), ('非个人限额', 'nonPersonLimit')], 3)}
        <h3>净调出 Top3</h3>
        {top_table([r for r in top_outflow if r['recentNetInflow'] < 0], [('基金', 'name'), ('方向', 'direction'), ('净调入', 'recentNetInflow'), ('近1月收益', 'return1m'), ('个人限额', 'personalLimit'), ('非个人限额', 'nonPersonLimit')], 3)}
      </article>
      <article class="panel">
        <h2>广发相关基金</h2>
        {top_table(sorted(top_gf, key=lambda x: (-x['strategyCount'], -x['totalWeight'])), [('基金', 'name'), ('方向', 'direction'), ('策略数', 'strategyCount'), ('近1月收益', 'return1m'), ('个人限额', 'personalLimit'), ('非个人限额', 'nonPersonLimit')], 6)}
      </article>
    </section>
    <section class="panel">
      <h2>每类 Top3 基金</h2>
      <div class="top3-grid">{direction_html}</div>
    </section>
    <section class="panel recommendation">
      <h2>给自建策略的使用建议</h2>
      <div class="rec-grid">
        <div><b>核心候选</b><span>优先看策略数和机构覆盖都靠前、且限额无明显风险的基金。</span></div>
        <div><b>卫星增强</b><span>近 1 月净调入明显、方向和现有组合互补的基金可进入观察池。</span></div>
        <div><b>谨慎使用</b><span>限额未确认或高共识但可买性不清晰的基金，不建议设为必要成分。</span></div>
      </div>
    </section>
    <footer class="foot">
      口径：严格 QDII/海外 QDII 主口径；测试、信号、已停止策略剔除；目标盈期次按同机构同系列合并为一个策略。同一基金 A/C/F 等份额合并统计，仓位、策略数、渗透率和调仓按合并份额汇总，业绩和限额取主基金份额。限额来自天天基金申购状态与东方财富/天天基金公告正文；个人和非个人限额优先用公告正文，公告未单独披露时用销售端日累计限额作为参考并标注来源。
    </footer>
  </main>
  <script>{info_script()}</script>
</body>
</html>"""
    return html


def js_data(data: dict[str, Any]) -> str:
    return json.dumps(data, ensure_ascii=False, separators=(",", ":")).replace("</", "<\\/")


def build_detail_html(data: dict[str, Any]) -> str:
    html = f"""<!doctype html>
<html lang="zh-CN">
<head>
  <meta charset="utf-8">
  <meta name="viewport" content="width=device-width, initial-scale=1">
  <title>QD基金配置明细</title>
  <style>{CSS}</style>
</head>
<body>
  {qd_nav_html('qd_detail')}
  <main class="page wide">
    <header class="compact-hero">
      <div>
        <p class="eyebrow">全市场投顾组合 · QD基金配置</p>
        <h1>QD基金配置明细</h1>
        <p class="lead">按基金查看配置该基金的策略、机构和仓位，点击基金行展开策略明细。</p>
      </div>
      <div>
        <a class="ghost-link" href="./qd-fund-report.html">返回一页报告</a>
        <a class="ghost-link" href="./qd-fund-detail.xlsx">下载XLSX明细</a>
      </div>
    </header>
    <section class="toolbar">
      <input id="search" type="search" placeholder="搜索代码、基金、公司、方向">
      <select id="direction"><option value="">全部方向</option></select>
      <select id="limit"><option value="">全部限额状态</option><option value="unknown">限额待核验</option><option value="risk">限额风险</option><option value="known">已采状态</option></select>
      <select id="sort">
        <option value="strategyCount">按配置策略数</option>
        <option value="totalWeightRatio">按配置总仓位比例</option>
        <option value="recentNetInflow">按近1月净调入</option>
        <option value="advisorCount">按覆盖机构数</option>
      </select>
    </section>
    <section id="table"></section>
    <section id="detail" class="panel detail-panel"><p class="muted">点击上方基金行查看配置策略明细。</p></section>
    <footer class="foot">
      目标盈系列已合并计数。目标盈系列的仓位取最新持仓日同系列期次的中位数代表仓位，避免一期一期重复放大。
    </footer>
  </main>
  <script>{info_script()}</script>
  <script>window.QD_FUND_DATA={js_data(data)};</script>
  <script>{DETAIL_JS}</script>
</body>
</html>"""
    return html


CSS = """
:root{color-scheme:light;--ink:#162033;--muted:#667085;--line:#dbe4ef;--soft:#f4f7fb;--accent:#0f766e;--red:#dc2626;--orange:#c77700;--blue:#2563eb}
*{box-sizing:border-box}body{margin:0;background:#f7fafc;color:var(--ink);font:15px/1.55 -apple-system,BlinkMacSystemFont,"Segoe UI","Microsoft YaHei",sans-serif}.page{max-width:1180px;margin:0 auto;padding:44px 32px 56px}.page.wide{max-width:1320px}.site-nav{display:flex;flex-wrap:wrap;justify-content:flex-end;gap:8px;margin:-18px 0 24px}.site-nav a{border:1px solid #dbe4ef;border-radius:999px;padding:6px 12px;color:#344054;background:white;text-decoration:none;font-weight:800;font-size:13px}.site-nav a.is-active,.site-nav a:hover{background:#eaf3fb;color:#0f766e}.hero,.compact-hero{display:flex;justify-content:space-between;gap:32px;align-items:flex-start;border-bottom:3px solid #d9e3ef;padding-bottom:26px}.compact-hero{align-items:center}.eyebrow{color:#0f766e;font-weight:800;margin:0 0 8px}h1{font-size:42px;line-height:1.1;margin:0 0 14px;letter-spacing:0}h2{font-size:24px;margin:0 0 10px}.lead{font-size:18px;color:#526173;max-width:760px;margin:0}.hero aside{min-width:260px;background:#eaf3fb;border:1px solid #cbddee;border-radius:8px;padding:18px 20px}.hero aside b,.hero aside span{display:block;color:#526173}.hero aside strong{display:block;font-size:26px;margin:6px 0 10px}.hero aside a,.ghost-link{display:inline-block;margin-top:14px;margin-right:14px;color:#0f766e;font-weight:800;text-decoration:none}.metrics{display:grid;grid-template-columns:repeat(auto-fit,minmax(150px,1fr));gap:12px;margin:22px 0}.metric{background:white;border:1px solid var(--line);border-radius:8px;padding:16px}.metric b{display:block;font-size:26px}.metric span{display:block;font-weight:800}.metric small{color:var(--muted)}.grid{display:grid;gap:16px;margin-top:18px}.grid.two{grid-template-columns:1fr 1fr}.panel{background:white;border:1px solid var(--line);border-radius:8px;padding:22px}.panel p,.muted{color:var(--muted);margin-top:0}table{width:100%;border-collapse:collapse}th{background:#edf3fa;text-align:left;color:#344054}th,td{padding:10px 9px;border-bottom:1px solid #e6edf5;vertical-align:top}td{font-size:14px}.pos{color:#dc2626;font-weight:800}.neg{color:#059669;font-weight:800}.neutral{color:#475467}.bars{display:grid;gap:11px}.bar-row{display:grid;grid-template-columns:150px 1fr 70px;gap:12px;align-items:center}.bar-row span{display:block;color:var(--muted);font-size:12px}.bar{height:12px;background:#edf2f7;border-radius:999px;overflow:hidden}.bar i{display:block;height:100%;background:var(--accent)}.bar.personal i{background:#2563eb}.bar.non-person i{background:#0f766e}.bar-row em{text-align:right;font-style:normal;font-weight:800}.grouped-bars-wrap{display:grid;gap:12px}.chart-legend{display:flex;gap:14px;align-items:center;color:#475467;font-size:13px}.chart-legend span{display:inline-flex;align-items:center;gap:6px}.chart-legend i{display:inline-block;width:10px;height:10px;border-radius:2px}.legend-personal{background:#2563eb}.legend-non-person{background:#0f766e}.grouped-bar-row{display:grid;grid-template-columns:96px 1fr 58px;gap:10px;align-items:center}.grouped-label b{font-size:13px}.grouped-bars{display:grid;gap:5px}.grouped-values{display:grid;gap:2px;font-size:12px;font-weight:800;text-align:right;color:#344054}.scatter-controls{display:flex;justify-content:flex-end;margin:8px 0 12px}.scatter-controls label{display:inline-flex;align-items:center;gap:8px;color:#475467;font-weight:800}.scatter-controls select{height:34px;border:1px solid #cbd5e1;border-radius:8px;background:white;padding:0 10px}.scatter-chart{width:100%;min-height:390px;overflow-x:auto}.scatter-chart svg{width:100%;min-width:760px;height:auto}.plot-bg{fill:#fbfdff;stroke:#e6edf5}.grid-line{stroke:#e6edf5;stroke-width:1}.axis-line{stroke:#94a3b8;stroke-width:1.2}.axis-text{fill:#64748b;font-size:12px}.axis-title{fill:#344054;font-size:13px;font-weight:800}.scatter-chart circle{fill:#2563eb;stroke:white;stroke-width:1.5;opacity:.72;cursor:pointer}.scatter-chart circle.pos{fill:#dc2626}.scatter-chart circle.neg{fill:#059669}.scatter-chart circle.neutral{fill:#64748b}.scatter-chart circle:hover,.scatter-chart circle:focus{opacity:1;stroke:#162033;stroke-width:2}.rec-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.rec-grid div{border-left:5px solid var(--accent);background:#f5faf9;border-radius:6px;padding:14px}.rec-grid b,.rec-grid span{display:block}.rec-grid span{color:#526173;margin-top:6px}.top3-grid{display:grid;grid-template-columns:repeat(3,1fr);gap:14px}.top3-card{border:1px solid #e4edf6;border-radius:8px;padding:14px;background:#fbfdff}.top3-card h3,.panel h3{margin:0 0 8px;font-size:16px}.top3-card ul{margin:0;padding-left:18px}.top3-card li{margin:7px 0}.top3-card b,.top3-card span{display:block}.top3-card span{color:#526173;font-size:12px}.foot{margin-top:18px;color:#667085;background:white;border:1px solid var(--line);border-radius:8px;padding:14px 16px;font-size:13px}.toolbar{display:grid;grid-template-columns:2fr 1fr 1fr 1fr;gap:10px;margin:22px 0}.toolbar input,.toolbar select{height:40px;border:1px solid #cbd5e1;border-radius:8px;padding:0 12px;background:white}.fund-link,.strategy-link{color:#0f766e;text-decoration:none;font-weight:800}.clickable{cursor:pointer}.tag{display:inline-flex;border-radius:999px;padding:2px 8px;background:#eef6ff;color:#1d4ed8;font-size:12px;font-weight:700}.tag.risk{background:#fef2f2;color:#b91c1c}.tag.unknown{background:#fff7ed;color:#a16207}.info-eye{position:relative;display:inline-block;width:18px;height:14px;margin-left:5px;padding:0;border:0;background:transparent;vertical-align:-2px;cursor:pointer}.info-eye:before{content:"";position:absolute;left:1px;top:2px;width:16px;height:10px;border:1.8px solid #64748b;border-radius:50%}.info-eye:after{content:"";position:absolute;left:7px;top:5px;width:4px;height:4px;background:#64748b;border-radius:50%}.info-modal{position:fixed;inset:0;z-index:1000}.info-backdrop{position:absolute;inset:0;background:rgba(15,23,42,.34)}.info-card{position:absolute;left:50%;top:50%;transform:translate(-50%,-50%);width:min(460px,calc(100vw - 36px));background:white;border:1px solid #dbe4ef;border-radius:8px;box-shadow:0 18px 45px rgba(15,23,42,.18);padding:22px}.info-card h3{margin:0 30px 10px 0}.info-card p{margin:0;color:#475467}.info-close{position:absolute;right:12px;top:10px;border:0;background:white;font-size:24px;line-height:1;cursor:pointer;color:#64748b}.detail-panel{margin-top:16px}.strategy-list{display:grid;gap:10px}.strategy-card{display:grid;grid-template-columns:1.4fr .9fr .7fr .7fr;gap:12px;border:1px solid #e6edf5;border-radius:8px;padding:12px}.small{font-size:12px;color:#667085}.num{text-align:right;font-variant-numeric:tabular-nums}.empty{padding:26px;text-align:center;color:#667085}@media(max-width:900px){.hero,.compact-hero{display:block}.site-nav{justify-content:flex-start;margin-top:0}.metrics,.grid.two,.rec-grid,.toolbar,.top3-grid{grid-template-columns:1fr}.bar-row,.grouped-bar-row{grid-template-columns:1fr}.strategy-card{grid-template-columns:1fr}h1{font-size:32px}.page{padding:28px 18px}}
"""

CSS += """
.topbar{position:sticky;top:0;z-index:20;background:rgba(255,255,255,.96);border-bottom:1px solid var(--line)}
.topbar-inner{max-width:1440px;margin:0 auto;padding:12px 24px;display:flex;align-items:center;justify-content:space-between;gap:20px}
.brand{display:inline-flex;align-items:center;gap:10px;min-width:220px;color:#182230;text-decoration:none}
.brand-mark{width:34px;height:34px;display:inline-grid;place-items:center;background:#166c77;color:#fff;border-radius:6px;font-weight:800}
.brand small{display:block;color:var(--muted);font-size:12px;margin-top:1px}
.nav{display:flex;gap:8px;flex-wrap:wrap;justify-content:flex-end}
.nav-link{padding:7px 10px;border-radius:6px;color:#405063;font-size:14px;text-decoration:none;font-weight:500}
.nav-link:hover,.nav-link.is-active{background:var(--soft);color:#0f4f58}
.scatter-legend{display:flex;flex-wrap:wrap;gap:8px 14px;margin:4px 0 12px;color:#475467;font-size:12px}
.scatter-controls{gap:12px;flex-wrap:wrap}
.scatter-legend span,.scatter-legend button{display:inline-flex;align-items:center;gap:6px}
.scatter-legend i{display:inline-block;width:10px;height:10px;border-radius:50%;box-shadow:0 0 0 1px rgba(15,23,42,.12)}
.legend-filter{border:1px solid #dbe4ef;background:#fff;color:#475467;border-radius:999px;padding:4px 9px;font-size:12px;font-weight:800;cursor:pointer}
.legend-filter.is-active{background:#eef6ff;border-color:#bfdbfe;color:#1d4ed8}
.legend-filter.is-muted{opacity:.42;background:#f8fafc}
.grid-line.strong{stroke:#d7e2ee;stroke-width:1.2}
.zero-line{stroke:#111827;stroke-width:1.35;stroke-dasharray:4 3}
.median-line{stroke:#64748b;stroke-width:1.1;stroke-dasharray:3 4}
.zero-label{fill:#334155;font-size:12px;font-weight:800}
.scatter-chart circle.fund-point{fill:var(--fund-color,#2563eb);stroke:white;stroke-width:1.5;opacity:.78;cursor:pointer}
.scatter-chart circle.fund-point:hover,.scatter-chart circle.fund-point:focus{opacity:1;stroke:#111827;stroke-width:2.4;outline:none}
.scatter-chart circle.fund-point.selected{opacity:1;stroke:#111827;stroke-width:3}
.scatter-detail{margin-top:14px;border:1px solid #dbe4ef;border-radius:8px;background:#fbfdff;padding:16px}
.scatter-detail.empty{color:#667085;text-align:center}
.scatter-detail-head{display:flex;justify-content:space-between;gap:18px;align-items:flex-start;margin-bottom:12px}
.scatter-detail-head h3{margin:0 0 4px;font-size:18px}
.scatter-type-pill{border-left:8px solid var(--type-color);background:white;border-radius:6px;padding:6px 10px;font-weight:800;color:#344054;white-space:nowrap;box-shadow:0 0 0 1px #e6edf5 inset}
.metrics.compact{grid-template-columns:repeat(auto-fit,minmax(130px,1fr));gap:10px;margin:10px 0 14px}
.metrics.compact .metric{padding:12px}
.metrics.compact .metric b{font-size:20px}
.metrics.compact .metric span{font-size:12px}
.metrics.compact .metric.limit-metric b{font-size:16px;line-height:1.3;word-break:break-word}
.metrics.compact .metric.limit-metric small{display:block;margin-top:4px}
.compact-table th,.compact-table td{font-size:13px;padding:8px}
.scatter-strategy-details{margin-top:8px}
.scatter-strategy-details summary{cursor:pointer;font-weight:900;color:#344054;padding:10px 0}
.sort-head{border:0;background:transparent;padding:0;color:#344054;font:inherit;font-weight:900;cursor:pointer;text-align:left}
th.num .sort-head{text-align:right;width:100%}
.sort-head:hover{color:#0f766e;text-decoration:underline}
@media(max-width:900px){.topbar-inner{align-items:flex-start;flex-direction:column}.scatter-detail-head{display:block}.scatter-type-pill{display:inline-block;margin-top:8px}.scatter-chart svg{min-width:700px}}
"""

CSS += SIDEBAR_CSS


DETAIL_JS = """
(function(){
  const data = window.QD_FUND_DATA || {funds:[]};
  const $ = (id) => document.getElementById(id);
  const esc = (v) => String(v ?? "").replace(/[&<>"']/g, s => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[s]));
  const pct = (v, signed=false) => v == null ? "--" : `${signed && v > 0 ? "+" : ""}${Number(v).toFixed(2)}%`;
  const pctHtml = (v, signed=false) => `<span class="${v == null || Number(v) === 0 ? "neutral" : Number(v) > 0 ? "pos" : "neg"}">${pct(v, signed)}</span>`;
  const info = (label) => window.qdInfoButton ? window.qdInfoButton(label) : esc(label);
  const directions = [...new Set(data.funds.map(x => x.direction).filter(Boolean))].sort();
  $("direction").innerHTML += directions.map(x => `<option>${esc(x)}</option>`).join("");
  let selectedCode = "";
  function filtered(){
    const q = $("search").value.trim().toLowerCase();
    const dir = $("direction").value;
    const lim = $("limit").value;
    const sort = $("sort").value;
    return data.funds.filter(row => {
      const hay = `${row.code} ${row.name} ${row.company} ${row.direction}`.toLowerCase();
      return (!q || hay.includes(q)) && (!dir || row.direction === dir) && (!lim || row.limitLevel === lim);
    }).sort((a,b) => (Number(b[sort] || 0) - Number(a[sort] || 0)) || (b.strategyCount - a.strategyCount) || a.code.localeCompare(b.code));
  }
  function limitTag(row){
    return `<span class="tag ${esc(row.limitLevel)}">${esc(row.limit || "未采集")}</span>`;
  }
  function renderTable(){
    const rows = filtered();
    $("table").innerHTML = `<div class="panel"><table><thead><tr>
      <th>主基金代码</th><th>QD基金名称</th><th>${info("合并份额")}</th><th>所属公司</th><th>QD方向</th><th class="num">配置策略数</th><th class="num">${info("QD策略均仓")}</th><th class="num">${info("近1月收益")}</th><th class="num">${info("近1月净调入")}</th><th>${info("个人限额")}</th><th>${info("非个人限额")}</th>
    </tr></thead><tbody>${rows.map(row => `<tr class="clickable" data-code="${esc(row.code)}">
      <td><a class="fund-link" href="./fund.html?code=${encodeURIComponent(row.code)}">${esc(row.code)}</a></td>
      <td>${esc(row.name)}<div class="small">${esc(row.fundType || "")}</div></td>
      <td>${esc((row.shareCodes || []).join(","))}</td>
      <td>${esc(row.company)}</td>
      <td>${esc(row.direction)}</td>
      <td class="num">${row.strategyCount}</td>
      <td class="num">${pct(row.averageWeight)}</td>
      <td class="num">${pctHtml(row.return1m, true)}<div class="small">${esc(row.latestNavDate || "")}</div></td>
      <td class="num">${pctHtml(row.recentNetInflow, true)}</td>
      <td>${esc(row.personalLimit || "未披露")}</td>
      <td>${esc(row.nonPersonLimit || "未披露")}</td>
    </tr>`).join("")}</tbody></table>${rows.length ? "" : '<div class="empty">没有匹配结果</div>'}</div>`;
    document.querySelectorAll("tr[data-code]").forEach(tr => tr.addEventListener("click", (ev) => {
      if (ev.target.closest("a")) return;
      selectedCode = tr.dataset.code;
      renderDetail();
    }));
  }
  function renderDetail(){
    const row = data.funds.find(x => x.code === selectedCode);
    if (!row) return;
    const examples = (row.recentExamples || []).map(x => `<li>${esc(x.date)} ${esc(x.advisor)} ${esc(x.strategyName)}：${pct(x.change, true)} ${esc(x.reason || "")}</li>`).join("");
    $("detail").innerHTML = `<h2>${esc(row.code)} ${esc(row.name)}</h2>
      <p>${esc(row.company)} · ${esc(row.direction)} · ${esc(row.fundType || "未披露分类")} · ${limitTag(row)}</p>
      <div class="metrics">
        <div class="metric"><b>${row.strategyCount}</b><span>配置策略数</span><small>目标盈系列合并</small></div>
        <div class="metric"><b>${row.advisorCount}</b><span>覆盖机构数</span><small>机构去重</small></div>
        <div class="metric"><b>${pct(row.totalWeightRatio)}</b><span>配置总仓位比例</span><small>占有效策略总仓位</small></div>
        <div class="metric"><b>${pct(row.averageWeight)}</b><span>平均配置仓位</span><small>配置该基金的策略</small></div>
        <div class="metric"><b>${pctHtml(row.return1m,true)}</b><span>近1月收益</span><small>${esc(row.latestNavDate || "净值日期缺失")}</small></div>
      </div>
      <table><tbody>
        <tr><th>申购状态</th><td>${esc(row.limit || "未披露")}</td></tr>
        <tr><th>${info("个人限额")}</th><td>${esc(row.personalLimit || "未单独披露")}</td></tr>
        <tr><th>${info("非个人限额")}</th><td>${esc(row.nonPersonLimit || "未单独披露")}</td></tr>
        <tr><th>${info("合并份额")}</th><td>${esc((row.shareNames || []).join("；"))}</td></tr>
        <tr><th>限额公告</th><td>${row.limitSourceUrl ? `<a class="fund-link" href="${esc(row.limitSourceUrl)}" target="_blank" rel="noopener">${esc(row.limitSourceDate || "")} ${esc(row.limitSourceTitle || "")}</a>` : "未解析到公告"}</td></tr>
      </tbody></table>
      ${examples ? `<h3>近期调仓信号</h3><ul>${examples}</ul>` : ""}
      <h3>配置该基金的策略</h3>
      <div class="strategy-list">${row.strategyUnits.map(item => {
        const id = item.strategyIds && item.strategyIds.length ? item.strategyIds[0] : "";
        return `<div class="strategy-card">
          <div><a class="strategy-link" href="./strategy.html?id=${encodeURIComponent(id)}">${esc(item.strategyName)}</a><div class="small">${esc(item.unitType)}${item.sourceStrategyCount > 1 ? ` · 合并${item.sourceStrategyCount}个期次` : ""} · ${esc((item.shareCodes || []).join(","))}</div></div>
          <div>${esc(item.advisor)}</div>
          <div>${esc(item.holdingDate || "--")}</div>
          <div class="num"><b>${pct(item.weight)}</b></div>
        </div>`;
      }).join("")}</div>`;
  }
  ["search","direction","limit","sort"].forEach(id => $(id).addEventListener("input", renderTable));
  renderTable();
})();
"""


def qd_workbook_purchase_status(fund: dict[str, Any]) -> str:
    purchase = fund.get("purchaseStatus") or {}
    return str(purchase.get("申购状态") or "").strip()


def qd_workbook_limit_source_text(value: Any) -> str:
    source = str(value or "").strip()
    labels = {
        "explicit": "公告明确披露",
        "general": "通用公告参考",
        "sales": "销售端参考",
    }
    return labels.get(source, source)


def qd_workbook_daily_limit_amount(fund: dict[str, Any], amount_key: str) -> int:
    purchase = fund.get("purchaseStatus") or {}
    status_text = str(purchase.get("申购状态") or fund.get("limit") or "")
    if "暂停申购" in status_text:
        return 0
    amount = as_float(fund.get(amount_key))
    if amount is None:
        amount = as_float(fund.get("salesDailyLimitAmount"))
    if amount is None:
        amount = as_float(purchase.get("日累计限定金额"))
    if amount is None or amount <= 0 or amount >= 100_000_000:
        return 100_000_000
    return int(round(amount))


def qd_workbook_advisor_summary(fund: dict[str, Any]) -> str:
    counts: dict[str, int] = defaultdict(int)
    for item in fund.get("strategyUnits", []) or []:
        advisor = str(item.get("advisor") or "未披露").strip() or "未披露"
        counts[advisor] += 1
    return "、".join(f"{advisor}（{count}）" for advisor, count in sorted(counts.items(), key=lambda row: (-row[1], row[0])))


def write_qd_fund_workbook(path: Path, data: dict[str, Any]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "QD基金汇总"
    header_fill = PatternFill("solid", fgColor="EAF3FB")
    header_font = Font(bold=True, color="162033")
    red_font = Font(color="C00000", bold=True)
    green_font = Font(color="008000", bold=True)
    gray_font = Font(color="475467")

    summary_columns = [
        ("主基金代码", "code"),
        ("基金名称", "name"),
        ("底层基金名称", "baseName"),
        ("所属公司", "company"),
        ("QD方向", "direction"),
        ("基金二级分类", "fundType"),
        ("合并份额代码", "shareCodes"),
        ("配置策略数", "strategyCount"),
        ("覆盖机构数", "advisorCount"),
        ("配置机构汇总", "advisorSummary"),
        ("配置总仓位比例(%)", "totalWeightRatio"),
        ("配置总仓位合计(%)", "totalWeight"),
        ("平均配置仓位(%)", "averageWeight"),
        ("最大单策略仓位(%)", "maxWeight"),
        ("近1周收益率(%)", "return1w"),
        ("近1月收益率(%)", "return1m"),
        ("近3月收益率(%)", "return3m"),
        ("近6月收益率(%)", "return6m"),
        ("近1年收益率(%)", "return1y"),
        ("今年以来收益率(%)", "returnYtd"),
        ("近1年最大回撤(%)", "maxDrawdown1y"),
        ("近1年年化波动率(%)", "volatility1y"),
        ("近1年夏普", "sharpeRatio1y"),
        ("最新净值日期", "latestNavDate"),
        ("基金申购状态", "purchaseApplyStatus"),
        ("个人单日限额", "personalDailyLimitAmountForWorkbook"),
        ("个人限额", "personalLimit"),
        ("个人限额来源", "personalLimitSource"),
        ("非个人单日限额", "nonPersonDailyLimitAmountForWorkbook"),
        ("非个人限额", "nonPersonLimit"),
        ("非个人限额来源", "nonPersonLimitSource"),
        ("限额状态", "limit"),
        ("限额公告日期", "limitSourceDate"),
        ("限额公告标题", "limitSourceTitle"),
        ("限额公告URL", "limitSourceUrl"),
        ("近1月净调入(百分点)", "recentNetInflow"),
        ("调入策略数", "recentAddStrategies"),
        ("调出策略数", "recentCutStrategies"),
        ("调整策略数", "recentChangedStrategies"),
        ("经济暴露质量", "exposureQuality"),
    ]

    ws.append([label for label, _key in summary_columns])
    for fund in data.get("funds", []):
        values = []
        for _label, key in summary_columns:
            value = fund.get(key)
            if key == "purchaseApplyStatus":
                value = qd_workbook_purchase_status(fund)
            elif key == "personalDailyLimitAmountForWorkbook":
                value = qd_workbook_daily_limit_amount(fund, "personalLimitAmount")
            elif key == "nonPersonDailyLimitAmountForWorkbook":
                value = qd_workbook_daily_limit_amount(fund, "nonPersonLimitAmount")
            elif key in {"personalLimitSource", "nonPersonLimitSource"}:
                value = qd_workbook_limit_source_text(value)
            elif key == "advisorSummary":
                value = qd_workbook_advisor_summary(fund)
            elif key in {"shareCodes"}:
                value = ",".join(value or [])
            values.append(value)
        ws.append(values)

    numeric_headers = {
        "配置总仓位比例(%)",
        "配置总仓位合计(%)",
        "平均配置仓位(%)",
        "最大单策略仓位(%)",
        "近1周收益率(%)",
        "近1月收益率(%)",
        "近3月收益率(%)",
        "近6月收益率(%)",
        "近1年收益率(%)",
        "今年以来收益率(%)",
        "近1年最大回撤(%)",
        "近1年年化波动率(%)",
        "近1年夏普",
        "近1月净调入(百分点)",
    }
    return_headers = {
        "近1周收益率(%)",
        "近1月收益率(%)",
        "近3月收益率(%)",
        "近6月收益率(%)",
        "近1年收益率(%)",
        "今年以来收益率(%)",
        "近1年最大回撤(%)",
        "近1月净调入(百分点)",
    }
    integer_headers = {"个人单日限额", "非个人单日限额"}

    def format_sheet(sheet) -> None:
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in sheet.columns:
            max_len = 0
            header = str(col[0].value or "")
            for cell in col:
                if cell.value is not None:
                    max_len = max(max_len, len(str(cell.value)))
                if cell.row > 1 and header in integer_headers and isinstance(cell.value, (int, float)):
                    cell.number_format = "0"
                if cell.row > 1 and header in numeric_headers and isinstance(cell.value, (int, float)):
                    cell.number_format = "0.00"
                if cell.row > 1 and header in return_headers and isinstance(cell.value, (int, float)):
                    cell.font = red_font if cell.value > 0 else green_font if cell.value < 0 else gray_font
            sheet.column_dimensions[get_column_letter(col[0].column)].width = min(max(max_len + 2, 10), 36)

    format_sheet(ws)

    detail = wb.create_sheet("策略配置明细")
    detail_columns = [
        "主基金代码",
        "基金名称",
        "所属公司",
        "QD方向",
        "基金二级分类",
        "策略名称",
        "投顾机构",
        "配置机构汇总",
        "策略类型",
        "持仓日期",
        "策略内仓位(%)",
        "合并份额代码",
        "合并份额名称",
        "近1月收益率(%)",
        "近3月收益率(%)",
        "近1年最大回撤(%)",
        "近1年年化波动率(%)",
        "近1年夏普",
        "基金申购状态",
        "个人单日限额",
        "非个人单日限额",
        "个人限额",
        "非个人限额",
    ]
    detail.append(detail_columns)
    for fund in data.get("funds", []):
        for item in fund.get("strategyUnits", []):
            detail.append(
                [
                    fund.get("code"),
                    fund.get("name"),
                    fund.get("company"),
                    fund.get("direction"),
                    fund.get("fundType"),
                    item.get("strategyName"),
                    item.get("advisor"),
                    qd_workbook_advisor_summary(fund),
                    item.get("unitType"),
                    item.get("holdingDate"),
                    item.get("weight"),
                    ",".join(item.get("shareCodes") or []),
                    "；".join(item.get("shareNames") or []),
                    fund.get("return1m"),
                    fund.get("return3m"),
                    fund.get("maxDrawdown1y"),
                    fund.get("volatility1y"),
                    fund.get("sharpeRatio1y"),
                    qd_workbook_purchase_status(fund),
                    qd_workbook_daily_limit_amount(fund, "personalLimitAmount"),
                    qd_workbook_daily_limit_amount(fund, "nonPersonLimitAmount"),
                    fund.get("personalLimit"),
                    fund.get("nonPersonLimit"),
                ]
            )
    format_sheet(detail)

    notes = wb.create_sheet("口径说明")
    notes.append(["字段", "口径说明"])
    for key in [
        "配置策略数",
        "配置机构汇总",
        "配置总仓位比例",
        "个人限额",
        "非个人限额",
        "基金申购状态",
        "个人单日限额",
        "非个人单日限额",
        "近1月收益",
        "近1年最大回撤",
        "近1年年化波动率",
        "近1年夏普",
        "最大限额",
        "保守限额容量",
    ]:
        notes.append([key, FIELD_DESCRIPTIONS.get(key, "")])
    notes.append(["生成时间", data.get("generatedAt")])
    notes.append(["数据来源", json.dumps(data.get("source") or {}, ensure_ascii=False)])
    format_sheet(notes)
    wb.save(path)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="生成 QD 基金配置一页报告和明细页。")
    parser.add_argument("--db-path", type=Path, default=DEFAULT_DB_PATH)
    parser.add_argument("--site-dir", type=Path, default=DEFAULT_SITE_DIR)
    parser.add_argument("--purchase-cache", type=Path, default=DEFAULT_PURCHASE_CACHE)
    parser.add_argument("--announcement-cache", type=Path, default=DEFAULT_LIMIT_ANNOUNCEMENT_CACHE)
    parser.add_argument("--skip-purchase-status", action="store_true", help="跳过天天基金申购状态/限额补充。")
    parser.add_argument("--skip-limit-announcements", action="store_true", help="跳过基金公告个人/非个人限额解析。")
    parser.add_argument("--refresh-limit-announcements", action="store_true", help="忽略旧缓存，重新抓取并解析 QD 基金限额公告。")
    parser.add_argument("--announcement-workers", type=int, default=3)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    site_dir = args.site_dir.resolve()
    conn = sqlite3.connect(args.db_path)
    conn.row_factory = sqlite3.Row
    all_rows = load_current_holding_rows(conn)
    units, total_weight = build_units(all_rows)
    qd_codes = {
        code
        for unit in units.values()
        for code, row in unit["fundMeta"].items()
        if is_qd_fund(row)
    }
    purchase_status, purchase_source = load_purchase_status(qd_codes, args.purchase_cache, args.skip_purchase_status)
    limit_announcements, announcement_source = load_limit_announcements(
        qd_codes,
        args.announcement_cache,
        skip=args.skip_limit_announcements,
        refresh=args.refresh_limit_announcements,
        workers=args.announcement_workers,
    )
    latest_rebalance_row = conn.execute(f"SELECT MAX({q('调仓日期')}) AS d FROM {q('策略调仓明细')}").fetchone()
    latest_rebalance = latest_rebalance_row["d"] if latest_rebalance_row else None
    recent_rebalance = load_recent_rebalance(conn, qd_codes, latest_rebalance)
    fund_returns = load_fund_returns(conn, qd_codes)
    data = aggregate(units, total_weight, recent_rebalance, purchase_status, limit_announcements, fund_returns)
    data["latestRebalanceDate"] = latest_rebalance
    data["source"] = {
        "db": str(args.db_path),
        "tables": ["策略当前持仓", "策略治理标签", "基金标准分类字典", "基金经济暴露快照", "策略调仓明细"],
        "purchaseStatus": purchase_source,
        "limitAnnouncements": announcement_source,
    }
    persist_limit_snapshot(conn, data)
    write_text(site_dir / "qd-fund-report.html", build_report_html(data))
    write_text(site_dir / "qd-fund-detail.html", build_detail_html(data))
    write_qd_fund_workbook(site_dir / "qd-fund-detail.xlsx", data)
    print(
        json.dumps(
            {
                "status": "ok",
                "siteDir": str(site_dir),
                "report": str(site_dir / "qd-fund-report.html"),
                "detail": str(site_dir / "qd-fund-detail.html"),
                "workbook": str(site_dir / "qd-fund-detail.xlsx"),
                "fundCount": data["summary"]["fundCount"],
                "strategyCount": data["summary"]["strategyCount"],
                "advisorCount": data["summary"]["advisorCount"],
                "latestRebalanceDate": latest_rebalance,
            },
            ensure_ascii=False,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
