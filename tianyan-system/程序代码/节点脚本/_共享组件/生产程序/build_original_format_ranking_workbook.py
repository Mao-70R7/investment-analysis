from __future__ import annotations

import json
import math
import shutil
from copy import copy
from pathlib import Path

import openpyxl


ROOT = next(parent for parent in Path(__file__).resolve().parents if (parent / "AGENTS.md").is_file())
DATA_PATH = ROOT / "outputs" / "fof_h1_strategy_ranking" / "20260630_recheck_20260702" / "fof_h1_strategy_ranking_data.json"
TEMPLATE_PATH = ROOT / "业务基线" / "策略排名.xlsx"
OUT_DIR = ROOT / "outputs" / "fof_strategy_business_report" / "20260703_final_business"
FORMAL_DIR = ROOT / "site" / "basic_data" / "reports"
OUT_NAME = "策略排名_同原表格式_2026H1.xlsx"


RANKING_HEADERS = [
    "代码",
    "策略名称",
    "策略基准",
    "累计收益(%)",
    "基准累计收益(%)",
    "年化收益(%)",
    "权益占比(%)",
    "固收占比(%)",
    "货币占比(%)",
    "另类占比(%)",
    "基金分类",
    "同类基金总数",
    "排名",
    "击败百分比",
    "排名位置百分位",
]

SUMMARY_HEADERS = [
    "基金分类",
    "策略数量",
    "对应公募基金类型",
    "同类基金总数",
    "策略平均YTD(%)",
    "策略平均排名百分位(%)",
    "策略中位数排名百分位(%)",
    "策略最高排名百分位(%)",
    "策略最低排名百分位(%)",
]

CATEGORY_PUBLIC_MAPPING = {
    "FOF-稳健型": "低波、固收、稳健配置型FOF",
    "FOF-均衡型": "平衡混合型FOF、均衡配置型FOF",
    "FOF-进取型": "偏股混合型FOF、高权益配置型FOF",
    "QDII-FOF": "海外、全球配置和QDII类FOF",
    "FOF-其他": "暂未形成有效同类样本的FOF",
}

CATEGORY_ORDER = {
    "FOF-稳健型": 1,
    "FOF-均衡型": 2,
    "FOF-进取型": 3,
    "QDII-FOF": 4,
    "FOF-其他": 99,
}


def clean_number(value, digits: int | None = 2):
    if value is None or value == "":
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return value
    if not math.isfinite(number):
        return None
    if digits is None:
        return number
    return round(number, digits)


def copy_cell_style(src, dst):
    if src.has_style:
        dst._style = copy(src._style)
    if src.number_format:
        dst.number_format = src.number_format
    if src.font:
        dst.font = copy(src.font)
    if src.fill:
        dst.fill = copy(src.fill)
    if src.border:
        dst.border = copy(src.border)
    if src.alignment:
        dst.alignment = copy(src.alignment)
    if src.protection:
        dst.protection = copy(src.protection)


def clone_row_style(ws, template_row: int, target_row: int):
    ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height
    for col_idx in range(1, ws.max_column + 1):
        copy_cell_style(ws.cell(template_row, col_idx), ws.cell(target_row, col_idx))


def capture_row_style(ws, row_idx: int, max_col: int):
    return {
        "height": ws.row_dimensions[row_idx].height,
        "cells": [
            {
                "style": copy(ws.cell(row_idx, col_idx)._style),
                "number_format": ws.cell(row_idx, col_idx).number_format,
                "font": copy(ws.cell(row_idx, col_idx).font),
                "fill": copy(ws.cell(row_idx, col_idx).fill),
                "border": copy(ws.cell(row_idx, col_idx).border),
                "alignment": copy(ws.cell(row_idx, col_idx).alignment),
                "protection": copy(ws.cell(row_idx, col_idx).protection),
            }
            for col_idx in range(1, max_col + 1)
        ],
    }


def apply_row_style(ws, row_style, target_row: int):
    ws.row_dimensions[target_row].height = row_style["height"]
    for col_idx, style in enumerate(row_style["cells"], 1):
        cell = ws.cell(target_row, col_idx)
        cell._style = copy(style["style"])
        cell.number_format = style["number_format"]
        cell.font = copy(style["font"])
        cell.fill = copy(style["fill"])
        cell.border = copy(style["border"])
        cell.alignment = copy(style["alignment"])
        cell.protection = copy(style["protection"])


def reset_sheet(ws, keep_rows: int, total_rows: int):
    if ws.max_row > keep_rows:
        ws.delete_rows(keep_rows + 1, ws.max_row - keep_rows)
    if total_rows > keep_rows:
        ws.insert_rows(keep_rows + 1, total_rows - keep_rows)


def build_ranking_rows(strategy_rows: list[dict]) -> list[list]:
    rows = []
    for row in strategy_rows:
        if row.get("策略H1收益率_百分比") is None or row.get("同类FOF样本数") in (None, "") or row.get("同类FOF排名") in (None, ""):
            continue
        alt_weight = clean_number(row.get("QDII权重_百分比"))
        peer_count = clean_number(row.get("同类FOF样本数"), 0)
        rank = clean_number(row.get("同类FOF排名"), 0)
        if peer_count is not None and rank is not None:
            peer_count = max(peer_count, rank)
        rows.append(
            [
                row.get("统一策略ID"),
                row.get("策略名称"),
                row.get("业绩基准"),
                clean_number(row.get("策略H1收益率_百分比")),
                clean_number(row.get("基准H1收益率_百分比")),
                clean_number(row.get("策略H1年化收益率_百分比")),
                clean_number(row.get("权益基金权重_百分比")),
                clean_number(row.get("债券基金权重_百分比")),
                clean_number(row.get("货币基金权重_百分比")),
                alt_weight,
                row.get("FOF可比分类"),
                peer_count,
                rank,
                None,
                None,
            ]
        )
    rows.sort(
        key=lambda r: (
            r[10] or "",
            1 if r[12] is None else 0,
            r[12] if r[12] is not None else 10**9,
            -999 if r[3] is None else -r[3],
            r[0] or "",
        )
    )
    return rows


def percentile_values(strategy_rows: list[dict], category: str) -> list[float]:
    values = []
    for row in strategy_rows:
        if row.get("FOF可比分类") != category:
            continue
        rank = clean_number(row.get("同类FOF排名"), None)
        peer_count = clean_number(row.get("同类FOF样本数"), None)
        if rank is None or peer_count is None:
            continue
        denominator = max(float(peer_count), float(rank))
        if denominator > 0:
            values.append(float(rank) / denominator * 100)
    return values


def build_summary_rows(category_rows: list[dict], strategy_rows: list[dict]) -> list[list]:
    rows = []
    for row in category_rows:
        category = row.get("FOF可比分类")
        if clean_number(row.get("策略数量"), 0) in (None, 0):
            continue
        values = percentile_values(strategy_rows, category)
        rows.append(
            [
                category,
                clean_number(row.get("策略数量"), 0),
                CATEGORY_PUBLIC_MAPPING.get(category, ""),
                clean_number(row.get("有收益FOF产品数") or row.get("FOF产品总数"), 0),
                clean_number(row.get("策略平均H1收益率_百分比")),
                clean_number((sum(values) / len(values)) if values else None, 1),
                clean_number(row.get("策略中位数排名百分位") * 100 if row.get("策略中位数排名百分位") is not None else None, 1),
                clean_number(max(values) if values else None, 1),
                clean_number(min(values) if values else None, 1),
            ]
        )
    rows.sort(key=lambda r: (CATEGORY_ORDER.get(r[0], 90), r[0] or ""))
    return rows


def fill_ranking_sheet(ws, rows: list[list]):
    total_rows = 1 + len(rows)
    row_style = capture_row_style(ws, 2, len(RANKING_HEADERS))
    reset_sheet(ws, 1, total_rows)
    for col_idx, header in enumerate(RANKING_HEADERS, 1):
        ws.cell(1, col_idx).value = header
    for row_idx in range(2, total_rows + 1):
        apply_row_style(ws, row_style, row_idx)
    for row_idx, values in enumerate(rows, 2):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx).value = value
        ws.cell(row_idx, 14).value = f"=1-O{row_idx}"
        ws.cell(row_idx, 15).value = f"=M{row_idx}/L{row_idx}"


def fill_summary_sheet(ws, rows: list[list]):
    total_rows = 1 + len(rows)
    row_style = capture_row_style(ws, 2, len(SUMMARY_HEADERS))
    reset_sheet(ws, 1, total_rows)
    for col_idx, header in enumerate(SUMMARY_HEADERS, 1):
        ws.cell(1, col_idx).value = header
    for row_idx in range(2, total_rows + 1):
        apply_row_style(ws, row_style, row_idx)
    for row_idx, values in enumerate(rows, 2):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx).value = value


def fill_note_sheet(ws, meta: dict):
    style_by_row = {idx: capture_row_style(ws, idx, ws.max_column) for idx in range(1, ws.max_row + 1)}
    notes = [
        ["分类规则和排名方法说明", None, None, None],
        [None, None, None, None],
        ["一、分类规则", None, None, None],
        ["按策略的资产配置特征映射到可比FOF分类，并在同类FOF产品中比较2026年上半年收益表现。", None, None, None],
        ["策略分类", "可比FOF范围", "说明", None],
        ["FOF-低波/现金管理", "货币型、短债型、低波稳健类FOF", "以现金管理、短债、低波稳健配置为主", None],
        ["FOF-固收型", "债券型FOF、固收为主FOF", "以固收资产为主，权益配置较低", None],
        ["FOF-均衡偏债", "偏债混合型FOF、稳健配置型FOF", "权益配置低于均衡型，兼顾收益与波动控制", None],
        ["FOF-均衡型", "平衡混合型FOF、均衡配置型FOF", "股债配置相对均衡", None],
        ["FOF-高权益型", "偏股混合型FOF、高权益配置型FOF", "权益资产占比较高，收益弹性和波动均较高", None],
        [None, None, None, None],
        ["二、排名方法", None, None, None],
        ["累计收益、基准累计收益和年化收益均取2026年上半年区间。", None, None, None],
        ["同类基金总数为该分类下有2026年上半年收益数据的FOF数量。", None, None, None],
        ["排名按同类FOF收益从高到低排序，数值越小表示排名越靠前。", None, None, None],
        ["击败百分比=1-排名位置百分位；排名位置百分位=排名/同类基金总数。", None, None, None],
        ["原模板未单列混合基金权重，因此占比列延续原表的权益、固收、货币和另类展示方式。", None, None, None],
        [None, None, None, None],
        ["三、数据范围", None, None, None],
        [f"策略收益截止日：{meta.get('策略收益截止锚点') or meta.get('统计截止日') or '2026-06-30'}", None, None, None],
        [f"FOF收益截止日：{meta.get('基金收益截止锚点') or meta.get('统计截止日') or '2026-06-30'}", None, None, None],
        [f"投顾策略数量：{meta.get('策略总数') or ''}", None, None, None],
        [f"有上半年收益FOF数量：{meta.get('有H1收益FOF数') or ''}", None, None, None],
    ]
    reset_sheet(ws, 0, len(notes))
    for row_idx in range(1, len(notes) + 1):
        apply_row_style(ws, style_by_row.get(row_idx) or style_by_row[1], row_idx)
    for row_idx, values in enumerate(notes, 1):
        for col_idx, value in enumerate(values, 1):
            ws.cell(row_idx, col_idx).value = value


def main():
    data = json.loads(DATA_PATH.read_text(encoding="utf-8"))
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    FORMAL_DIR.mkdir(parents=True, exist_ok=True)
    dev_path = OUT_DIR / OUT_NAME
    formal_path = FORMAL_DIR / OUT_NAME
    shutil.copy2(TEMPLATE_PATH, dev_path)
    wb = openpyxl.load_workbook(dev_path)
    ranking_rows = build_ranking_rows(data["strategyRows"])
    summary_rows = build_summary_rows(data["categoryRows"], data["strategyRows"])
    fill_ranking_sheet(wb["投顾组合排名"], ranking_rows)
    fill_summary_sheet(wb["分类汇总"], summary_rows)
    fill_note_sheet(wb["计算说明"], data.get("meta", {}))
    wb.calculation.fullCalcOnLoad = True
    wb.calculation.forceFullCalc = True
    wb.save(dev_path)
    shutil.copy2(dev_path, formal_path)
    print(json.dumps({"dev_path": str(dev_path), "formal_path": str(formal_path), "ranking_rows": len(ranking_rows), "summary_rows": len(summary_rows)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
