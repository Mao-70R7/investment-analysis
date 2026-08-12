from __future__ import annotations

import argparse
import hashlib
import sqlite3
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


PROJECT_ROOT = next(parent for parent in Path(__file__).resolve().parents if (parent / "AGENTS.md").is_file())
FORMAL_REPORT_ROOT = PROJECT_ROOT / "site" / "reports"
DEFAULT_DB = PROJECT_ROOT / "data" / "analysis_zh_current.sqlite"
DEFAULT_FUND_FILE = FORMAL_REPORT_ROOT / "副本全部基金0630.xlsx"
DEFAULT_STRATEGY_FILE = FORMAL_REPORT_ROOT / "投顾策略上半年业绩_20260701.xlsx"


def clean(value: Any) -> str:
    return "" if value is None else str(value).strip()


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for block in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def read_funds(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["全部基金"]
    rows: list[tuple[Any, ...]] = []
    for code, name, h1_return, benchmark, *_ in sheet.iter_rows(min_row=2, values_only=True):
        fund_code = clean(code).upper()
        if fund_code.endswith(".OF"):
            fund_code = fund_code[:-3]
        if not fund_code:
            continue
        rows.append((fund_code, clean(name), float(h1_return) if isinstance(h1_return, (int, float)) else None, clean(benchmark)))
    workbook.close()
    return rows


def read_strategies(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook["Sheet2"]
    rows: list[tuple[Any, ...]] = []
    for code, name, benchmark, cumulative_return, benchmark_return, annualized_return, *_ in sheet.iter_rows(min_row=3, values_only=True):
        strategy_code = clean(code).upper()
        if not strategy_code:
            continue
        rows.append(
            (
                strategy_code,
                clean(name),
                clean(benchmark),
                float(cumulative_return) * 100.0 if isinstance(cumulative_return, (int, float)) else None,
                float(benchmark_return) * 100.0 if isinstance(benchmark_return, (int, float)) else None,
                float(annualized_return) * 100.0 if isinstance(annualized_return, (int, float)) else None,
            )
        )
    workbook.close()
    return rows


def import_tables(db_path: Path, fund_path: Path, strategy_path: Path) -> dict[str, Any]:
    fund_rows = read_funds(fund_path)
    strategy_rows = read_strategies(strategy_path)
    generated_at = datetime.now(timezone.utc).astimezone().isoformat(timespec="seconds")
    conn = sqlite3.connect(db_path)
    try:
        conn.execute('DROP TABLE IF EXISTS "外部基金0630核对"')
        conn.execute(
            '''
            CREATE TABLE "外部基金0630核对" (
              "基金代码" TEXT PRIMARY KEY,
              "基金名称" TEXT,
              "上半年复权收益率_百分比" REAL,
              "业绩比较基准" TEXT,
              "截止日期" TEXT,
              "来源文件" TEXT,
              "来源SHA256" TEXT,
              "导入时间" TEXT
            )
            '''
        )
        conn.executemany(
            'INSERT INTO "外部基金0630核对" VALUES (?,?,?,?,?,?,?,?)',
            [(*row, "2026-06-30", str(fund_path), file_sha256(fund_path), generated_at) for row in fund_rows],
        )
        conn.execute('DROP TABLE IF EXISTS "外部广发策略0630核对"')
        conn.execute(
            '''
            CREATE TABLE "外部广发策略0630核对" (
              "策略代码" TEXT PRIMARY KEY,
              "策略名称" TEXT,
              "策略基准" TEXT,
              "上半年收益率_百分比" REAL,
              "基准上半年收益率_百分比" REAL,
              "年化收益率_百分比" REAL,
              "截止日期" TEXT,
              "来源文件" TEXT,
              "来源SHA256" TEXT,
              "导入时间" TEXT
            )
            '''
        )
        conn.executemany(
            'INSERT INTO "外部广发策略0630核对" VALUES (?,?,?,?,?,?,?,?,?,?)',
            [(*row, "2026-06-30", str(strategy_path), file_sha256(strategy_path), generated_at) for row in strategy_rows],
        )
        conn.commit()
    finally:
        conn.close()
    return {
        "fund_rows": len(fund_rows),
        "strategy_rows": len(strategy_rows),
        "fund_file": str(fund_path),
        "strategy_file": str(strategy_path),
        "generated_at": generated_at,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Import the two external 2026H1 audit workbooks into SQLite staging tables.")
    parser.add_argument("--db", type=Path, default=DEFAULT_DB)
    parser.add_argument("--fund-file", type=Path, default=DEFAULT_FUND_FILE)
    parser.add_argument("--strategy-file", type=Path, default=DEFAULT_STRATEGY_FILE)
    args = parser.parse_args()
    print(import_tables(args.db, args.fund_file, args.strategy_file))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
